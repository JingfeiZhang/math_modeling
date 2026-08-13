from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from copy import deepcopy
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

try:
    from src.workflow.prompt_policy import assemble_packet, format_receipt
except ModuleNotFoundError:  # Direct file execution from scripts/workflow.ps1.
    from prompt_policy import assemble_packet, format_receipt


QUESTION_RE = re.compile(r"(?:问题\s*[一二三四五六七八九十0-9]+|第\s*[一二三四五六七八九十0-9]+\s*问|\bQ\s*[1-9][0-9]*\b)", re.I)
CHINESE_NUMBERS = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}
QUESTION_V2_ROOT_FIELDS = {
    "schema_version", "problem_id", "question_id", "source_problem", "problem", "model_selection",
    "method", "assumptions", "risk_probes", "decisions", "evidence", "paper", "status",
}
QUESTION_V3_ROOT_FIELDS = QUESTION_V2_ROOT_FIELDS | {"literature"}
QUESTION_ARGUMENT_FIELDS = (
    "objective_interface", "model_choice", "formulation", "algorithm", "result", "validation", "conclusion",
)
SPRINT_ROOT_ONLY_PATHS = (
    "state/decision_log.json",
    "results/{problem}/claims.json",
    "paper/main.tex",
    "paper/figure_contracts.yaml",
    "output/",
)
SPRINT_STATUS_VALUES = {"SUCCESS", "PARTIAL", "FAILED", "INTERRUPTED"}
EXPERIMENT_LEVELS = ("scratch", "candidate", "formal", "paper-evidence")
NONFORMAL_LEVELS = {"scratch", "candidate"}
LIFECYCLE_RECEIPT_NAMES = {
    "quickcheck": "probe_receipt.json",
    "checkpoint": "candidate_receipt.json",
    "archive-work": "archive_receipt.json",
}
FORMAL_FIGURE_FIELDS = (
    "core_message", "visual_hierarchy", "target_size_profile", "statistics_report",
    "data_integrity", "label_strategy", "rasterized_layers", "palette_id", "color_encoding",
)
FIGURE_SIZE_WIDTHS = {"contest-body": 158.0, "journal-single": 89.0, "journal-double": 183.0}
VISUAL_DESIGN_STATES = (
    "DATA_READY", "INTENT_READY", "BRIEF_READY", "DESIGN_APPROVED",
    "RENDERED", "QA_PASSED", "CONTRACT_READY", "STALE",
)
LITERATURE_STATES = (
    "NOT_STARTED", "PLAN_READY", "DISCOVERED", "SOURCES_VERIFIED", "CARDS_READY",
    "SYNTHESIS_READY", "CITATION_READY", "STALE",
)
LITERATURE_REVIEW_DEPTHS = {"METADATA_ONLY", "ABSTRACT_SCREENED", "TARGETED_READ", "DEEP_READ"}
LITERATURE_ALLOWED_PUBLICATION_TYPES = {"journal", "conference", "preprint", "thesis"}
LITERATURE_PUBLICATION_TYPES = {"journal-article", "conference-paper", "preprint", "thesis"}
LITERATURE_PROVIDERS = {
    "crossref", "openalex", "arxiv", "semantic-scholar", "manual-google-scholar",
    "manual-cnki", "manual-wanfang", "user-supplied",
}
LITERATURE_MANAGED_BIB_BEGIN = "% BEGIN LITERATURE-GUIDED-MODELING"
LITERATURE_MANAGED_BIB_END = "% END LITERATURE-GUIDED-MODELING"


class ReopenRequiredError(ValueError):
    """Signal that one question must return to its local G3/G4 cycle."""
FIGURE_BRIEF_CONTRACT_FIELDS = (
    "claim_id", "core_conclusion", "core_message", "evidence_chain", "backend", "source_data",
    "source_script", "outputs", "baseline", "axes", "caption", "panel_map", "statistics",
    "statistics_report", "data_integrity", "label_strategy", "rasterized_layers", "review_risks",
    "target_size_profile", "final_width_mm", "min_font_pt", "palette_id", "color_encoding",
    "visual_hierarchy",
)
Q1_SPRINT_OUTPUTS = {
    "forecast-q1": (
        "handoff.json",
        "run_forecast.py",
        "run_manifest.json",
        "metrics_summary.json",
        "metrics_by_series.csv",
        "baseline_selection.csv",
        "validation_predictions.csv",
        "blind_test_predictions.csv",
        "workload_panel.csv",
        "claim_proposal.json",
        "hash_manifest.json",
    ),
    "scheduling-q1": (
        "handoff.json",
        "run_scheduling_q1.py",
        "run_manifest.json",
        "scheduling_metrics.json",
        "baseline_schedule.csv",
        "optimized_schedule.csv",
        "resource_audit.csv",
        "carry_in.json",
        "constraint_audit.json",
        "fallback_test.json",
        "claim_proposal.json",
        "hash_manifest.json",
    ),
    "writer-q1": (
        "handoff.json",
        "q1_draft.tex",
        "figure_contract_proposal.yaml",
        "figure_manifest.json",
    ),
    "reviewer-q1": ("handoff.json", "q1_review.json"),
}


def load_yaml(path: Path) -> dict[str, Any]:
    import yaml

    value = yaml.safe_load(path.read_text(encoding="utf-8"))
    return value if isinstance(value, dict) else {}


def dump_yaml(path: Path, value: dict[str, Any]) -> None:
    import yaml

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(value, allow_unicode=True, sort_keys=False), encoding="utf-8")


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")


def dump_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2, default=_json_default) + "\n", encoding="utf-8")


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def canonical_sha256(value: Any) -> str:
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=_json_default).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def question_interface_sha256(payload: dict[str, Any]) -> str:
    """Hash only the P1 problem interface, excluding downstream evidence and handoffs."""

    normalized = {
        key: deepcopy(payload.get(key))
        for key in ("schema_version", "problem_id", "question_id", "source_problem", "problem")
    }
    return canonical_sha256(normalized)


def _deep_merge(base: dict[str, Any], override: dict[str, Any]) -> dict[str, Any]:
    merged = deepcopy(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = _deep_merge(merged[key], value)
        else:
            merged[key] = deepcopy(value)
    return merged


def safe_token(value: str, label: str) -> str:
    if not re.fullmatch(r"[A-Za-z0-9._-]+", value):
        raise ValueError(f"{label} may contain only letters, digits, dot, underscore, and hyphen")
    return value


def workspace_path(root: Path, value: str, label: str) -> Path:
    candidate = Path(value)
    if candidate.is_absolute():
        raise ValueError(f"absolute {label} path is forbidden: {value}")
    resolved = (root / candidate).resolve()
    try:
        resolved.relative_to(root.resolve())
    except ValueError as exc:
        raise ValueError(f"{label} path escapes workspace: {value}") from exc
    return resolved


def relative_path(root: Path, path: Path) -> str:
    return path.resolve().relative_to(root.resolve()).as_posix()


def _project_requires_visual_handoff(root: Path) -> bool:
    project_path = root / "project.yaml"
    if project_path.is_file():
        project = load_yaml(project_path)
        if int(project.get("workflow_contract_version", 0) or 0) >= 5:
            return True
    local_workflow = root / "config" / "workflow.yaml"
    if local_workflow.is_file():
        workflow = load_yaml(local_workflow)
        visualization = workflow.get("visualization_design") if isinstance(workflow.get("visualization_design"), dict) else {}
        return visualization.get("strict_handoff") is True
    return False


def _project_requires_literature_handoff(root: Path) -> bool:
    project_path = root / "project.yaml"
    if not project_path.is_file():
        return False
    project = load_yaml(project_path)
    if int(project.get("workflow_contract_version", 0) or 0) >= 6:
        return True
    local_workflow = root / "config" / "workflow.yaml"
    if not local_workflow.is_file():
        return False
    workflow = load_yaml(local_workflow)
    literature = workflow.get("literature_guided_modeling") if isinstance(workflow.get("literature_guided_modeling"), dict) else {}
    return literature.get("strict_g5") is True and int(project.get("workflow_contract_version", 0) or 0) >= 6


def _rewrite_visual_path(value: Any, old_prefix: str | None, new_prefix: str) -> Any:
    if not isinstance(value, str) or not old_prefix:
        return value
    if value == old_prefix:
        return new_prefix
    if value.startswith(old_prefix + "/"):
        return new_prefix + value[len(old_prefix):]
    if ":" in value:
        path_value, selector = value.split(":", 1)
        if path_value == old_prefix or path_value.startswith(old_prefix + "/"):
            rewritten = new_prefix + path_value[len(old_prefix):]
            return rewritten + ":" + selector
    return value


def _sync_visual_handoff_after_lifecycle_transition(
    root: Path,
    manifest_path: Path,
    manifest: dict[str, Any],
    old_prefix: str | None = None,
) -> None:
    run_root = manifest_path.parent
    new_prefix = relative_path(root, run_root)
    level = str(manifest.get("run_mode") or manifest.get("level") or run_root.parent.name)
    eligible = _visual_evidence_eligible(level, manifest, run_root)
    data_path = run_root / "figure_data_manifest.yaml"
    if not data_path.is_file():
        return

    data = load_yaml(data_path)
    for item in data.get("source_artifacts", []):
        if isinstance(item, dict):
            item["path"] = _rewrite_visual_path(item.get("path"), old_prefix, new_prefix)
    data.update({
        "level": level,
        "source_run_manifest": relative_path(root, manifest_path),
        "source_run_manifest_sha256": sha256(manifest_path),
        "contest_evidence_eligible": eligible,
    })
    data_current = True
    for item in data.get("source_artifacts", []):
        if not isinstance(item, dict):
            data_current = False
            continue
        try:
            source = workspace_path(root, str(item.get("path", "")), "figure-data source")
        except ValueError:
            data_current = False
            continue
        if not source.is_file() or item.get("sha256") != sha256(source):
            data_current = False
    data["status"] = "DATA_READY" if data_current else "STALE"
    dump_yaml(data_path, data)

    intent_path = run_root / "visual_intent.yaml"
    intent: dict[str, Any] | None = None
    if intent_path.is_file():
        intent = load_yaml(intent_path)
        intent.update({
            "source_data_manifest": relative_path(root, data_path),
            "source_data_manifest_sha256": sha256(data_path),
            "contest_evidence_eligible": eligible,
        })
        if not data_current:
            intent["status"] = "STALE"
        dump_yaml(intent_path, intent)

    briefs_root = run_root / "figure_briefs"
    for brief_path in sorted(briefs_root.glob("*.yaml")) if briefs_root.is_dir() else []:
        brief = load_yaml(brief_path)
        for field in ("source_data",):
            if isinstance(brief.get(field), list):
                brief[field] = [_rewrite_visual_path(item, old_prefix, new_prefix) for item in brief[field]]
        brief["source_script"] = _rewrite_visual_path(brief.get("source_script"), old_prefix, new_prefix)
        if isinstance(brief.get("render_command"), list):
            brief["render_command"] = [
                _rewrite_visual_path(item, old_prefix, new_prefix) for item in brief["render_command"]
            ]
        try:
            source_script = workspace_path(root, str(brief.get("source_script", "")), "figure source script")
        except ValueError:
            source_script = root / "__invalid_figure_source_script__"
        script_current = source_script.is_file() and brief.get("source_script_sha256") == sha256(source_script)
        outputs = brief.get("outputs") if isinstance(brief.get("outputs"), dict) else {}
        for key, value in list(outputs.items()):
            outputs[key] = _rewrite_visual_path(value, old_prefix, new_prefix)
        for evidence in brief.get("evidence_chain", []):
            if isinstance(evidence, dict):
                evidence["locator"] = _rewrite_visual_path(evidence.get("locator"), old_prefix, new_prefix)
        integrity = brief.get("data_integrity") if isinstance(brief.get("data_integrity"), dict) else {}
        for source_hash in integrity.get("source_hashes", []):
            if isinstance(source_hash, dict):
                source_hash["path"] = _rewrite_visual_path(source_hash.get("path"), old_prefix, new_prefix)
        brief.update({
            "source_data_manifest": relative_path(root, data_path),
            "source_data_manifest_sha256": sha256(data_path),
            "contest_evidence_eligible": eligible,
        })
        if intent_path.is_file():
            brief["visual_intent"] = relative_path(root, intent_path)
            brief["visual_intent_sha256"] = sha256(intent_path)
        if not data_current or not script_current or (intent is not None and intent.get("status") == "STALE"):
            brief["status"] = "STALE"
        dump_yaml(brief_path, brief)


def shared_asset(root: Path, workspace_root: Path | None, relative: str) -> Path:
    """Resolve a project-local asset first, then a shared workbench asset."""

    candidates = [root / relative]
    if workspace_root is not None and workspace_root.resolve() != root.resolve():
        candidates.append(workspace_root / relative)
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    return candidates[0]


def path_is_within(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
    except ValueError:
        return False
    return True


def protected_sprint_target(root: Path, problem: str, value: str) -> bool:
    target = workspace_path(root, value, "proposed target")
    protected = [item.format(problem=problem) for item in SPRINT_ROOT_ONLY_PATHS]
    for item in protected:
        protected_path = workspace_path(root, item.rstrip("/"), "root-only")
        if item.endswith("/") and path_is_within(target, protected_path):
            return True
        if target == protected_path:
            return True
    return False


def locator_path(root: Path, locator: str) -> Path:
    raw = locator.split(":", 1)[0].strip()
    candidate = Path(raw)
    if candidate.is_absolute():
        raise ValueError(f"absolute evidence path is forbidden: {locator}")
    resolved = (root / candidate).resolve()
    try:
        resolved.relative_to(root.resolve())
    except ValueError as exc:
        raise ValueError(f"evidence path escapes workspace: {locator}") from exc
    return resolved


def locator_value(root: Path, locator: str) -> Any:
    parts = locator.split(":", 1)
    path = locator_path(root, locator)
    if not path.is_file() or len(parts) != 2 or path.suffix.lower() != ".json":
        return None
    value: Any = load_json(path)
    selector = parts[1].strip().removeprefix("$").strip(".")
    if not selector:
        return None
    for token in selector.split("."):
        if isinstance(value, dict) and token in value:
            value = value[token]
        elif isinstance(value, list) and token.isdigit() and int(token) < len(value):
            value = value[int(token)]
        else:
            return None
    return value if isinstance(value, (str, int, float, bool)) else None


def extract_problem_text(path: Path) -> str:
    if path.suffix.lower() == ".pdf":
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("pypdf is required to initialize from a PDF problem file") from exc
        return "\n".join(page.extract_text() or "" for page in PdfReader(str(path)).pages)
    return path.read_text(encoding="utf-8", errors="replace")


def question_number(token: str) -> int | None:
    digit = re.search(r"[0-9]+", token)
    if digit:
        return int(digit.group())
    for char, number in CHINESE_NUMBERS.items():
        if char in token:
            return number
    return None


def extract_questions(text: str) -> list[tuple[str, str]]:
    matches = list(QUESTION_RE.finditer(text))
    found: dict[int, str] = {}
    for index, match in enumerate(matches):
        number = question_number(match.group())
        if number is None or number in found:
            continue
        end = matches[index + 1].start() if index + 1 < len(matches) else min(len(text), match.end() + 800)
        target = re.sub(r"\s+", " ", text[match.start():end]).strip()
        found[number] = target[:800]
    return [(f"Q{number}", found[number]) for number in sorted(found)]


def contest_config(root: Path) -> dict[str, Any]:
    path = root / "contest.yaml"
    if not path.is_file():
        raise FileNotFoundError(f"contest configuration is missing: {path}")
    return load_yaml(path)


def question_paths(root: Path, problem: str) -> list[Path]:
    folder = root / "problems" / problem / "questions"
    return sorted(folder.glob("Q*/question.yaml")) if folder.is_dir() else []


def _mapping_issues(value: Any, label: str, required: set[str]) -> tuple[dict[str, Any], list[str]]:
    if not isinstance(value, dict):
        return {}, [f"{label} must be an object"]
    missing = sorted(required - set(value))
    unexpected = sorted(set(value) - required)
    issues = [f"{label} missing {missing}"] if missing else []
    if unexpected:
        issues.append(f"{label} has unexpected fields {unexpected}")
    return value, issues


def _string_list_issues(value: Any, label: str) -> list[str]:
    if not isinstance(value, list) or any(not isinstance(item, str) for item in value):
        return [f"{label} must be an array of strings"]
    return []


def question_v2_shape_issues(payload: dict[str, Any]) -> list[str]:
    """Validate the strict v2 handoff shape without adding a runtime dependency."""

    issues: list[str] = []
    if payload.get("schema_version") != 2:
        return ["schema_version must be 2"]
    missing = sorted(QUESTION_V2_ROOT_FIELDS - set(payload))
    unexpected = sorted(set(payload) - QUESTION_V2_ROOT_FIELDS)
    if missing:
        issues.append(f"root missing {missing}")
    if unexpected:
        issues.append(f"root has unexpected fields {unexpected}")
    if not isinstance(payload.get("problem_id"), str) or not str(payload.get("problem_id", "")):
        issues.append("problem_id must be a non-empty string")
    if not re.fullmatch(r"Q[1-9][0-9]*", str(payload.get("question_id", ""))):
        issues.append("question_id must match Q<number>")
    if not isinstance(payload.get("source_problem"), str):
        issues.append("source_problem must be a string")

    problem_fields = {"target", "type", "inputs", "outputs", "constraints", "evaluation_metrics", "dependencies", "key_conflicts"}
    problem, nested = _mapping_issues(payload.get("problem"), "problem", problem_fields)
    issues.extend(nested)
    for field in ("target", "type"):
        if field in problem and not isinstance(problem[field], str):
            issues.append(f"problem.{field} must be a string")
    for field in problem_fields - {"target", "type"}:
        if field in problem:
            issues.extend(_string_list_issues(problem[field], f"problem.{field}"))
    dependencies = problem.get("dependencies", [])
    if isinstance(dependencies, list) and any(not re.fullmatch(r"Q[1-9][0-9]*", item) for item in dependencies if isinstance(item, str)):
        issues.append("problem.dependencies entries must match Q<number>")

    selection_fields = {"primary", "rationale", "baseline", "rejected_alternatives"}
    selection, nested = _mapping_issues(payload.get("model_selection"), "model_selection", selection_fields)
    issues.extend(nested)
    for field in ("primary", "rationale", "baseline"):
        if field in selection and not isinstance(selection[field], str):
            issues.append(f"model_selection.{field} must be a string")
    if "rejected_alternatives" in selection:
        issues.extend(_string_list_issues(selection["rejected_alternatives"], "model_selection.rejected_alternatives"))

    method, nested = _mapping_issues(payload.get("method"), "method", {"main", "baseline", "fallback"})
    issues.extend(nested)
    main, nested = _mapping_issues(method.get("main"), "method.main", {"name", "rationale", "implementation"})
    issues.extend(nested)
    baseline, nested = _mapping_issues(method.get("baseline"), "method.baseline", {"name", "implementation", "comparable_output"})
    issues.extend(nested)
    for label, item, fields in (
        ("method.main", main, ("name", "rationale", "implementation")),
        ("method.baseline", baseline, ("name", "implementation")),
    ):
        for field in fields:
            if field in item and not isinstance(item[field], str):
                issues.append(f"{label}.{field} must be a string")
    if "comparable_output" in baseline and not isinstance(baseline["comparable_output"], bool):
        issues.append("method.baseline.comparable_output must be boolean")
    fallback = method.get("fallback")
    if fallback is not None:
        fallback_value, nested = _mapping_issues(fallback, "method.fallback", {"name", "trigger"})
        issues.extend(nested)
        if any(not isinstance(fallback_value.get(field), str) for field in ("name", "trigger")):
            issues.append("method.fallback name and trigger must be strings")

    for field in ("assumptions", "risk_probes", "decisions"):
        value = payload.get(field)
        if not isinstance(value, list) or any(not isinstance(item, dict) for item in value):
            issues.append(f"{field} must be an array of objects")

    evidence_fields = {"runs", "paper_evidence_runs", "robustness", "figures", "result_claim_ids", "validation_claim_ids", "boundary_claim_ids"}
    evidence, nested = _mapping_issues(payload.get("evidence"), "evidence", evidence_fields)
    issues.extend(nested)
    for field in evidence_fields - {"robustness"}:
        if field in evidence:
            issues.extend(_string_list_issues(evidence[field], f"evidence.{field}"))
    if "robustness" in evidence and not isinstance(evidence["robustness"], str):
        issues.append("evidence.robustness must be a string")

    paper_fields = {"section", "table_ids", "figure_ids", "code_refs", "downstream_interfaces", "argument_contract"}
    paper, nested = _mapping_issues(payload.get("paper"), "paper", paper_fields)
    issues.extend(nested)
    if "section" in paper and not isinstance(paper["section"], str):
        issues.append("paper.section must be a string")
    for field in paper_fields - {"section", "argument_contract"}:
        if field in paper:
            issues.extend(_string_list_issues(paper[field], f"paper.{field}"))
    argument, nested = _mapping_issues(payload.get("paper", {}).get("argument_contract") if isinstance(payload.get("paper"), dict) else None, "paper.argument_contract", set(QUESTION_ARGUMENT_FIELDS))
    issues.extend(nested)
    for field in QUESTION_ARGUMENT_FIELDS:
        if argument.get(field) not in {"pending", "complete"}:
            issues.append(f"paper.argument_contract.{field} must be pending or complete")
    if payload.get("status") not in {"DRAFT", "SCREENED", "HUMAN_DECIDED", "RUN", "FROZEN", "PAPER_READY"}:
        issues.append("status is invalid")
    return issues


def _literature_ref_issues(value: Any, label: str, allow_null: bool = False) -> list[str]:
    if value is None and allow_null:
        return []
    if not isinstance(value, dict) or set(value) != {"path", "sha256"}:
        return [f"{label} must contain only path and sha256"]
    issues: list[str] = []
    path_value = value.get("path")
    hash_value = value.get("sha256")
    if not isinstance(path_value, str) or not path_value.strip() or Path(path_value).is_absolute():
        issues.append(f"{label}.path must be a non-empty relative path")
    if not isinstance(hash_value, str) or (hash_value and not re.fullmatch(r"[0-9a-f]{64}", hash_value)):
        issues.append(f"{label}.sha256 must be empty or a lowercase SHA-256")
    return issues


def question_v3_shape_issues(payload: dict[str, Any]) -> list[str]:
    """Validate the V3 question handoff while preserving the V2 behavioral contract."""

    if payload.get("schema_version") != 3:
        return ["schema_version must be 3"]
    missing = sorted(QUESTION_V3_ROOT_FIELDS - set(payload))
    unexpected = sorted(set(payload) - QUESTION_V3_ROOT_FIELDS)
    issues: list[str] = []
    if missing:
        issues.append(f"root missing {missing}")
    if unexpected:
        issues.append(f"root has unexpected fields {unexpected}")
    legacy = deepcopy(payload)
    legacy.pop("literature", None)
    legacy["schema_version"] = 2
    issues.extend(question_v2_shape_issues(legacy))

    expected = {"search_plan", "search_receipts", "evidence_cards", "model_evidence_brief", "bib_keys", "status"}
    literature, nested = _mapping_issues(payload.get("literature"), "literature", expected)
    issues.extend(nested)
    issues.extend(_literature_ref_issues(literature.get("search_plan"), "literature.search_plan", allow_null=True))
    issues.extend(_literature_ref_issues(literature.get("model_evidence_brief"), "literature.model_evidence_brief", allow_null=True))
    for field in ("search_receipts", "evidence_cards"):
        values = literature.get(field)
        if not isinstance(values, list):
            issues.append(f"literature.{field} must be an array")
            continue
        for index, item in enumerate(values):
            issues.extend(_literature_ref_issues(item, f"literature.{field}[{index}]"))
    issues.extend(_string_list_issues(literature.get("bib_keys"), "literature.bib_keys"))
    if literature.get("status") not in LITERATURE_STATES:
        issues.append("literature.status is invalid")
    return issues


def _literature_schema_path(root: Path, name: str) -> Path:
    local = root / "config" / "schemas" / name
    if local.is_file():
        return local
    return Path(__file__).resolve().parents[2] / "config" / "schemas" / name


def _validate_literature_payload(root: Path, schema_name: str, payload: dict[str, Any]) -> None:
    try:
        from jsonschema import Draft202012Validator
    except ImportError as exc:
        raise RuntimeError("jsonschema is required for literature workflow validation") from exc
    schema_path = _literature_schema_path(root, schema_name)
    if not schema_path.is_file():
        raise FileNotFoundError(f"literature schema is missing: {schema_path}")
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    errors = sorted(Draft202012Validator(schema).iter_errors(payload), key=lambda item: list(item.path))
    if errors:
        details = []
        for error in errors[:8]:
            location = ".".join(str(item) for item in error.path) or "root"
            details.append(f"{location}: {error.message}")
        raise ValueError(f"{schema_name} validation failed: " + "; ".join(details))


def _question_manifest(root: Path, problem: str, question: str) -> tuple[Path, dict[str, Any]]:
    problem = safe_token(problem, "problem")
    question = safe_token(question, "question")
    if not re.fullmatch(r"Q[1-9][0-9]*", question):
        raise ValueError("question must use Q<number> form")
    path = root / "problems" / problem / "questions" / question / "question.yaml"
    if not path.is_file():
        raise FileNotFoundError(f"question manifest is missing: {path}")
    payload = load_yaml(path)
    if payload.get("problem_id") != problem or payload.get("question_id") != question:
        raise ValueError("question manifest identity does not match the selected problem/question")
    return path, payload


def _literature_root(root: Path, problem: str, question: str) -> Path:
    return root / "problems" / problem / "questions" / question / "literature"


def _literature_cache_root(root: Path, problem: str, question: str) -> Path:
    return root / "work" / "cache" / "literature" / problem / question


def _hashed_ref(root: Path, path: Path) -> dict[str, str]:
    if not path.is_file() or not path_is_within(path, root):
        raise ValueError(f"literature reference must be a project-local file: {path}")
    return {"path": relative_path(root, path), "sha256": sha256(path)}


def _current_hashed_ref(root: Path, value: Any, label: str) -> tuple[Path | None, list[str]]:
    issues = _literature_ref_issues(value, label)
    if issues:
        return None, issues
    try:
        path = workspace_path(root, str(value["path"]), label)
    except (KeyError, ValueError) as exc:
        return None, [str(exc)]
    if not path.is_file():
        return path, [f"{label} file is missing: {value.get('path')}"]
    if sha256(path) != value.get("sha256"):
        return path, [f"{label} SHA-256 drifted: {value.get('path')}"]
    return path, []


def _update_question_literature(root: Path, question_path: Path, **updates: Any) -> dict[str, Any]:
    payload = load_yaml(question_path)
    if payload.get("schema_version") != 3:
        raise ValueError("literature handoff updates require question schema v3")
    literature = payload.setdefault("literature", {})
    literature.update(deepcopy(updates))
    dump_yaml(question_path, payload)
    return payload


def normalize_doi(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip().lower()
    text = re.sub(r"^(?:https?://(?:dx\.)?doi\.org/|doi\s*:\s*)", "", text, flags=re.I)
    text = text.rstrip(".,;)]}")
    match = re.fullmatch(r"10\.[0-9]{4,9}/\S+", text)
    return text if match else None


_TITLE_STOPWORDS = {"a", "an", "the", "in", "of", "for", "on", "to", "and", "with", "by", "et", "al"}


def _title_tokens(value: Any) -> set[str]:
    tokens = re.findall(r"[\w\u4e00-\u9fff]+", str(value or "").lower(), flags=re.UNICODE)
    return {token for token in tokens if token not in _TITLE_STOPWORDS}


def _first_author_key(authors: Any) -> str:
    if not isinstance(authors, list) or not authors:
        return ""
    first = str(authors[0]).strip().lower()
    if "," in first:
        first = first.split(",", 1)[0]
    else:
        parts = first.split()
        first = parts[-1] if parts else ""
    return re.sub(r"[^\w\u4e00-\u9fff]", "", first)


def _records_match(left: dict[str, Any], right: dict[str, Any]) -> bool:
    left_doi = normalize_doi(left.get("doi"))
    right_doi = normalize_doi(right.get("doi"))
    if left_doi and right_doi:
        return left_doi == right_doi
    if left.get("year") != right.get("year") or _first_author_key(left.get("authors")) != _first_author_key(right.get("authors")):
        return False
    left_tokens = _title_tokens(left.get("title"))
    right_tokens = _title_tokens(right.get("title"))
    union = left_tokens | right_tokens
    return bool(union) and len(left_tokens & right_tokens) / len(union) >= 0.90


def _record_completeness(record: dict[str, Any]) -> tuple[int, int, int]:
    fields = ("doi", "url", "abstract", "venue", "volume", "pages", "arxiv_id")
    complete = sum(record.get(field) not in (None, "", []) for field in fields)
    publisher = 0 if record.get("publication_type") == "preprint" else 1
    citations = int(record.get("citation_count", 0) or 0)
    return complete, publisher, citations


def deduplicate_literature_records(records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    unique: list[dict[str, Any]] = []
    merged: list[dict[str, Any]] = []
    for record in records:
        match_index = next((index for index, item in enumerate(unique) if _records_match(item, record)), None)
        if match_index is None:
            unique.append(record)
            continue
        existing = unique[match_index]
        keep, drop = (record, existing) if _record_completeness(record) > _record_completeness(existing) else (existing, record)
        unique[match_index] = keep
        kept_id = str(keep.get("canonical_id") or keep.get("doi") or keep.get("title"))
        dropped_id = str(drop.get("canonical_id") or drop.get("doi") or drop.get("title"))
        bucket = next((item for item in merged if item["kept"] == kept_id), None)
        if bucket is None:
            bucket = {"kept": kept_id, "merged": []}
            merged.append(bucket)
        if dropped_id not in bucket["merged"]:
            bucket["merged"].append(dropped_id)
    return unique, merged


def _academic_source_record(record: dict[str, Any]) -> bool:
    title = str(record.get("title", ""))
    venue = str(record.get("venue", ""))
    combined = f"{title} {venue}".lower()
    contest_markers = ("数学建模竞赛", "mathematical modeling contest", "mcm/icm", "cumcm", "华数杯")
    return record.get("publication_type") in LITERATURE_PUBLICATION_TYPES and not any(marker in combined for marker in contest_markers)


def _is_literature_evidence_path(root: Path, path: Path) -> bool:
    try:
        relative = relative_path(root, path).lower()
    except ValueError:
        return True
    return "/literature/" in f"/{relative}/" or relative.startswith("work/cache/literature/")


def write_paper_blueprint(root: Path, problem: str, source_problem: str, questions: list[tuple[str, str]]) -> tuple[Path, Path]:
    """Write deterministic derived paper structure files; neither file owns workflow state."""

    generated = root / "paper" / "generated"
    generated.mkdir(parents=True, exist_ok=True)
    structure_path = generated / "question_structure.tex"
    structure_lines = [
        "% Generated from the real problem markers by competition_workflow.py.",
        "% Do not edit; update the question manifests and regenerate instead.",
        f"\\providecommand{{\\MathModelQuestionCount}}{{{len(questions)}}}",
    ]
    blueprint_questions: list[dict[str, Any]] = []
    for question_id, target in questions:
        number = int(question_id[1:])
        section = f"sections/question_{number}.tex"
        structure_lines.append(f"\\input{{{section}}}")
        manifest_path = root / "problems" / problem / "questions" / question_id / "question.yaml"
        payload = load_yaml(manifest_path) if manifest_path.is_file() else {}
        paper = payload.get("paper", {}) if isinstance(payload.get("paper"), dict) else {}
        problem_data = payload.get("problem", {}) if isinstance(payload.get("problem"), dict) else {}
        blueprint_questions.append({
            "question_id": question_id,
            "manifest": relative_path(root, manifest_path),
            "section": str(paper.get("section") or section),
            "target": str(problem_data.get("target") or target),
            "inputs": list(problem_data.get("inputs", [])) if isinstance(problem_data.get("inputs"), list) else [],
            "outputs": list(problem_data.get("outputs", [])) if isinstance(problem_data.get("outputs"), list) else [],
            "dependencies": list(problem_data.get("dependencies", [])) if isinstance(problem_data.get("dependencies"), list) else [],
            "argument_contract": dict(paper.get("argument_contract", {})) if isinstance(paper.get("argument_contract"), dict) else {},
        })
    structure_path.write_text("\n".join(structure_lines) + "\n", encoding="utf-8")

    blueprint_path = generated / "paper_blueprint.yaml"
    dump_yaml(blueprint_path, {
        "schema_version": 1,
        "contract": "CUMCM-paper-structure-v3",
        "derived": True,
        "problem_id": problem,
        "source_problem": source_problem,
        "question_count": len(questions),
        "body_order": [
            "abstract", "problem_restatement", "problem_analysis", "assumptions_notation",
            "data_processing", "questions", "model_evaluation", "ai_statement_if_required",
            "references", "appendix",
        ],
        "questions": blueprint_questions,
    })
    return structure_path, blueprint_path


def decision_template(root: Path, workspace_root: Path | None = None) -> dict[str, Any]:
    candidates = [
        root / "skill_staging" / "handsomeZR-mathmodel-skill" / "templates" / "shared" / "decision_log.json",
        *(([workspace_root / "skill_staging" / "handsomeZR-mathmodel-skill" / "templates" / "shared" / "decision_log.json"]) if workspace_root else []),
        Path.home() / ".codex" / "skills" / "mathmodel-skill" / "templates" / "shared" / "decision_log.json",
    ]
    for path in candidates:
        if path.is_file():
            return load_json(path)
    raise FileNotFoundError("mathmodel-skill decision_log.json template was not found")


def initialize(root: Path, problem: str, problem_file: Path, workspace_root: Path | None = None) -> dict[str, Any]:
    problem = safe_token(problem, "problem")
    if problem.upper() == "TBD":
        raise ValueError("formal initialization is forbidden while the problem is TBD")
    problem_file = problem_file.resolve()
    if not problem_file.is_file():
        raise FileNotFoundError(f"problem file does not exist: {problem_file}")
    text = extract_problem_text(problem_file)
    questions = extract_questions(text)
    if not questions:
        raise ValueError("no explicit subproblem markers were found; initialization created no state")

    state_path = root / "state" / "decision_log.json"
    contest = contest_config(root)
    existing = state_path.is_file()
    if existing:
        state = load_json(state_path)
        if str(state.get("problem") or "") != problem:
            raise ValueError(f"existing decision state belongs to problem {state.get('problem')!r}")
        recorded_count = state.get("stages", {}).get("5", {}).get("qi_count")
        if recorded_count not in (None, len(questions)):
            raise ValueError(f"problem marker count changed from {recorded_count} to {len(questions)}; review before resuming")
    else:
        state = decision_template(root, workspace_root)
        state["competition"] = str(contest.get("competition", "CUMCM")).lower()
        state["problem"] = problem
        state["started_at"] = datetime.now(UTC).isoformat()
        state.setdefault("problem_meta", {})["year"] = contest.get("year")
        state["problem_meta"]["letter"] = problem
        state["problem_meta"]["deadline_iso"] = contest.get("deadline")
        stage0 = state.setdefault("stages", {}).setdefault("0", {})
        stage0.setdefault("problem_scan", {})["problem_id"] = problem
        stage0["problem_scan"]["subproblem_count"] = len(questions)
        stage5 = state["stages"].setdefault("5", {})
        stage5["qi_count"] = len(questions)
        stage5["qi_weights"] = [1.0] * len(questions)
        stage5["qi_status"] = {question_id: "pending" for question_id, _ in questions}

    question_template_path = shared_asset(root, workspace_root, "templates/workflow/question.yaml")
    template = load_yaml(question_template_path)
    if not template:
        raise FileNotFoundError(f"question template is missing: {question_template_path}")
    project_contract = load_yaml(root / "project.yaml") if (root / "project.yaml").is_file() else {}
    if template.get("schema_version") == 3 and int(project_contract.get("workflow_contract_version", 0) or 0) < 6:
        template = deepcopy(template)
        template["schema_version"] = 2
        template.pop("literature", None)
    source_relative = problem_file.relative_to(root).as_posix() if problem_file.is_relative_to(root) else problem_file.name
    expected_ids = {question_id for question_id, _ in questions}
    existing_ids = {path.parent.name for path in question_paths(root, problem)}
    stale_ids = sorted(existing_ids - expected_ids)
    if stale_ids:
        raise ValueError(f"existing question manifests are not present in the current problem markers: {', '.join(stale_ids)}")
    created: list[str] = []
    for question_id, target in questions:
        path = root / "problems" / problem / "questions" / question_id / "question.yaml"
        if path.exists():
            continue
        payload = deepcopy(template)
        payload["problem_id"] = problem
        payload["question_id"] = question_id
        payload["source_problem"] = source_relative
        payload["problem"]["target"] = target
        payload["paper"]["section"] = f"sections/question_{int(question_id[1:])}.tex"
        template_version = payload.get("schema_version")
        issues = question_v3_shape_issues(payload) if template_version == 3 else question_v2_shape_issues(payload)
        if issues:
            raise ValueError(f"question template does not satisfy schema v{template_version}: " + "; ".join(issues))
        dump_yaml(path, payload)
        created.append(path.relative_to(root).as_posix())

    claims_path = root / "results" / problem / "claims.json"
    if not claims_path.exists():
        dump_json(claims_path, {"schema_version": 1, "problem_id": problem, "claims": []})
    contracts_path = root / "paper" / "figure_contracts.yaml"
    if not contracts_path.exists():
        dump_yaml(contracts_path, {"schema_version": "2.0", "figures": []})
    code_manifest_path = root / "paper" / "code_manifest.yaml"
    if not code_manifest_path.exists():
        code_template = shared_asset(root, workspace_root, "templates/submission/code_manifest.yaml")
        if code_template.is_file():
            dump_yaml(code_manifest_path, load_yaml(code_template))
    paper_main = root / "paper" / "main.tex"
    if not paper_main.exists() and workspace_root is not None:
        family = str(contest.get("paper", {}).get("template_family") or "cumcm-2026")
        source_template = workspace_root / "templates" / "paper" / family
        if source_template.is_dir():
            generated_suffixes = {".aux", ".fdb_latexmk", ".fls", ".log", ".out", ".pdf", ".xdv", ".synctex.gz"}
            shutil.copytree(
                source_template,
                root / "paper",
                dirs_exist_ok=True,
                ignore=shutil.ignore_patterns(*generated_suffixes, "*.aux", "*.log", "*.pdf", "*.xdv", "*.fls", "*.out", "*.fdb_latexmk"),
            )
    references_bib = root / "paper" / "references.bib"
    if not references_bib.exists():
        references_bib.parent.mkdir(parents=True, exist_ok=True)
        references_bib.write_text("", encoding="utf-8")
    literature_plans: list[str] = []
    for question_id, _ in questions:
        question_path = root / "problems" / problem / "questions" / question_id / "question.yaml"
        question_payload = load_yaml(question_path)
        if question_payload.get("schema_version") != 3:
            continue
        literature = question_payload.get("literature", {}) if isinstance(question_payload.get("literature"), dict) else {}
        if literature.get("search_plan") is None:
            plan_result = literature_plan(root, problem, question_id)
            literature_plans.append(str(plan_result["search_plan"]))
    structure_path, blueprint_path = write_paper_blueprint(root, problem, source_relative, questions)
    if not existing:
        dump_json(state_path, state)

    contest["problem"] = problem
    dump_yaml(root / "contest.yaml", contest)
    return {
        "status": "INITIALIZED" if not existing else "RESUMED",
        "problem": problem,
        "questions": [item[0] for item in questions],
        "created": created,
        "literature_plans": literature_plans,
        "state": state_path.relative_to(root).as_posix(),
        "question_structure": structure_path.relative_to(root).as_posix(),
        "paper_blueprint": blueprint_path.relative_to(root).as_posix(),
    }


def add_check(checks: list[dict[str, Any]], name: str, passed: bool, detail: str, evidence: str | None = None) -> None:
    item = {"name": name, "passed": bool(passed), "detail": detail}
    if evidence:
        item["evidence"] = evidence
    checks.append(item)


def add_warning(warnings: list[dict[str, Any]], name: str, detail: str, evidence: str | None = None) -> None:
    item = {"name": name, "detail": detail}
    if evidence:
        item["evidence"] = evidence
    warnings.append(item)


def validate_run_manifest(root: Path, path: Path, checks: list[dict[str, Any]]) -> None:
    if not path.is_file():
        add_check(checks, "run_manifest_exists", False, str(path))
        return
    try:
        manifest = load_json(path)
    except Exception as exc:
        add_check(checks, "run_manifest_valid_json", False, str(exc), str(path))
        return
    required = ("run_id", "problem_id", "question_id", "engine", "command", "environment", "code", "random_seed", "methods", "artifacts", "metrics", "started_at_utc", "duration_seconds", "status")
    missing = [key for key in required if manifest.get(key) in (None, "", [])]
    add_check(checks, "run_manifest_fields", not missing, "missing: " + ", ".join(missing), str(path))
    if manifest.get("schema_version") == 2:
        lifecycle = manifest.get("lifecycle", {}) if isinstance(manifest.get("lifecycle"), dict) else {}
        formal = manifest.get("run_mode") == "formal" and manifest.get("mode") == "formal" and lifecycle.get("state") == "FORMAL" and lifecycle.get("formal") is True
        add_check(checks, "formal_run_manifest", formal, f"run_mode={manifest.get('run_mode')} state={lifecycle.get('state')}", str(path))
        primary_name = str(manifest.get("primary_metric") or "")
        snapshot = manifest.get("metric_snapshot", [])
        primary = next((item for item in snapshot if isinstance(item, dict) and item.get("name") == primary_name), None)
        add_check(checks, "formal_primary_metric_snapshot", bool(primary) and primary.get("value") is not None, primary_name or "missing", str(path))
        lifecycle_issues = _lifecycle_check_issues(manifest, ("input_output_match", "units_defined", "core_constraints_passed", "deterministic", "baseline_comparable"))
        replay = manifest.get("replay") if isinstance(manifest.get("replay"), dict) else {}
        formal_contract = not lifecycle_issues and replay.get("required") is True and int(replay.get("count", 0)) >= 2
        add_check(checks, "formal_lifecycle_contract", formal_contract, "; ".join(lifecycle_issues) if lifecycle_issues else f"replays={replay.get('count')}", str(path))
    roles = {str(item.get("role")) for item in manifest.get("methods", []) if isinstance(item, dict)}
    add_check(checks, "main_and_baseline_ran", {"main", "baseline"}.issubset(roles), f"roles: {sorted(roles)}", str(path))
    add_check(checks, "run_status", manifest.get("status") == "PASS", str(manifest.get("status")), str(path))
    code = manifest.get("code", {})
    code_path = root / str(code.get("runner", "")) if isinstance(code, dict) else root
    add_check(checks, "runner_hash", code_path.is_file() and code.get("sha256") == sha256(code_path), str(code_path), str(path))
    for artifact in manifest.get("artifacts", []):
        if not isinstance(artifact, dict) or not artifact.get("path"):
            add_check(checks, "artifact_record", False, "artifact path is missing", str(path))
            continue
        target = root / str(artifact["path"])
        valid = target.is_file() and artifact.get("sha256") == sha256(target)
        add_check(checks, "artifact_hash", valid, str(target), str(path))


def load_claims(root: Path, problem: str) -> tuple[Path, dict[str, Any]]:
    path = root / "results" / problem / "claims.json"
    return path, load_json(path) if path.is_file() else {"schema_version": 1, "problem_id": problem, "claims": []}


def _add_run_visual_design_checks(
    root: Path,
    run_path: Path,
    figure_ids: set[str],
    gate: str,
    checks: list[dict[str, Any]],
    required: bool = False,
) -> None:
    if not figure_ids or not run_path.is_file():
        return
    run_manifest = load_json(run_path)
    lifecycle = run_manifest.get("lifecycle") if isinstance(run_manifest.get("lifecycle"), dict) else {}
    if run_manifest.get("schema_version") != 2 or run_manifest.get("run_mode") != "formal" or lifecycle.get("formal") is not True:
        return
    data_path = run_path.parent / "figure_data_manifest.yaml"
    if not required and not data_path.is_file() and not (run_path.parent / "visual_intent.yaml").is_file() and not (run_path.parent / "figure_briefs").is_dir():
        return
    data_ok = data_path.is_file()
    data: dict[str, Any] = {}
    if data_ok:
        try:
            data = load_yaml(data_path)
            _validate_visual_payload(root, "figure_data_manifest.schema.json", data)
            source = _current_visual_source(root, str(data["source_run_manifest"]), str(data["source_run_manifest_sha256"]), "source run manifest")
            data_ok = source.resolve() == run_path.resolve() and bool(data.get("contest_evidence_eligible"))
        except (OSError, ValueError, KeyError):
            data_ok = False
    add_check(checks, f"{gate}_figure_data_manifest", data_ok, relative_path(root, data_path) if data_path.exists() else str(data_path), str(run_path))
    for figure_id in sorted(figure_ids):
        brief_path = run_path.parent / "figure_briefs" / f"{figure_id}.yaml"
        brief_ok = brief_path.is_file()
        status_value = "missing"
        if brief_ok:
            try:
                brief = load_yaml(brief_path)
                _validate_visual_payload(root, "figure_brief.schema.json", brief)
                status_value = str(brief.get("status"))
                allowed = {"REVIEWED", "APPROVED", "RENDERED", "QA_PASSED", "CONTRACT_READY"}
                if gate in {"G4", "G5", "G6"}:
                    allowed = {"APPROVED", "RENDERED", "QA_PASSED", "CONTRACT_READY"}
                brief_ok = (
                    status_value in allowed
                    and brief.get("source_data_manifest_sha256") == sha256(data_path)
                    and bool(brief.get("contest_evidence_eligible"))
                )
                _current_visual_source(root, str(brief["visual_intent"]), str(brief["visual_intent_sha256"]), "visual intent")
            except (OSError, ValueError, KeyError):
                brief_ok = False
        add_check(checks, f"{gate}_figure_brief", brief_ok, f"{figure_id}: {status_value}", str(brief_path))


def _design_handoff_valid(root: Path, contract: dict[str, Any]) -> tuple[bool, str]:
    handoff = contract.get("design_handoff")
    if not isinstance(handoff, dict):
        return False, "missing design_handoff"
    required = (
        ("data_manifest", "data_manifest_sha256"),
        ("visual_intent", "visual_intent_sha256"),
        ("figure_brief", "figure_brief_sha256"),
        ("render_qa", "render_qa_sha256"),
    )
    resolved: dict[str, Path] = {}
    try:
        for path_field, hash_field in required:
            resolved[path_field] = _current_visual_source(root, str(handoff.get(path_field, "")), str(handoff.get(hash_field, "")), path_field)
    except ValueError as exc:
        return False, str(exc)
    if handoff.get("design_status") != "APPROVED":
        return False, "design_status must be APPROVED"
    try:
        data = load_yaml(resolved["data_manifest"])
        intent = load_yaml(resolved["visual_intent"])
        brief = load_yaml(resolved["figure_brief"])
        qa = load_json(resolved["render_qa"])
        _validate_visual_payload(root, "figure_data_manifest.schema.json", data)
        _validate_visual_payload(root, "visual_intent.schema.json", intent)
        _validate_visual_payload(root, "figure_brief.schema.json", brief)
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        return False, str(exc)
    if data.get("status") != "DATA_READY" or not data.get("contest_evidence_eligible"):
        return False, "figure data is not formal evidence"
    if data.get("level") == "paper-evidence" and not _paper_evidence_is_ready(resolved["data_manifest"].parent):
        return False, "paper evidence parent or child run is stale"
    if intent.get("status") != "READY" or not intent.get("contest_evidence_eligible"):
        return False, "visual intent is not current formal evidence"
    if brief.get("status") not in {"QA_PASSED", "CONTRACT_READY"} or not brief.get("contest_evidence_eligible"):
        return False, "figure brief is not QA-passed formal evidence"
    if qa.get("passed") is not True or qa.get("status") != "QA_PASSED":
        return False, "render QA did not pass"
    if qa.get("brief_sha256") != sha256(resolved["figure_brief"]):
        return False, "render QA references a stale figure brief"
    if qa.get("data_manifest_sha256") != sha256(resolved["data_manifest"]):
        return False, "render QA references a stale data manifest"
    if qa.get("visual_intent_sha256") != sha256(resolved["visual_intent"]):
        return False, "render QA references a stale visual intent"
    try:
        expected_run = workspace_path(root, str(data.get("source_run_manifest", "")), "source run manifest")
    except ValueError as exc:
        return False, str(exc)
    data_issues = _figure_data_manifest_issues(root, resolved["data_manifest"], expected_run)
    brief_issues = _figure_brief_integrity_issues(root, brief, resolved["data_manifest"], resolved["visual_intent"])
    if data_issues or brief_issues:
        return False, "; ".join([*data_issues, *brief_issues])
    if brief.get("source_data_manifest_sha256") != sha256(resolved["data_manifest"]):
        return False, "figure brief references a stale data manifest"
    if brief.get("visual_intent_sha256") != sha256(resolved["visual_intent"]):
        return False, "figure brief references a stale visual intent"
    qa_hashes = {str(item.get("sha256")) for item in qa.get("outputs", []) if isinstance(item, dict) and item.get("sha256")}
    for value in contract.get("outputs", {}).values():
        if not isinstance(value, str):
            continue
        path = workspace_path(root, value, "figure output")
        if path.suffix.lower() in {".pdf", ".svg", ".png"} and (not path.is_file() or sha256(path) not in qa_hashes):
            return False, f"promoted output does not match render QA: {value}"
    return True, "approved brief and render QA are current"


def gate_checks(root: Path, problem: str, gate: str, question: str | None = None, strict: bool = False) -> list[dict[str, Any]]:
    visual_strict = strict or _project_requires_visual_handoff(root)
    literature_strict = strict or _project_requires_literature_handoff(root)
    paths = question_paths(root, problem)
    if question:
        paths = [path for path in paths if path.parent.name == question]
    checks: list[dict[str, Any]] = []
    add_check(checks, "question_manifests_exist", bool(paths), f"found {len(paths)} question manifest(s)")
    payloads: list[tuple[Path, dict[str, Any]]] = []
    for path in paths:
        try:
            payloads.append((path, load_yaml(path)))
        except Exception as exc:
            add_check(checks, "question_manifest_readable", False, str(exc), str(path))
    for path, payload in payloads:
        version = payload.get("schema_version")
        if version == 3:
            issues = question_v3_shape_issues(payload)
            add_check(checks, "question_manifest_v3_shape", not issues, "; ".join(issues) if issues else "schema v3", str(path))
        elif version == 2:
            issues = question_v2_shape_issues(payload)
            add_check(checks, "question_manifest_v2_shape", not issues, "; ".join(issues) if issues else "schema v2", str(path))
        elif version == 1:
            add_check(
                checks,
                "question_manifest_schema_version",
                not strict,
                "legacy schema v1 compatibility" if not strict else "strict validation requires schema v2 or v3",
                str(path),
            )
        else:
            add_check(checks, "question_manifest_schema_version", False, f"unsupported schema version: {version!r}", str(path))
        if version == 3 and gate in {"G0", "G1", "G2", "G3", "G4"}:
            literature = payload.get("literature", {}) if isinstance(payload.get("literature"), dict) else {}
            ready = literature.get("status") == "CITATION_READY"
            detail = "CITATION_READY" if ready else f"LITERATURE_INCOMPLETE warning: status={literature.get('status', 'NOT_STARTED')}"
            add_check(checks, "LITERATURE_INCOMPLETE", True, f"{payload.get('question_id')}: {detail}", str(path))

    if gate in {"G0", "G1", "G2", "G3", "G4", "G5", "G6"}:
        for path, payload in payloads:
            problem_data = payload.get("problem", {}) if isinstance(payload.get("problem"), dict) else {}
            add_check(checks, "G0_target", bool(str(problem_data.get("target", "")).strip()), payload.get("question_id", ""), str(path))
            add_check(checks, "G0_inputs", bool(problem_data.get("inputs")), payload.get("question_id", ""), str(path))
            add_check(checks, "G0_constraints", isinstance(problem_data.get("constraints"), list), payload.get("question_id", ""), str(path))
            add_check(checks, "G0_metrics", bool(problem_data.get("evaluation_metrics")), payload.get("question_id", ""), str(path))
            if payload.get("schema_version") in {2, 3}:
                paper = payload.get("paper", {}) if isinstance(payload.get("paper"), dict) else {}
                argument = paper.get("argument_contract", {}) if isinstance(paper.get("argument_contract"), dict) else {}
                expected_section = f"sections/question_{str(payload.get('question_id', ''))[1:]}.tex"
                add_check(checks, "G0_problem_type", bool(str(problem_data.get("type", "")).strip()), payload.get("question_id", ""), str(path))
                add_check(checks, "G0_outputs", bool(problem_data.get("outputs")), payload.get("question_id", ""), str(path))
                add_check(checks, "G0_dependencies", isinstance(problem_data.get("dependencies"), list), payload.get("question_id", ""), str(path))
                add_check(checks, "G0_key_conflicts", isinstance(problem_data.get("key_conflicts"), list), payload.get("question_id", ""), str(path))
                add_check(checks, "G0_paper_section", paper.get("section") == expected_section, str(paper.get("section", "")), str(path))
                add_check(checks, "G0_argument_objective_interface", argument.get("objective_interface") == "complete", payload.get("question_id", ""), str(path))
    if gate in {"G1", "G2", "G3", "G4", "G5", "G6"}:
        for path, payload in payloads:
            method = payload.get("method", {})
            main_value = method.get("main", {}) if isinstance(method, dict) else {}
            baseline_value = method.get("baseline", {}) if isinstance(method, dict) else {}
            main = main_value if isinstance(main_value, dict) else {}
            baseline = baseline_value if isinstance(baseline_value, dict) else {}
            fallback = method.get("fallback") if isinstance(method, dict) else None
            add_check(checks, "G1_main", all(str(main.get(key, "")).strip() for key in ("name", "rationale", "implementation")), payload.get("question_id", ""), str(path))
            add_check(checks, "G1_baseline", all(str(baseline.get(key, "")).strip() for key in ("name", "implementation")) and baseline.get("comparable_output") is True, payload.get("question_id", ""), str(path))
            add_check(checks, "G1_fallback", fallback is None or (isinstance(fallback, dict) and bool(str(fallback.get("name", "")).strip()) and bool(str(fallback.get("trigger", "")).strip())), payload.get("question_id", ""), str(path))
            add_check(checks, "G1_risk_probes", bool(payload.get("risk_probes")), payload.get("question_id", ""), str(path))
            if payload.get("schema_version") in {2, 3}:
                selection = payload.get("model_selection", {}) if isinstance(payload.get("model_selection"), dict) else {}
                paper = payload.get("paper", {}) if isinstance(payload.get("paper"), dict) else {}
                argument = paper.get("argument_contract", {}) if isinstance(paper.get("argument_contract"), dict) else {}
                selection_complete = all(str(selection.get(key, "")).strip() for key in ("primary", "rationale", "baseline"))
                add_check(checks, "G1_model_selection", selection_complete, payload.get("question_id", ""), str(path))
                add_check(checks, "G1_rejected_alternative", bool(selection.get("rejected_alternatives")), payload.get("question_id", ""), str(path))
                add_check(checks, "G1_argument_model_choice", argument.get("model_choice") == "complete", payload.get("question_id", ""), str(path))
    if gate in {"G2", "G3", "G4", "G5", "G6"}:
        for path, payload in payloads:
            decisions = payload.get("decisions", [])
            confirmed = any(isinstance(item, dict) and item.get("status") == "confirmed" and item.get("evidence_ref") for item in decisions)
            add_check(checks, "G2_human_decision", confirmed, payload.get("question_id", ""), str(path))
            if payload.get("schema_version") in {2, 3}:
                add_check(checks, "G2_testable_assumptions", bool(payload.get("assumptions")), payload.get("question_id", ""), str(path))
    if gate in {"G3", "G4", "G5", "G6"}:
        for path, payload in payloads:
            evidence = payload.get("evidence", {}) if isinstance(payload.get("evidence"), dict) else {}
            runs = evidence.get("runs", [])
            add_check(checks, "G3_run_refs", bool(runs), payload.get("question_id", ""), str(path))
            paper = payload.get("paper", {}) if isinstance(payload.get("paper"), dict) else {}
            declared_figures = {str(item) for item in paper.get("figure_ids", []) if item}
            for run_ref in runs:
                run_path = root / str(run_ref)
                validate_run_manifest(root, run_path, checks)
                _add_run_visual_design_checks(root, run_path, declared_figures, gate, checks, required=visual_strict)
            if payload.get("schema_version") in {2, 3}:
                argument = paper.get("argument_contract", {}) if isinstance(paper.get("argument_contract"), dict) else {}
                add_check(checks, "G3_result_claim_refs", bool(evidence.get("result_claim_ids")), payload.get("question_id", ""), str(path))
                has_result_asset = bool(paper.get("table_ids") or paper.get("figure_ids"))
                if visual_strict and not has_result_asset:
                    has_result_asset, _, _ = _question_visual_decision(root, payload, {"text", "none"})
                add_check(checks, "G3_paper_result_asset", has_result_asset, payload.get("question_id", ""), str(path))
                add_check(checks, "G3_code_refs", bool(paper.get("code_refs")), payload.get("question_id", ""), str(path))
                complete = all(argument.get(field) == "complete" for field in ("formulation", "algorithm", "result"))
                add_check(checks, "G3_argument_model_result", complete, payload.get("question_id", ""), str(path))
    claims_path, claims_payload = load_claims(root, problem)
    selected_question_ids = {str(payload.get("question_id")) for _, payload in payloads if payload.get("question_id")}
    relevant_claims = [item for item in claims_payload.get("claims", []) if not question or item.get("question_id") == question]
    claims_by_question = {
        question_id: [item for item in claims_payload.get("claims", []) if str(item.get("question_id")) == question_id]
        for question_id in selected_question_ids
    }
    if gate in {"G3", "G4", "G5", "G6"}:
        for path, payload in payloads:
            if payload.get("schema_version") not in {2, 3}:
                continue
            question_id = str(payload.get("question_id", ""))
            question_claims = claims_by_question.get(question_id, [])
            known_claim_ids = {str(item.get("id")) for item in question_claims if item.get("id")}
            evidence = payload.get("evidence", {}) if isinstance(payload.get("evidence"), dict) else {}
            fields = ["result_claim_ids"]
            if gate in {"G4", "G5", "G6"}:
                fields.extend(("validation_claim_ids", "boundary_claim_ids"))
            declared = {str(item) for field in fields for item in evidence.get(field, []) if item}
            unknown = sorted(declared - known_claim_ids)
            add_check(checks, "claim_handoff_ids_resolve", not unknown, f"unknown: {unknown or 'none'}", str(path))
            if gate in {"G4", "G5", "G6"}:
                frozen_ids = {str(item.get("id")) for item in question_claims if item.get("status") == "frozen" and item.get("id")}
                unfrozen = sorted(declared - frozen_ids)
                add_check(checks, "claim_handoff_ids_frozen", not unfrozen, f"unfrozen: {unfrozen or 'none'}", str(path))
    if gate in {"G4", "G5", "G6"}:
        for path, payload in payloads:
            evidence = payload.get("evidence", {}) if isinstance(payload.get("evidence"), dict) else {}
            robustness = evidence.get("robustness")
            valid = bool(robustness) and locator_path(root, str(robustness)).is_file()
            add_check(checks, "G4_robustness", valid, payload.get("question_id", ""), str(path))
            if payload.get("schema_version") in {2, 3}:
                paper = payload.get("paper", {}) if isinstance(payload.get("paper"), dict) else {}
                argument = paper.get("argument_contract", {}) if isinstance(paper.get("argument_contract"), dict) else {}
                add_check(checks, "G4_validation_claim_refs", bool(evidence.get("validation_claim_ids")), payload.get("question_id", ""), str(path))
                add_check(checks, "G4_boundary_claim_refs", bool(evidence.get("boundary_claim_ids")), payload.get("question_id", ""), str(path))
                complete = all(argument.get(field) == "complete" for field in ("validation", "conclusion"))
                add_check(checks, "G4_argument_validation_conclusion", complete, payload.get("question_id", ""), str(path))
        frozen = [item for item in relevant_claims if item.get("status") == "frozen"]
        add_check(checks, "G4_frozen_claims", bool(frozen), f"found {len(frozen)} frozen claim(s)", str(claims_path))
        for claim in frozen:
            try:
                evidence = locator_path(root, str(claim.get("locator", "")))
                current_value = locator_value(root, str(claim.get("locator", "")))
                valid = (
                    evidence.is_file()
                    and not _is_literature_evidence_path(root, evidence)
                    and claim.get("evidence_sha256") == sha256(evidence)
                    and current_value == claim.get("value")
                )
            except ValueError:
                valid = False
            add_check(checks, "G4_frozen_hash", valid, str(claim.get("id")), str(claims_path))
    if gate in {"G5", "G6"}:
        literature_report = literature_audit(root, problem, question, strict=literature_strict, write=False)
        for item in literature_report.get("checks", []):
            copied = dict(item)
            copied["name"] = f"G5_{item.get('name', 'literature')}"
            checks.append(copied)
        contracts_path = root / "paper" / "figure_contracts.yaml"
        contracts = load_yaml(contracts_path).get("figures", []) if contracts_path.is_file() else []
        frozen_ids = {str(item.get("id")) for item in relevant_claims if item.get("status") == "frozen"}
        formal_contracts = [item for item in contracts if str(item.get("question_id")) in selected_question_ids]
        for path, payload in payloads:
            if payload.get("schema_version") not in {2, 3}:
                continue
            question_id = str(payload.get("question_id", ""))
            paper = payload.get("paper", {}) if isinstance(payload.get("paper"), dict) else {}
            declared_figures = {str(item) for item in paper.get("figure_ids", []) if item}
            declared_tables = {str(item) for item in paper.get("table_ids", []) if item}
            contract_ids = {
                str(item.get("id")) for item in formal_contracts
                if str(item.get("question_id")) == question_id and item.get("id")
            }
            strict_visual = _project_requires_visual_handoff(root)
            nonfigure_ok = False
            nonfigure_detail = "not required"
            nonfigure_decision = ""
            if strict_visual and not declared_figures and not declared_tables:
                nonfigure_ok, nonfigure_detail, nonfigure_decision = _question_visual_decision(
                    root,
                    payload,
                    {"text", "none"},
                )
            add_check(
                checks,
                "G5_result_artifact_binding",
                bool(declared_figures or declared_tables or nonfigure_ok),
                (
                    f"{question_id}: figures={sorted(declared_figures)}, tables={sorted(declared_tables)}, "
                    f"nonfigure={nonfigure_decision or 'none'}"
                ),
                str(path),
            )
            if strict_visual and declared_tables and not declared_figures:
                table_decision_ok, table_decision_detail, _ = _question_visual_decision(root, payload, {"table"})
                add_check(checks, "G5_table_design_decision", table_decision_ok, table_decision_detail, str(path))
            if strict_visual and not declared_figures and not declared_tables:
                add_check(checks, "G5_nonfigure_design_decision", nonfigure_ok, nonfigure_detail, str(path))
            missing_contracts = sorted(declared_figures - contract_ids)
            undeclared_contracts = sorted(contract_ids - declared_figures)
            add_check(
                checks,
                "G5_figure_contracts",
                not missing_contracts and not undeclared_contracts,
                f"{question_id}: missing={missing_contracts or 'none'}, undeclared={undeclared_contracts or 'none'}",
                str(contracts_path),
            )
        if any(payload.get("schema_version") == 1 for _, payload in payloads):
            add_check(checks, "G5_figure_contracts", bool(formal_contracts), f"legacy manifests: found {len(formal_contracts)} formal contract(s)", str(contracts_path))
        for contract in formal_contracts:
            identifier = str(contract.get("id", ""))
            add_check(checks, "G5_figure_claim_frozen", str(contract.get("claim_id")) in frozen_ids, identifier, str(contracts_path))
            required = (
                "contract_version", "id", "question_id", "claim_id", "core_conclusion", "evidence_chain",
                "kind", "archetype", "backend", "source_data", "source_script", "outputs", "baseline",
                "axes", "caption", "panel_map", "statistics", "review_risks", "final_width_mm", "min_font_pt",
                *FORMAL_FIGURE_FIELDS,
            )
            # An empty rasterized_layers list is an explicit declaration that all
            # layers remain vector. Other required list fields must carry content.
            missing = [
                key for key in required
                if key not in contract
                or contract.get(key) in (None, "")
                or (contract.get(key) == [] and key != "rasterized_layers")
            ]
            version_ok = contract.get("contract_version") == "2.0"
            add_check(checks, "G5_figure_fields", not missing and version_ok, f"{identifier}: missing {missing or 'none'}; version={contract.get('contract_version')}", str(contracts_path))
            target_profile = str(contract.get("target_size_profile", ""))
            expected_width = FIGURE_SIZE_WIDTHS.get(target_profile)
            actual_width = contract.get("final_width_mm")
            size_profile_ok = (
                expected_width is not None
                and isinstance(actual_width, (int, float))
                and abs(float(actual_width) - expected_width) <= 0.6
                and isinstance(contract.get("min_font_pt"), (int, float))
                and float(contract["min_font_pt"]) >= 8.0
            )
            add_check(checks, "G5_figure_physical_size", size_profile_ok, f"{identifier}: profile={target_profile}, width={actual_width}, min_font={contract.get('min_font_pt')}", str(contracts_path))
            label_strategy = contract.get("label_strategy")
            label_ok = (
                isinstance(label_strategy, dict)
                and label_strategy.get("mode") in {"direct", "external-legend", "none"}
                and label_strategy.get("collision_checked") is True
            )
            add_check(checks, "G5_figure_label_strategy", label_ok, identifier, str(contracts_path))
            integrity = contract.get("data_integrity")
            integrity_ok = (
                isinstance(integrity, dict)
                and integrity.get("manual_values_forbidden") is True
                and bool(str(integrity.get("transformation", "")).strip())
                and isinstance(integrity.get("source_hashes"), list)
                and bool(integrity.get("source_hashes"))
            )
            if integrity_ok:
                for source_hash in integrity["source_hashes"]:
                    if not isinstance(source_hash, dict) or not re.fullmatch(r"[0-9a-f]{64}", str(source_hash.get("sha256", ""))):
                        integrity_ok = False
                        continue
                    try:
                        integrity_path = locator_path(root, str(source_hash.get("path", "")))
                        integrity_ok = integrity_ok and integrity_path.is_file() and sha256(integrity_path) == source_hash.get("sha256")
                    except ValueError:
                        integrity_ok = False
            add_check(checks, "G5_figure_data_integrity", integrity_ok, identifier, str(contracts_path))
            panel_map = contract.get("panel_map")
            single_or_justified = isinstance(panel_map, list) and (len(panel_map) <= 1 or bool(str(contract.get("multipanel_justification", "")).strip()))
            add_check(checks, "G5_figure_panel_policy", single_or_justified, identifier, str(contracts_path))
            try:
                script = locator_path(root, str(contract.get("source_script", "")))
            except ValueError:
                script = root / "__invalid_source_script__"
            source_data = contract.get("source_data")
            source_data_ok = isinstance(source_data, list) and bool(source_data)
            for item in source_data if isinstance(source_data, list) else []:
                try:
                    source_data_ok = source_data_ok and locator_path(root, str(item)).is_file()
                except ValueError:
                    source_data_ok = False
            add_check(checks, "G5_figure_sources", script.is_file() and source_data_ok, identifier, str(contracts_path))
            outputs_value = contract.get("outputs")
            output_paths: list[Path] = []
            output_paths_ok = isinstance(outputs_value, dict)
            for key in ("pdf", "svg", "png"):
                try:
                    output_paths.append(locator_path(root, str(outputs_value.get(key, ""))))
                except (AttributeError, ValueError):
                    output_paths_ok = False
            output_extensions = {path.suffix.lower() for path in output_paths if path.is_file()}
            exports_ok = output_paths_ok and outputs_value.get("png_dpi") == 400 and {".pdf", ".svg", ".png"}.issubset(output_extensions)
            add_check(checks, "G5_figure_exports", exports_ok, identifier, str(contracts_path))
            axes_ok = isinstance(contract.get("axes"), list) and all(isinstance(axis, dict) and axis.get("variable") and axis.get("unit") for axis in contract.get("axes", []))
            if contract.get("kind") == "schematic":
                axes_ok = axes_ok or bool(contract.get("axes_not_applicable_reason"))
            add_check(checks, "G5_figure_axes", axes_ok, identifier, str(contracts_path))
            evidence_entries = contract.get("evidence_chain", []) if isinstance(contract.get("evidence_chain"), list) else []
            evidence_ok = bool(evidence_entries)
            for chain in evidence_entries:
                if not isinstance(chain, dict) or not chain.get("locator") or not re.fullmatch(r"[0-9a-f]{64}", str(chain.get("sha256", ""))):
                    evidence_ok = False
                    continue
                try:
                    evidence_path = locator_path(root, str(chain["locator"]))
                    evidence_ok = evidence_ok and evidence_path.is_file() and sha256(evidence_path) == chain.get("sha256") and bool(chain.get("fields"))
                except ValueError:
                    evidence_ok = False
            add_check(checks, "G5_figure_evidence_hash", evidence_ok, identifier, str(contracts_path))
            if visual_strict or contract.get("design_handoff") is not None:
                handoff_ok, handoff_detail = _design_handoff_valid(root, contract)
                add_check(checks, "G5_figure_design_handoff", handoff_ok, f"{identifier}: {handoff_detail}", str(contracts_path))
            else:
                add_check(checks, "G5_figure_design_handoff_compat", True, f"{identifier}: legacy contract without V5 design handoff", str(contracts_path))
        exploratory = [item.get("id") for item in relevant_claims if item.get("status") in {"exploratory", "verified"}]
        add_check(checks, "G5_no_unfrozen_active_claims", not exploratory, ", ".join(map(str, exploratory)), str(claims_path))
        if question is None:
            audit_script = Path(__file__).resolve().parents[1] / "utils" / "audit_latex.py"
            sentinel_log = root / "output" / "_verification" / "__g5_static_no_log__.log"
            command = [
                sys.executable,
                str(audit_script),
                "--paper-dir",
                str(root / "paper"),
                "--log",
                str(sentinel_log),
                "--structure-strict",
            ]
            contest_path = root / "contest.yaml"
            if contest_path.is_file():
                command.extend(("--contest-config", str(contest_path)))
            try:
                completed = subprocess.run(command, cwd=root, capture_output=True, text=True, check=False)
                static_report = json.loads(completed.stdout)
                static_passed = static_report.get("passed") is True
                error_codes = [str(item.get("code")) for item in static_report.get("errors", []) if item.get("code")]
                detail = "passed" if static_passed else f"errors: {error_codes or ['audit process failed']}"
            except (OSError, json.JSONDecodeError) as exc:
                static_passed = False
                detail = str(exc)
            add_check(checks, "G5_paper_static_audit", static_passed, detail, str(root / "paper" / "main.tex"))
    if gate == "G6":
        for name in ("paper_audit.json", "figure_audit.json", "figure_style_audit.json", "pdf_visual_audit.json", "code_parity_audit.json", "ai_usage_audit.json", "package_audit.json", "audit.json"):
            path = root / "output" / name
            passed = False
            if path.is_file():
                report = load_json(path)
                passed = report.get("passed", report.get("status") == "PASS") is True
            add_check(checks, "G6_audit_report", passed, name, str(path))
    return checks


def validate(root: Path, problem: str, gate: str, question: str | None = None, write: bool = True, strict: bool = False) -> dict[str, Any]:
    gate = gate.upper()
    if gate not in {f"G{index}" for index in range(7)}:
        raise ValueError(f"unsupported gate: {gate}")
    checks = gate_checks(root, problem, gate, question, strict=strict)
    report = {"schema_version": 1, "problem": problem, "question": question, "gate": gate, "strict": strict, "passed": bool(checks) and all(item["passed"] for item in checks), "checks": checks, "generated_at_utc": datetime.now(UTC).isoformat()}
    if write:
        dump_json(root / "output" / "workflow_status.json", report)
    return report


def freeze(root: Path, problem: str, question: str, decision_id: str) -> dict[str, Any]:
    decision_id = safe_token(decision_id, "decision-id")
    prerequisite = validate(root, problem, "G3", question, write=False)
    if not prerequisite["passed"]:
        raise ValueError("G3 must pass before claims can be frozen")
    question_path = root / "problems" / problem / "questions" / question / "question.yaml"
    question_data = load_yaml(question_path)
    robustness = question_data.get("evidence", {}).get("robustness")
    if not robustness or not locator_path(root, str(robustness)).is_file():
        raise ValueError("robustness evidence is required before freezing claims")
    claims_path, payload = load_claims(root, problem)
    changed: list[str] = []
    for claim in payload.get("claims", []):
        if claim.get("question_id") != question or claim.get("status") != "verified":
            continue
        evidence = locator_path(root, str(claim.get("locator", "")))
        if _is_literature_evidence_path(root, evidence):
            raise ValueError(f"claim {claim.get('id')} cannot use literature or search-cache content as project evidence")
        if not evidence.is_file():
            raise FileNotFoundError(f"claim evidence does not exist: {evidence}")
        value = locator_value(root, str(claim.get("locator", "")))
        if value is None:
            raise ValueError(f"claim locator must select a primitive JSON value: {claim.get('locator')}")
        claim["evidence_sha256"] = sha256(evidence)
        claim["value"] = value
        claim["decision_id"] = decision_id
        claim["frozen_at_utc"] = datetime.now(UTC).isoformat()
        claim["status"] = "frozen"
        changed.append(str(claim.get("id")))
    if not changed:
        raise ValueError(f"no verified claims are available for {question}")
    dump_json(claims_path, payload)
    question_data["status"] = "FROZEN"
    dump_yaml(question_path, question_data)
    report = validate(root, problem, "G4", question)
    return {"status": "FROZEN", "claims": changed, "gate": report}


def status(root: Path) -> dict[str, Any]:
    contest = contest_config(root)
    project_meta = load_yaml(root / "project.yaml") if (root / "project.yaml").is_file() else {}
    state_path = root / "state" / "decision_log.json"
    if not state_path.is_file():
        report = {
            "schema_version": 1,
            "project_id": project_meta.get("project_id", contest.get("project_id", "legacy-root")),
            "competition": contest.get("competition"),
            "year": contest.get("year"),
            "phase": "PRECONTEST",
            "problem": contest.get("problem", "TBD"),
            "state_exists": False,
            "allowed_actions": ["preflight", "status", "corpus", "template", "tooling"],
            "generated_at_utc": datetime.now(UTC).isoformat(),
        }
    else:
        state = load_json(state_path)
        problem = str(state.get("problem") or contest.get("problem"))
        gates: dict[str, Any] = {gate: validate(root, problem, gate, write=False)["passed"] for gate in (f"G{i}" for i in range(5))}
        for gate, candidates in {
            "G5": (root / "output" / "paper_audit.json", root / "output" / "figure_audit.json"),
            "G6": (root / "output" / "audit.json", root / "output" / "package_audit.json"),
        }.items():
            existing = [path for path in candidates if path.is_file()]
            if not existing:
                gates[gate] = "NOT_RUN"
            else:
                gates[gate] = all(load_json(path).get("passed", load_json(path).get("status") == "PASS") is True for path in existing)
        report = {"schema_version": 1, "project_id": project_meta.get("project_id", contest.get("project_id", "legacy-root")), "competition": contest.get("competition"), "year": contest.get("year"), "phase": "ACTIVE", "problem": problem, "state_exists": True, "current_stage": state.get("current_stage"), "gates": gates, "generated_at_utc": datetime.now(UTC).isoformat()}
    dump_json(root / "output" / "workflow_status.json", report)
    return report


def prompt(
    root: Path,
    project_id: str,
    stage: str,
    role: str,
    question: str | None,
    workspace_root: Path | None = None,
) -> dict[str, Any]:
    """Assemble a prompt preview without mutating competition state or evidence."""

    packet = assemble_packet(root, project_id, stage, role, question, workspace_root)
    folder = root / "output" / "_verification" / "prompts" / stage / role / (question or "project")
    packet_path = folder / "prompt_packet.yaml"
    receipt_path = folder / "prompt_receipt.json"
    dump_yaml(packet_path, packet)
    receipt = format_receipt({
        "status": "READY",
        "objective": packet["objective"],
        "conclusion": "Prompt packet assembled; no competition state or formal evidence was modified.",
        "evidence": [f"{relative_path(root, packet_path)}#sha256={sha256(packet_path)}"],
        "warnings": [],
        "next_action": "Run the role task using this packet and return the compact receipt.",
        "decision_request": None,
    })
    dump_json(receipt_path, receipt)
    return {
        "status": "READY",
        "packet": relative_path(root, packet_path),
        "receipt": relative_path(root, receipt_path),
        "project_id": project_id,
        "stage": stage,
        "role": role,
        "question_id": question or "",
    }


def resolve_run_config(root: Path, path: Path) -> dict[str, Any]:
    config = load_yaml(path)
    required = ("experiment_id", "problem", "question", "engine", "runner", "seed", "output_root", "methods", "metrics")
    missing = [key for key in required if config.get(key) in (None, "", [])]
    if missing:
        raise ValueError("run configuration is missing: " + ", ".join(missing))
    config["experiment_id"] = safe_token(str(config["experiment_id"]), "experiment-id")
    config["problem"] = safe_token(str(config["problem"]), "problem")
    config["question"] = safe_token(str(config["question"]), "question")
    if not re.fullmatch(r"Q[1-9][0-9]*", config["question"]):
        raise ValueError("question must use Q<number> format")
    if config["engine"] not in {"python", "matlab"}:
        raise ValueError("engine must be python or matlab")
    version = int(config.get("schema_version", 1))
    if version not in {1, 2}:
        raise ValueError(f"unsupported experiment schema version: {version}")
    run_mode = str(config.get("run_mode") or config.get("level") or ("formal" if version == 1 else "scratch"))
    if run_mode not in EXPERIMENT_LEVELS:
        raise ValueError(f"run_mode must be one of {', '.join(EXPERIMENT_LEVELS)}")
    level = str(config.get("level") or run_mode)
    if level != run_mode:
        raise ValueError("level and run_mode must match")
    mode = str(config.get("mode") or ("formal" if run_mode == "formal" else "probe"))
    if mode not in {"probe", "formal"}:
        raise ValueError("experiment mode must be probe or formal")
    if run_mode in NONFORMAL_LEVELS and mode != "probe":
        raise ValueError(f"{run_mode} experiments must use probe mode")
    if run_mode == "formal" and version == 2:
        raise ValueError("formal experiments must be created through promote, not run directly")
    if run_mode == "paper-evidence":
        required_paper_fields = ("source_run_id", "source_manifest_sha256", "evidence_scope")
        missing_paper = [field for field in required_paper_fields if not config.get(field)]
        if missing_paper:
            raise ValueError("paper-evidence experiment is missing: " + ", ".join(missing_paper))
        if config.get("evidence_scope") not in {"diagnostic", "sensitivity", "mechanism", "figure_support"}:
            raise ValueError("invalid paper-evidence scope")
    runner = workspace_path(root, str(config["runner"]), "runner")
    if not runner.is_file():
        raise FileNotFoundError(f"runner does not exist: {runner}")
    output_root = workspace_path(root, str(config["output_root"]), "output-root")
    for item in config.get("inputs", []):
        workspace_path(root, str(item), "input")
    experiment_root = (output_root / config["experiment_id"]).resolve()
    try:
        experiment_root.relative_to(root.resolve())
    except ValueError as exc:
        raise ValueError("experiment output escapes the selected project root") from exc
    config["config_path"] = path.resolve().relative_to(root).as_posix()
    config["runner_path"] = runner.relative_to(root).as_posix()
    config["experiment_root"] = experiment_root.relative_to(root).as_posix()
    config["schema_version"] = version
    config["arguments"] = [str(item) for item in config.get("arguments", [])]
    config["diagnostic_arguments"] = [str(item) for item in config.get("diagnostic_arguments", [])]
    config["run_mode"] = run_mode
    config["level"] = level
    config["mode"] = mode
    config["purpose"] = str(config.get("purpose") or ("paper" if run_mode == "paper-evidence" else "exploration"))
    if config["purpose"] not in {"exploration", "candidate", "paper"}:
        raise ValueError("experiment purpose must be exploration, candidate, or paper")
    config["formal_candidate"] = bool(config.get("formal_candidate", run_mode == "candidate"))
    config["parent_run_id"] = config.get("parent_run_id")
    config["source_run_id"] = config.get("source_run_id")
    config["source_manifest"] = config.get("source_manifest")
    config["source_manifest_sha256"] = config.get("source_manifest_sha256")
    config["evidence_scope"] = config.get("evidence_scope")
    config["checkpoint_id"] = config.get("checkpoint_id")
    config["primary_metric"] = config.get("primary_metric")
    default_checks = {
        "input_output_match": False,
        "units_defined": False,
        "core_constraints_passed": False,
        "deterministic": False,
        "baseline_comparable": False,
    }
    checks = config.get("checks") if isinstance(config.get("checks"), dict) else {}
    config["checks"] = {key: bool(checks.get(key, default)) for key, default in default_checks.items()}
    if version == 2 and set(checks) != set(default_checks):
        raise ValueError("schema v2 experiment checks must define all lifecycle checks")
    default_reuse = {"seed": False, "environment": False, "code": False, "inputs": False, "methods": False, "parameters": False}
    reuse = config.get("reuse_contract") if isinstance(config.get("reuse_contract"), dict) else {}
    config["reuse_contract"] = {key: bool(reuse.get(key, default)) for key, default in default_reuse.items()}
    if run_mode == "paper-evidence" and not all(config["reuse_contract"].values()):
        raise ValueError("paper-evidence reuse_contract must require seed, environment, code, inputs, methods, and parameters")
    if run_mode != "paper-evidence" and config["diagnostic_arguments"]:
        raise ValueError("diagnostic_arguments are reserved for paper-evidence runs")
    replay = config.get("replay") if isinstance(config.get("replay"), dict) else {}
    config["replay"] = {"required": bool(replay.get("required", False)), "count": int(replay.get("count", 1))}
    if config["replay"]["count"] < 1:
        raise ValueError("replay count must be at least 1")
    return config


def _metric_locator(root: Path, experiment_root: Path, locator: str) -> tuple[str, Any]:
    if not locator:
        return "", None
    selector = ""
    raw = locator
    if ":" in locator:
        raw, selector = locator.split(":", 1)
    raw = raw.strip()
    candidates: list[Path] = []
    try:
        candidates.append(workspace_path(root, raw, "metric locator"))
    except ValueError:
        raise
    if raw and not Path(raw).is_absolute():
        candidate = (experiment_root / raw).resolve()
        if path_is_within(candidate, root) and candidate not in candidates:
            candidates.append(candidate)
    for candidate in candidates:
        if not candidate.is_file():
            continue
        canonical = relative_path(root, candidate)
        canonical_locator = canonical + (f":{selector}" if selector else "")
        return canonical_locator, locator_value(root, canonical_locator) if selector else None
    return locator, None


def snapshot_metrics(root: Path, experiment_root: Path, metrics: list[Any], primary_name: str | None = None) -> list[dict[str, Any]]:
    snapshots: list[dict[str, Any]] = []
    dictionaries = [item for item in metrics if isinstance(item, dict)]
    selected_primary = primary_name or next((str(item.get("name")) for item in dictionaries if item.get("primary")), None)
    if not selected_primary and dictionaries:
        selected_primary = str(dictionaries[0].get("name", ""))
    for item in dictionaries:
        locator, value = _metric_locator(root, experiment_root, str(item.get("locator", "")))
        snapshots.append({
            "name": str(item.get("name", "")),
            "value": value,
            "unit": str(item.get("unit", "")),
            "locator": locator,
            "primary": str(item.get("name", "")) == selected_primary,
        })
    return snapshots


def _manifest_path_for_run(root: Path, problem: str, question: str, run_id: str) -> Path:
    problem = safe_token(problem, "problem")
    question = safe_token(question, "question")
    run_id = safe_token(run_id, "run-id")
    candidates = [root / "experiments" / problem / question / level / run_id / "run_manifest.json" for level in EXPERIMENT_LEVELS]
    candidates.append(root / "experiments" / problem / question / run_id / "run_manifest.json")
    found = [path for path in candidates if path.is_file()]
    if not found:
        raise FileNotFoundError(f"run manifest not found for {problem}/{question}/{run_id}")
    if len(found) > 1:
        raise ValueError(f"run id is ambiguous across lifecycle levels: {run_id}")
    return found[0]


def _manifest_integrity_issues(
    root: Path,
    manifest: dict[str, Any],
    require_baseline: bool = False,
    *,
    verify_hashes: bool = True,
    verify_artifact_hashes: bool = True,
) -> list[str]:
    issues: list[str] = []
    if manifest.get("status") != "PASS":
        issues.append(f"run status is {manifest.get('status')!r}, expected PASS")
    code = manifest.get("code") if isinstance(manifest.get("code"), dict) else {}
    runner_value = str(code.get("runner", ""))
    try:
        runner = workspace_path(root, runner_value, "runner")
    except ValueError as exc:
        issues.append(str(exc))
    else:
        if not runner.is_file() or (verify_hashes and code.get("sha256") != sha256(runner)):
            issues.append(f"runner hash mismatch: {runner_value}")
    for collection in ("inputs", "artifacts"):
        for item in manifest.get(collection, []):
            if not isinstance(item, dict) or not item.get("path"):
                issues.append(f"{collection} contains an invalid record")
                continue
            try:
                target = workspace_path(root, str(item["path"]), collection[:-1])
            except ValueError as exc:
                issues.append(str(exc))
                continue
            expected = item.get("sha256")
            check_hash = verify_hashes and (collection == "inputs" or verify_artifact_hashes)
            if not target.is_file() or (check_hash and (not expected or expected != sha256(target))):
                issues.append(f"{collection[:-1]} hash mismatch: {item['path']}")
    roles = {str(item.get("role")) for item in manifest.get("methods", []) if isinstance(item, dict)}
    if "main" not in roles:
        issues.append("main method is missing")
    if require_baseline and "baseline" not in roles:
        issues.append("comparable baseline is missing")
    return issues


def _lifecycle_check_issues(manifest: dict[str, Any], required_fields: tuple[str, ...]) -> list[str]:
    checks = manifest.get("checks") if isinstance(manifest.get("checks"), dict) else {}
    return [f"lifecycle check failed: {field}" for field in required_fields if checks.get(field) is not True]


def _probe_receipt(
    root: Path,
    manifest_path: Path,
    manifest: dict[str, Any],
    receipt_type: str,
    status_value: str,
    note: str = "",
) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "receipt_type": receipt_type,
        "receipt_id": f"{manifest['run_id']}-{receipt_type}",
        "problem_id": manifest["problem_id"],
        "question_id": manifest["question_id"],
        "experiment_id": manifest["run_id"],
        "run_id": manifest["run_id"],
        "source_manifest": relative_path(root, manifest_path),
        "source_manifest_sha256": sha256(manifest_path),
        "status": status_value,
        "formal": False,
        "code": deepcopy(manifest.get("code", {})),
        "inputs": deepcopy(manifest.get("inputs", [])),
        "artifacts": deepcopy(manifest.get("artifacts", [])),
        "metrics": deepcopy(manifest.get("metric_snapshot", manifest.get("metrics", []))),
        "note": note,
        "created_at_utc": datetime.now(UTC).isoformat(),
    }


def _upgrade_manifest_v2(root: Path, manifest_path: Path, manifest: dict[str, Any]) -> dict[str, Any]:
    if manifest.get("schema_version") == 2:
        return deepcopy(manifest)
    upgraded = deepcopy(manifest)
    experiment_root = manifest_path.parent
    primary_name = next((str(item.get("name")) for item in manifest.get("metrics", []) if isinstance(item, dict) and item.get("primary")), None)
    if not primary_name:
        primary_name = next((str(item.get("name")) for item in manifest.get("metrics", []) if isinstance(item, dict)), None)
    upgraded.update({
        "schema_version": 2,
        "run_mode": "candidate",
        "level": "candidate",
        "mode": "probe",
        "purpose": "candidate",
        "formal_candidate": True,
        "parent_run_id": None,
        "source_run_id": None,
        "source_manifest": None,
        "source_manifest_sha256": None,
        "evidence_scope": None,
        "checkpoint_id": None,
        "primary_metric": primary_name,
        "checks": {
            "input_output_match": True,
            "units_defined": all(bool(str(item.get("unit", ""))) for item in manifest.get("metrics", []) if isinstance(item, dict)),
            "core_constraints_passed": manifest.get("status") == "PASS",
            "deterministic": True,
            "baseline_comparable": any(isinstance(item, dict) and item.get("role") == "baseline" for item in manifest.get("methods", [])),
        },
        "arguments": [],
        "diagnostic_arguments": [],
        "reuse_contract": {"seed": False, "environment": False, "code": False, "inputs": False, "methods": False, "parameters": False},
        "reuse_validation": {"seed": True, "environment": True, "code": True, "inputs": True, "methods": True, "parameters": True, "source_manifest": True},
        "replay": {"required": False, "count": 1},
        "metric_snapshot": snapshot_metrics(root, experiment_root, list(manifest.get("metrics", [])), primary_name),
        "lifecycle": {
            "state": "QUICKCHECK",
            "formal": False,
            "updated_at_utc": datetime.now(UTC).isoformat(),
            "receipt": None,
            "promoted_at_utc": None,
            "archived_at_utc": None,
        },
    })
    return upgraded


def record_run(root: Path, config_path: Path, command: list[str], environment: dict[str, Any], started_at: str, duration: float, success: bool) -> dict[str, Any]:
    config = resolve_run_config(root, config_path)
    experiment_root = root / config["experiment_root"]
    artifacts = []
    if experiment_root.is_dir():
        for path in sorted(experiment_root.rglob("*")):
            if path.is_file() and path.name not in {"run_manifest.json", *LIFECYCLE_RECEIPT_NAMES.values()}:
                artifacts.append({"path": path.relative_to(root).as_posix(), "bytes": path.stat().st_size, "sha256": sha256(path)})
    inputs = []
    for item in config.get("inputs", []):
        path = root / str(item)
        inputs.append({"path": str(item), "sha256": sha256(path) if path.is_file() else None})
    if config["schema_version"] == 1:
        manifest = {
            "schema_version": 1,
            "run_id": config["experiment_id"],
            "problem_id": config["problem"],
            "question_id": config["question"],
            "engine": config["engine"],
            "command": command,
            "environment": environment,
            "code": {"runner": config["runner_path"], "sha256": sha256(root / config["runner_path"])},
            "random_seed": int(config["seed"]),
            "methods": config["methods"],
            "inputs": inputs,
            "artifacts": artifacts,
            "metrics": config["metrics"],
            "started_at_utc": started_at,
            "duration_seconds": round(duration, 6),
            "status": "PASS" if success else "FAIL",
        }
        manifest_path = experiment_root / "run_manifest.json"
        dump_json(manifest_path, manifest)
        question_path = root / "problems" / config["problem"] / "questions" / config["question"] / "question.yaml"
        if question_path.is_file():
            question_data = load_yaml(question_path)
            ref = relative_path(root, manifest_path)
            runs = question_data.setdefault("evidence", {}).setdefault("runs", [])
            if ref not in runs:
                runs.append(ref)
            if success:
                question_data["status"] = "RUN"
            dump_yaml(question_path, question_data)
        return manifest
    metric_snapshot = snapshot_metrics(root, experiment_root, list(config["metrics"]), config.get("primary_metric"))
    primary_metric = next((item["name"] for item in metric_snapshot if item.get("primary")), None)
    run_mode = str(config["run_mode"])
    source_manifest_ref = config.get("source_manifest")
    source_manifest_hash = config.get("source_manifest_sha256")
    reuse_validation = {"seed": True, "environment": True, "code": True, "inputs": True, "methods": True, "parameters": True, "source_manifest": True}
    if run_mode == "paper-evidence":
        source_path = _manifest_path_for_run(root, config["problem"], config["question"], str(config["source_run_id"]))
        source = load_json(source_path)
        lifecycle = source.get("lifecycle", {}) if isinstance(source.get("lifecycle"), dict) else {}
        actual_source_ref = relative_path(root, source_path)
        actual_source_hash = sha256(source_path)
        source_manifest_ref = actual_source_ref
        reuse_validation = {
            "seed": int(config["seed"]) == int(source.get("random_seed", -1)),
            "environment": environment == source.get("environment"),
            "code": sha256(root / config["runner_path"]) == source.get("code", {}).get("sha256"),
            "inputs": {(item.get("path"), item.get("sha256")) for item in inputs} == {(item.get("path"), item.get("sha256")) for item in source.get("inputs", []) if isinstance(item, dict)},
            "methods": config["methods"] == source.get("methods"),
            "parameters": config["arguments"] == source.get("arguments", []),
            "source_manifest": lifecycle.get("formal") is True and lifecycle.get("state") == "FORMAL" and actual_source_hash == source_manifest_hash and (not config.get("source_manifest") or config.get("source_manifest") == actual_source_ref),
        }
        source_manifest_hash = actual_source_hash
        success = success and all(reuse_validation.values())
    receipt_path = experiment_root / LIFECYCLE_RECEIPT_NAMES["quickcheck"]
    manifest = {
        "schema_version": 2,
        "run_id": config["experiment_id"],
        "problem_id": config["problem"],
        "question_id": config["question"],
        "engine": config["engine"],
        "command": command,
        "arguments": deepcopy(config["arguments"]),
        "diagnostic_arguments": deepcopy(config["diagnostic_arguments"]),
        "environment": environment,
        "code": {"runner": config["runner_path"], "sha256": sha256(root / config["runner_path"])},
        "random_seed": int(config["seed"]),
        "methods": config["methods"],
        "inputs": inputs,
        "artifacts": artifacts,
        "metrics": config["metrics"],
        "metric_snapshot": metric_snapshot,
        "started_at_utc": started_at,
        "duration_seconds": round(duration, 6),
        "status": "PASS" if success else "FAIL",
        "run_mode": run_mode,
        "level": run_mode,
        "mode": "probe",
        "purpose": config["purpose"],
        "formal_candidate": config["formal_candidate"],
        "parent_run_id": config.get("parent_run_id"),
        "source_run_id": config.get("source_run_id"),
        "source_manifest": source_manifest_ref,
        "source_manifest_sha256": source_manifest_hash,
        "evidence_scope": config.get("evidence_scope"),
        "checkpoint_id": config.get("checkpoint_id"),
        "primary_metric": primary_metric,
        "checks": deepcopy(config["checks"]),
        "reuse_contract": deepcopy(config["reuse_contract"]),
        "reuse_validation": reuse_validation,
        "replay": deepcopy(config["replay"]),
        "lifecycle": {
            "state": "QUICKCHECK",
            "formal": False,
            "updated_at_utc": datetime.now(UTC).isoformat(),
            "receipt": relative_path(root, receipt_path),
            "promoted_at_utc": None,
            "archived_at_utc": None,
        },
    }
    manifest_path = experiment_root / "run_manifest.json"
    dump_json(manifest_path, manifest)
    receipt = _probe_receipt(root, manifest_path, manifest, "quickcheck", manifest["status"], "non-formal run record")
    dump_json(receipt_path, receipt)
    return manifest


def quickcheck(root: Path, problem: str, question: str | None = None, strict: bool = False) -> dict[str, Any]:
    manifests = lifecycle_manifests(root, problem, question, levels=NONFORMAL_LEVELS)
    checks: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    for path in manifests:
        manifest = load_json(path)
        issues = _manifest_integrity_issues(
            root,
            manifest,
            require_baseline=False,
            verify_hashes=strict,
            verify_artifact_hashes=strict,
        )
        if strict and manifest.get("schema_version") != 2:
            issues.append("strict quickcheck requires run manifest schema v2")
        if manifest.get("schema_version") == 2:
            lifecycle_warnings = _lifecycle_check_issues(manifest, ("input_output_match", "units_defined", "core_constraints_passed"))
            if strict:
                issues.extend(lifecycle_warnings)
            elif lifecycle_warnings:
                add_warning(warnings, "quickcheck_contract_incomplete", "; ".join(lifecycle_warnings), relative_path(root, path))
        data_path = path.parent / "figure_data_manifest.yaml"
        intent_path = path.parent / "visual_intent.yaml"
        if data_path.is_file():
            visual_issues = _figure_data_manifest_issues(root, data_path, path)
            if visual_issues:
                add_warning(warnings, "quickcheck_visual_handoff", "; ".join(visual_issues), relative_path(root, data_path))
        if intent_path.is_file():
            visual_issues = _visual_intent_issues(root, intent_path, data_path)
            if visual_issues:
                add_warning(warnings, "quickcheck_visual_intent", "; ".join(visual_issues), relative_path(root, intent_path))
        run_warnings = [item["detail"] for item in warnings if item.get("evidence") == relative_path(root, path)]
        note = "; ".join(issues or run_warnings)
        receipt = _probe_receipt(root, path, manifest, "quickcheck", "PASS" if not issues else "FAIL", note)
        dump_json(path.parent / LIFECYCLE_RECEIPT_NAMES["quickcheck"], receipt)
        add_check(checks, "quickcheck_run", not issues, "; ".join(issues) if issues else str(manifest.get("run_id")), relative_path(root, path))
    for question_path in question_paths(root, problem):
        if question and question_path.parent.name != question:
            continue
        payload = load_yaml(question_path)
        if payload.get("schema_version") != 3:
            continue
        literature_issues = _existing_literature_issues(root, payload)
        if literature_issues:
            add_warning(warnings, "LITERATURE_INCOMPLETE", "; ".join(literature_issues), relative_path(root, question_path))
    add_check(checks, "quickcheck_scope", bool(manifests), f"found {len(manifests)} non-formal run(s)")
    passed = all(item["passed"] for item in checks)
    return {
        "schema_version": 1,
        "action": "quickcheck",
        "problem": problem,
        "question": question,
        "strict": strict,
        "passed": passed,
        "outcome": "BLOCK_TRANSITION" if not passed else "PASS_WITH_WARNINGS" if warnings else "PASS",
        "checks": checks,
        "warnings": warnings,
    }


def lifecycle_manifests(root: Path, problem: str, question: str | None = None, levels: set[str] | None = None) -> list[Path]:
    base = root / "experiments" / safe_token(problem, "problem")
    questions = [safe_token(question, "question")] if question else [path.name for path in base.glob("Q*") if path.is_dir()]
    found: list[Path] = []
    for question_id in sorted(questions):
        question_root = base / question_id
        for level in EXPERIMENT_LEVELS:
            if levels is not None and level not in levels:
                continue
            found.extend(sorted((question_root / level).glob("*/run_manifest.json")))
        if levels is None or "candidate" in levels:
            found.extend(sorted(path for path in question_root.glob("*/run_manifest.json") if path.parent.name not in EXPERIMENT_LEVELS))
    return sorted(set(found))


def checkpoint(root: Path, problem: str, question: str | None = None, strict: bool = False) -> dict[str, Any]:
    manifests = lifecycle_manifests(root, problem, question, levels=NONFORMAL_LEVELS)
    receipts: list[str] = []
    checks: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    for manifest_path in manifests:
        original = load_json(manifest_path)
        original_version = original.get("schema_version")
        manifest = _upgrade_manifest_v2(root, manifest_path, original)
        issues = _manifest_integrity_issues(
            root,
            manifest,
            require_baseline=True,
            verify_hashes=True,
            verify_artifact_hashes=strict,
        )
        issues.extend(_lifecycle_check_issues(manifest, ("input_output_match", "units_defined", "core_constraints_passed", "baseline_comparable")))
        determinism_issues = _lifecycle_check_issues(manifest, ("deterministic",))
        if strict:
            issues.extend(determinism_issues)
        elif determinism_issues:
            add_warning(
                warnings,
                "candidate_determinism_deferred",
                "; ".join(determinism_issues) + "; resolve before Formal promotion",
                relative_path(root, manifest_path),
            )
        if strict and original_version != 2:
            issues.append("strict checkpoint requires run manifest schema v2")
        replay = manifest.get("replay") if isinstance(manifest.get("replay"), dict) else {}
        if strict and (replay.get("required") is not True or int(replay.get("count", 0)) < 2):
            issues.append("strict checkpoint requires at least two deterministic replays")
        elif not strict and (replay.get("required") is not True or int(replay.get("count", 0)) < 2):
            add_warning(
                warnings,
                "candidate_replay_deferred",
                "independent deterministic replay is not complete; Formal G3 will require it",
                relative_path(root, manifest_path),
            )
        data_path = manifest_path.parent / "figure_data_manifest.yaml"
        intent_path = manifest_path.parent / "visual_intent.yaml"
        if data_path.is_file():
            visual_issues = _figure_data_manifest_issues(root, data_path, manifest_path)
            if visual_issues:
                add_warning(warnings, "checkpoint_visual_handoff", "; ".join(visual_issues), relative_path(root, data_path))
        if intent_path.is_file():
            visual_issues = _visual_intent_issues(root, intent_path, data_path)
            if visual_issues:
                add_warning(warnings, "checkpoint_visual_intent", "; ".join(visual_issues), relative_path(root, intent_path))
        question_path = root / "problems" / problem / "questions" / str(manifest.get("question_id")) / "question.yaml"
        question_data = load_yaml(question_path) if question_path.is_file() else {}
        paper = question_data.get("paper") if isinstance(question_data.get("paper"), dict) else {}
        declared_figures = {str(item) for item in paper.get("figure_ids", []) if item}
        declared_tables = {str(item) for item in paper.get("table_ids", []) if item}
        if manifest.get("run_mode") == "candidate" and declared_figures:
            if not data_path.is_file() or not intent_path.is_file():
                add_warning(warnings, "candidate_figure_design_deferred", "figure data or visual intent is not ready; complete it before G5", relative_path(root, manifest_path))
            for figure_id in sorted(declared_figures):
                brief_path = manifest_path.parent / "figure_briefs" / f"{figure_id}.yaml"
                if not brief_path.is_file():
                    add_warning(warnings, "candidate_figure_brief_deferred", f"candidate figure brief is missing: {figure_id}", relative_path(root, manifest_path))
                    continue
                brief = load_yaml(brief_path)
                brief_issues = _figure_brief_integrity_issues(root, brief, data_path, intent_path)
                if brief_issues:
                    add_warning(warnings, "candidate_figure_brief_incomplete", "; ".join(brief_issues), relative_path(root, brief_path))
                if brief.get("status") != "REVIEWED":
                    add_warning(warnings, "candidate_figure_brief_unreviewed", f"candidate figure brief is not reviewed: {figure_id}", relative_path(root, brief_path))
        if manifest.get("run_mode") == "candidate" and declared_tables and not declared_figures:
            if not intent_path.is_file():
                add_warning(warnings, "candidate_table_design_deferred", "table-only candidate has no visual intent decision", relative_path(root, manifest_path))
            else:
                intent = load_yaml(intent_path)
                if intent.get("artifact_decision") != "table":
                    add_warning(warnings, "candidate_table_design_mismatch", "table-only candidate visual intent should choose table", relative_path(root, intent_path))
        status_value = "PASS" if not issues else "FAIL"
        original_prefix = relative_path(root, manifest_path.parent)
        if status_value == "PASS" and manifest_path.parent.parent.name != "candidate":
            destination = root / "experiments" / str(manifest["problem_id"]) / str(manifest["question_id"]) / "candidate" / str(manifest["run_id"])
            if destination.exists():
                raise FileExistsError(f"candidate checkpoint already exists: {destination}")
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(manifest_path.parent), str(destination))
            manifest_path = destination / "run_manifest.json"
            destination_prefix = relative_path(root, destination)
            for artifact in manifest.get("artifacts", []):
                if isinstance(artifact, dict) and str(artifact.get("path", "")).startswith(original_prefix + "/"):
                    artifact["path"] = destination_prefix + str(artifact["path"])[len(original_prefix):]
            for collection in (manifest.get("metrics", []), manifest.get("metric_snapshot", [])):
                for metric in collection:
                    if not isinstance(metric, dict):
                        continue
                    locator = str(metric.get("locator", ""))
                    if locator.startswith(original_prefix + "/"):
                        metric["locator"] = destination_prefix + locator[len(original_prefix):]
        receipt_path = manifest_path.parent / LIFECYCLE_RECEIPT_NAMES["checkpoint"]
        manifest["run_mode"] = "candidate" if status_value == "PASS" else str(manifest.get("run_mode") or manifest.get("level") or "scratch")
        manifest["level"] = manifest["run_mode"]
        manifest["mode"] = "probe"
        manifest["purpose"] = "candidate" if status_value == "PASS" else str(manifest.get("purpose") or "exploration")
        manifest["formal_candidate"] = status_value == "PASS"
        manifest["checkpoint_id"] = f"{manifest['run_id']}-checkpoint"
        manifest["lifecycle"] = {
            "state": "CHECKPOINT",
            "formal": False,
            "updated_at_utc": datetime.now(UTC).isoformat(),
            "receipt": relative_path(root, receipt_path),
            "promoted_at_utc": None,
            "archived_at_utc": None,
        }
        dump_json(manifest_path, manifest)
        transition_prefix = original_prefix if original_prefix != relative_path(root, manifest_path.parent) else None
        _sync_visual_handoff_after_lifecycle_transition(root, manifest_path, manifest, transition_prefix)
        manifest_prefix = relative_path(root, manifest_path)
        run_warnings = [item["detail"] for item in warnings if item.get("evidence") in {manifest_prefix, relative_path(root, manifest_path.parent)}]
        receipt = _probe_receipt(root, manifest_path, manifest, "checkpoint", status_value, "; ".join(issues or run_warnings))
        dump_json(receipt_path, receipt)
        receipts.append(relative_path(root, receipt_path))
        add_check(checks, "checkpoint_run", not issues, "; ".join(issues) if issues else str(manifest["run_id"]), relative_path(root, manifest_path))
    add_check(checks, "checkpoint_scope", bool(manifests), f"found {len(manifests)} non-formal run(s)")
    passed = all(item["passed"] for item in checks)
    return {
        "schema_version": 1,
        "action": "checkpoint",
        "problem": problem,
        "question": question,
        "strict": strict,
        "passed": passed,
        "outcome": "BLOCK_TRANSITION" if not passed else "PASS_WITH_WARNINGS" if warnings else "PASS",
        "checks": checks,
        "warnings": warnings,
        "receipts": receipts,
    }


def promote(root: Path, problem: str, question: str, run_id: str) -> dict[str, Any]:
    source_path = _manifest_path_for_run(root, problem, question, run_id)
    manifest = _upgrade_manifest_v2(root, source_path, load_json(source_path))
    lifecycle = manifest.get("lifecycle", {}) if isinstance(manifest.get("lifecycle"), dict) else {}
    if lifecycle.get("state") == "ARCHIVED":
        raise ValueError("archived work cannot be promoted")
    issues = _manifest_integrity_issues(
        root,
        manifest,
        require_baseline=True,
        verify_hashes=True,
        verify_artifact_hashes=False,
    )
    issues.extend(_lifecycle_check_issues(manifest, ("input_output_match", "units_defined", "core_constraints_passed", "deterministic", "baseline_comparable")))
    receipt_path = source_path.parent / LIFECYCLE_RECEIPT_NAMES["checkpoint"]
    if manifest.get("run_mode") != "candidate" or lifecycle.get("state") != "CHECKPOINT" or manifest.get("formal_candidate") is not True:
        issues.append("run has not passed the candidate checkpoint")
    if not receipt_path.is_file():
        issues.append("candidate receipt is missing")
    else:
        receipt = load_json(receipt_path)
        if receipt.get("status") != "PASS" or receipt.get("formal") is not False:
            issues.append("candidate receipt did not pass")
    primary_name = str(manifest.get("primary_metric") or "")
    primary = next((item for item in manifest.get("metric_snapshot", []) if item.get("name") == primary_name), None)
    if not primary or primary.get("value") is None:
        issues.append("primary metric is missing or unresolved")
    if issues:
        raise ValueError("run is not promotion-ready: " + "; ".join(issues))
    question_path = root / "problems" / problem / "questions" / question / "question.yaml"
    if not question_path.is_file():
        raise FileNotFoundError(f"question manifest is missing: {question_path}")
    destination = root / "experiments" / problem / question / "formal" / run_id / "run_manifest.json"
    destination.parent.parent.mkdir(parents=True, exist_ok=True)
    if destination.resolve() != source_path.resolve() and destination.parent.exists():
        raise FileExistsError(f"formal run already exists: {destination}")
    promoted_at = datetime.now(UTC).isoformat()
    manifest["run_mode"] = "formal"
    manifest["level"] = "formal"
    manifest["mode"] = "formal"
    manifest["purpose"] = "candidate"
    manifest["formal_candidate"] = True
    manifest["lifecycle"] = {
        "state": "FORMAL",
        "formal": True,
        "updated_at_utc": promoted_at,
        "receipt": None,
        "promoted_at_utc": promoted_at,
        "archived_at_utc": None,
    }
    if destination.resolve() == source_path.resolve():
        dump_json(destination, manifest)
        _sync_visual_handoff_after_lifecycle_transition(root, destination, manifest)
    else:
        shutil.move(str(source_path.parent), str(destination.parent))
        source_prefix = relative_path(root, source_path.parent)
        destination_prefix = relative_path(root, destination.parent)
        for artifact in manifest.get("artifacts", []):
            if isinstance(artifact, dict) and str(artifact.get("path", "")).startswith(source_prefix + "/"):
                artifact["path"] = destination_prefix + str(artifact["path"])[len(source_prefix):]
        for collection in (manifest.get("metrics", []), manifest.get("metric_snapshot", [])):
            for metric in collection:
                if not isinstance(metric, dict):
                    continue
                locator = str(metric.get("locator", ""))
                if locator.startswith(source_prefix + "/"):
                    metric["locator"] = destination_prefix + locator[len(source_prefix):]
        dump_json(destination, manifest)
        _sync_visual_handoff_after_lifecycle_transition(root, destination, manifest, source_prefix)
    question_data = load_yaml(question_path)
    ref = relative_path(root, destination)
    runs = question_data.setdefault("evidence", {}).setdefault("runs", [])
    if ref not in runs:
        runs.append(ref)
    question_data["status"] = "RUN"
    dump_yaml(question_path, question_data)
    return {"schema_version": 1, "status": "FORMAL", "problem": problem, "question": question, "run_id": run_id, "manifest": ref, "question_evidence_updated": True}


def paper_evidence(root: Path, problem: str, question: str, config_path: Path, strict: bool = False) -> dict[str, Any]:
    if not path_is_within(config_path, root):
        raise ValueError(f"paper-evidence config must remain inside the selected project root: {config_path}")
    config = resolve_run_config(root, config_path)
    if config["problem"] != problem or config["question"] != question:
        raise ValueError("paper-evidence config does not match the selected problem/question")
    if config["run_mode"] != "paper-evidence":
        raise ValueError("paper-evidence config must use run_mode paper-evidence")
    source_run_id = str(config["source_run_id"])
    evidence_id = safe_token(str(config["experiment_id"]), "evidence-id")
    child_path = root / config["experiment_root"] / "run_manifest.json"
    if not child_path.is_file():
        raise FileNotFoundError("paper-evidence run must execute before evidence review")
    child = load_json(child_path)
    if child.get("run_mode") != "paper-evidence":
        raise ValueError("executed run is not marked as paper-evidence")
    source_path = _manifest_path_for_run(root, problem, question, source_run_id)
    source = load_json(source_path)
    lifecycle = source.get("lifecycle", {}) if isinstance(source.get("lifecycle"), dict) else {}
    if source.get("schema_version") != 2 or lifecycle.get("formal") is not True or lifecycle.get("state") != "FORMAL":
        raise ValueError("paper evidence requires a formal schema v2 source run")
    if strict and source.get("run_mode") != "formal":
        raise ValueError("strict paper evidence requires source run_mode formal")
    snapshot = deepcopy(source.get("metric_snapshot", []))
    current_snapshot = snapshot_metrics(root, source_path.parent, list(source.get("metrics", [])), source.get("primary_metric"))
    derived = deepcopy(child.get("metric_snapshot", []))
    primary_name = str(source.get("primary_metric") or next((item.get("name") for item in snapshot if item.get("primary")), ""))
    source_primary = next((item for item in snapshot if item.get("name") == primary_name), None)
    current_primary = next((item for item in current_snapshot if item.get("name") == primary_name), None)
    derived_primary = next((item for item in derived if item.get("name") == primary_name), None)
    reasons: list[str] = []
    reasons.extend("formal source " + issue for issue in _manifest_integrity_issues(root, source, require_baseline=True))
    reasons.extend("paper-evidence run " + issue for issue in _manifest_integrity_issues(root, child, require_baseline=True))
    reasons.extend("paper-evidence run " + issue for issue in _lifecycle_check_issues(child, ("input_output_match", "units_defined", "core_constraints_passed", "deterministic", "baseline_comparable")))
    actual_source_ref = relative_path(root, source_path)
    actual_source_hash = sha256(source_path)
    if config.get("source_manifest") and config.get("source_manifest") != actual_source_ref:
        reasons.append("configured source manifest path drifted")
    if config.get("source_manifest_sha256") != actual_source_hash:
        reasons.append("configured source manifest hash drifted")
    if child.get("source_manifest") != actual_source_ref or child.get("source_manifest_sha256") != actual_source_hash:
        reasons.append("recorded parent manifest provenance drifted")
    reuse_contract = child.get("reuse_contract") if isinstance(child.get("reuse_contract"), dict) else {}
    reuse_validation = child.get("reuse_validation") if isinstance(child.get("reuse_validation"), dict) else {}
    if not all(reuse_contract.get(field) is True for field in ("seed", "environment", "code", "inputs", "methods", "parameters")):
        reasons.append("paper-evidence reuse contract is incomplete")
    if not all(reuse_validation.get(field) is True for field in ("seed", "environment", "code", "inputs", "methods", "parameters", "source_manifest")):
        reasons.append("paper-evidence seed/environment/code/input/method/parameter reuse check failed")
    if child.get("evidence_scope") != config.get("evidence_scope"):
        reasons.append("paper-evidence scope drifted")
    if child.get("random_seed") != source.get("random_seed"):
        reasons.append("paper-evidence seed drifted")
    if child.get("environment") != source.get("environment"):
        reasons.append("paper-evidence environment drifted")
    if child.get("code", {}).get("sha256") != source.get("code", {}).get("sha256"):
        reasons.append("paper-evidence runner drifted")
    child_inputs = {(item.get("path"), item.get("sha256")) for item in child.get("inputs", []) if isinstance(item, dict)}
    source_inputs = {(item.get("path"), item.get("sha256")) for item in source.get("inputs", []) if isinstance(item, dict)}
    if child_inputs != source_inputs:
        reasons.append("paper-evidence inputs drifted")
    if child.get("methods") != source.get("methods"):
        reasons.append("paper-evidence methods drifted")
    if child.get("arguments", []) != source.get("arguments", []):
        reasons.append("paper-evidence solver/model parameters drifted")
    for metric, current in zip(derived, snapshot_metrics(root, child_path.parent, list(child.get("metrics", [])), child.get("primary_metric"))):
        if metric.get("name") != current.get("name") or metric.get("value") != current.get("value"):
            reasons.append(f"paper-evidence metric artifact drifted: {metric.get('name')}")
    if not source_primary or source_primary.get("value") is None:
        reasons.append("formal primary metric is missing")
        source_primary = {"name": primary_name, "value": None, "unit": "", "locator": ""}
    if not current_primary or current_primary.get("value") != source_primary.get("value"):
        reasons.append("formal source primary metric drifted")
    if not derived_primary or derived_primary.get("value") != source_primary.get("value") or derived_primary.get("unit") != source_primary.get("unit"):
        reasons.append("paper-evidence primary metric drifted")
    status_value = "REOPEN_REQUIRED" if reasons else "READY"
    hashes: list[dict[str, str]] = []
    for manifest in (source, child):
        code = manifest.get("code", {}) if isinstance(manifest.get("code"), dict) else {}
        if code.get("runner") and code.get("sha256"):
            hashes.append({"path": str(code["runner"]), "sha256": str(code["sha256"])})
    for item in [*source.get("inputs", []), *source.get("artifacts", []), *child.get("inputs", []), *child.get("artifacts", [])]:
        if isinstance(item, dict) and item.get("path") and item.get("sha256"):
            hashes.append({"path": str(item["path"]), "sha256": str(item["sha256"])})
    hashes = list({(item["path"], item["sha256"]): item for item in hashes}.values())
    payload = {
        "schema_version": 1,
        "evidence_id": evidence_id,
        "problem_id": problem,
        "question_id": question,
        "source_run_id": source_run_id,
        "source_manifest": relative_path(root, source_path),
        "source_manifest_sha256": actual_source_hash,
        "child_manifest": relative_path(root, child_path),
        "child_manifest_sha256": sha256(child_path),
        "source_status": "FORMAL",
        "evidence_scope": str(child.get("evidence_scope")),
        "reuse_contract": deepcopy(reuse_contract),
        "source_metric_snapshot": snapshot,
        "derived_metrics": derived,
        "source_hashes": hashes,
        "primary_metric": {"name": primary_name, "value": source_primary.get("value"), "unit": str(source_primary.get("unit", "")), "source_locator": str(source_primary.get("locator", ""))},
        "status": status_value,
        "reason": "; ".join(reasons) if reasons else "primary metric matches the formal source snapshot",
        "created_at_utc": datetime.now(UTC).isoformat(),
    }
    target = child_path.parent / "paper_evidence_manifest.json"
    dump_json(target, payload)
    target_ref = relative_path(root, target)
    if status_value == "READY":
        question_path = root / "problems" / problem / "questions" / question / "question.yaml"
        question_data = load_yaml(question_path)
        evidence_runs = question_data.setdefault("evidence", {}).setdefault("paper_evidence_runs", [])
        if target_ref not in evidence_runs:
            evidence_runs.append(target_ref)
        dump_yaml(question_path, question_data)
    payload["manifest"] = target_ref
    return payload


def _literature_template(root: Path, name: str) -> dict[str, Any]:
    candidates = (
        root / "templates" / "workflow" / name,
        Path(__file__).resolve().parents[2] / "templates" / "workflow" / name,
    )
    for path in candidates:
        if path.is_file():
            return load_yaml(path)
    raise FileNotFoundError(f"literature workflow template is missing: {name}")


def _literature_config(root: Path, path: Path | None, label: str) -> dict[str, Any]:
    if path is None:
        return {}
    resolved = path.resolve()
    if not path_is_within(resolved, root) or not resolved.is_file():
        raise ValueError(f"{label} must be a project-local file")
    return load_yaml(resolved)


def _slug(value: Any, fallback: str) -> str:
    slug = re.sub(r"[^a-z0-9._-]+", "-", str(value or "").strip().lower()).strip("-._")
    return slug or fallback


def _normalized_task_type(value: Any) -> str:
    raw = str(value or "").strip().lower()
    aliases = {
        "forecast": "prediction", "regression": "prediction", "预测": "prediction",
        "评价": "evaluation", "ranking": "evaluation", "assessment": "evaluation",
        "优化": "optimization", "decision": "optimization", "决策": "optimization",
        "分类": "classification", "判别": "classification",
        "调度": "scheduling", "排程": "scheduling",
        "机理": "mechanism", "physical": "mechanism",
        "聚类": "clustering", "统计": "statistical-inference", "network-science": "network",
    }
    if raw in {"prediction", "evaluation", "optimization", "classification", "scheduling", "mechanism", "clustering", "statistical-inference", "network", "other"}:
        return raw
    return next((mapped for token, mapped in aliases.items() if token in raw), "other")


def literature_plan(root: Path, problem: str, question: str, config: Path | None = None) -> dict[str, Any]:
    question_path, question_payload = _question_manifest(root, problem, question)
    if question_payload.get("schema_version") != 3:
        raise ValueError("literature-plan requires question schema v3")
    override = _literature_config(root, config, "literature-plan config")
    payload = _deep_merge(_literature_template(root, "literature_search_plan.yaml"), override)
    problem_data = question_payload.get("problem", {}) if isinstance(question_payload.get("problem"), dict) else {}
    target = str(problem_data.get("target") or f"Answer {question}").strip()
    inputs = [str(item) for item in problem_data.get("inputs", []) if str(item).strip()]
    outputs = [str(item) for item in problem_data.get("outputs", []) if str(item).strip()] or [target]
    constraints = [str(item) for item in problem_data.get("constraints", []) if str(item).strip()]
    conflicts = [str(item) for item in problem_data.get("key_conflicts", []) if str(item).strip()]
    task_type = _normalized_task_type(problem_data.get("type"))
    plan_id = f"litplan-{question.lower()}"
    query_context = " ".join([target, task_type, *constraints[:3]]).strip()
    payload.update({
        "schema_version": 1,
        "plan_id": plan_id,
        "problem_id": problem,
        "question_id": question,
        "source_question_manifest": relative_path(root, question_path),
        "source_question_manifest_sha256": question_interface_sha256(question_payload),
        "status": "PLAN_READY",
        "created_at_utc": datetime.now(UTC).isoformat(),
    })
    payload["domain_problem"] = _deep_merge(
        {
            "subject": target,
            "context": str(question_payload.get("source_problem") or problem),
            "research_question": target,
        },
        override.get("domain_problem", {}) if isinstance(override.get("domain_problem"), dict) else {},
    )
    payload["math_task"] = _deep_merge(
        {"type": task_type, "inputs": inputs, "outputs": outputs},
        override.get("math_task", {}) if isinstance(override.get("math_task"), dict) else {},
    )
    payload["key_constraints"] = override.get("key_constraints", [*constraints, *conflicts])
    if not override.get("queries"):
        payload["queries"] = [
            {
                "query_id": "query-scenario-task-zh",
                "kind": "scenario-task",
                "language": "zh",
                "query": query_context or f"{problem} {question} 建模",
                "rationale": "检索任务、输出和场景相近的学术方法。",
            },
            {
                "query_id": "query-method-constraint-en",
                "kind": "method-constraint",
                "language": "en",
                "query": " ".join([task_type, *constraints[:3], "model algorithm validation"]).strip(),
                "rationale": "Find methods addressing the dominant mathematical constraints and validation needs.",
            },
        ]
    _validate_literature_payload(root, "literature_search_plan.schema.json", payload)
    target_path = _literature_root(root, problem, question) / "search_plan.yaml"
    dump_yaml(target_path, payload)
    question_payload = _update_question_literature(
        root,
        question_path,
        search_plan=_hashed_ref(root, target_path),
        search_receipts=[],
        evidence_cards=[],
        model_evidence_brief=None,
        bib_keys=[],
        status="PLAN_READY",
    )
    return {
        "schema_version": 1,
        "status": "PLAN_READY",
        "problem": problem,
        "question": question,
        "search_plan": relative_path(root, target_path),
        "search_plan_sha256": sha256(target_path),
        "question_interface_sha256": question_interface_sha256(question_payload),
    }


def _publication_type(value: Any) -> str:
    raw = str(value or "").strip().lower()
    mapping = {
        "journal": "journal-article", "journal-article": "journal-article", "article-journal": "journal-article",
        "proceedings-article": "conference-paper", "conference": "conference-paper", "conference-paper": "conference-paper",
        "posted-content": "preprint", "preprint": "preprint", "report": "preprint",
        "dissertation": "thesis", "thesis": "thesis", "phdthesis": "thesis", "mastersthesis": "thesis",
    }
    return mapping.get(raw, "journal-article")


def _normalize_search_record(record: dict[str, Any], index: int) -> dict[str, Any]:
    doi_raw = record.get("doi") or record.get("DOI")
    doi = normalize_doi(doi_raw)
    if doi_raw not in (None, "") and doi is None:
        raise ValueError(f"invalid DOI in search result {index}: {doi_raw}")
    title = str(record.get("title") or "").strip()
    if not title:
        raise ValueError(f"search result {index} has no title")
    authors_raw = record.get("authors") or record.get("author") or []
    if isinstance(authors_raw, str):
        authors = [item.strip() for item in re.split(r"\s+and\s+|;", authors_raw) if item.strip()]
    else:
        authors = []
        for item in authors_raw if isinstance(authors_raw, list) else []:
            if isinstance(item, dict):
                name = item.get("full_name") or item.get("name") or " ".join(filter(None, (item.get("given"), item.get("family"))))
                if name:
                    authors.append(str(name).strip())
            elif str(item).strip():
                authors.append(str(item).strip())
    if not authors:
        raise ValueError(f"search result {index} has no authors")
    year_value = record.get("year")
    try:
        year = int(year_value) if year_value not in (None, "") else None
    except (TypeError, ValueError) as exc:
        raise ValueError(f"search result {index} has invalid year: {year_value}") from exc
    publication_type = _publication_type(record.get("publication_type") or record.get("type"))
    arxiv_id = record.get("arxiv_id") or record.get("arxivId")
    url = record.get("url") or (f"https://doi.org/{doi}" if doi else None)
    canonical = f"doi-{_slug(doi, 'unknown')}" if doi else f"title-{canonical_sha256({'title': title, 'authors': authors[:1], 'year': year})[:16]}"
    return {
        "canonical_id": canonical,
        "title": title,
        "authors": authors,
        "year": year,
        "publication_type": publication_type,
        "doi": doi,
        "arxiv_id": str(arxiv_id).strip() if arxiv_id else None,
        "url": str(url).strip() if url else None,
        "abstract_available": bool(record.get("abstract_available", record.get("abstract"))),
        "selected_for_screening": bool(record.get("selected_for_screening", True)),
        "exclusion_reason": record.get("exclusion_reason"),
        "venue": str(record.get("venue") or record.get("journal") or record.get("container_title") or "").strip(),
        "abstract": str(record.get("abstract") or "").strip(),
        "citation_count": int(record.get("citation_count", 0) or 0),
    }


def _http_json(url: str, headers: dict[str, str] | None = None) -> dict[str, Any]:
    request = urllib.request.Request(url, headers={"User-Agent": "math-modeling-workflow/6", **(headers or {})})
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            value = json.loads(response.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"literature provider request failed: {exc}") from exc
    return value if isinstance(value, dict) else {"results": value}


def _provider_search(provider: str, query: str, limit: int) -> tuple[list[dict[str, Any]], Any]:
    encoded = urllib.parse.quote(query)
    if provider == "crossref":
        raw = _http_json(f"https://api.crossref.org/works?query={encoded}&rows={limit}")
        items = raw.get("message", {}).get("items", [])
        records = []
        for item in items:
            title = (item.get("title") or [""])[0]
            year_parts = item.get("published-print", item.get("published-online", {})).get("date-parts", [[]])
            authors = [" ".join(filter(None, (author.get("given"), author.get("family")))) for author in item.get("author", [])]
            records.append({
                "title": title, "authors": authors, "year": year_parts[0][0] if year_parts and year_parts[0] else None,
                "type": item.get("type"), "doi": item.get("DOI"), "url": item.get("URL"),
                "venue": (item.get("container-title") or [""])[0], "abstract": item.get("abstract", ""),
            })
        return records, raw
    if provider == "openalex":
        raw = _http_json(f"https://api.openalex.org/works?search={encoded}&per-page={limit}")
        records = []
        for item in raw.get("results", []):
            records.append({
                "title": item.get("display_name"),
                "authors": [entry.get("author", {}).get("display_name") for entry in item.get("authorships", [])],
                "year": item.get("publication_year"), "type": item.get("type"),
                "doi": item.get("doi"), "url": item.get("primary_location", {}).get("landing_page_url"),
                "venue": item.get("primary_location", {}).get("source", {}).get("display_name"),
                "citation_count": item.get("cited_by_count", 0),
            })
        return records, raw
    if provider == "semantic-scholar":
        fields = "title,authors,year,venue,externalIds,url,abstract,citationCount"
        raw = _http_json(f"https://api.semanticscholar.org/graph/v1/paper/search?query={encoded}&limit={limit}&fields={fields}")
        records = []
        for item in raw.get("data", []):
            ids = item.get("externalIds") or {}
            records.append({
                "title": item.get("title"), "authors": [author.get("name") for author in item.get("authors", [])],
                "year": item.get("year"), "publication_type": "journal-article", "doi": ids.get("DOI"),
                "arxiv_id": ids.get("ArXiv"), "url": item.get("url"), "venue": item.get("venue"),
                "abstract": item.get("abstract"), "citation_count": item.get("citationCount", 0),
            })
        return records, raw
    if provider == "arxiv":
        request = urllib.request.Request(
            f"https://export.arxiv.org/api/query?search_query=all:{encoded}&start=0&max_results={limit}",
            headers={"User-Agent": "math-modeling-workflow/6"},
        )
        try:
            with urllib.request.urlopen(request, timeout=20) as response:
                content = response.read()
        except (urllib.error.URLError, TimeoutError) as exc:
            raise RuntimeError(f"literature provider request failed: {exc}") from exc
        root_xml = ET.fromstring(content)
        namespace = {"atom": "http://www.w3.org/2005/Atom"}
        records = []
        for entry in root_xml.findall("atom:entry", namespace):
            identifier = (entry.findtext("atom:id", default="", namespaces=namespace).rsplit("/", 1)[-1])
            published = entry.findtext("atom:published", default="", namespaces=namespace)
            records.append({
                "title": entry.findtext("atom:title", default="", namespaces=namespace).strip(),
                "authors": [author.findtext("atom:name", default="", namespaces=namespace) for author in entry.findall("atom:author", namespace)],
                "year": int(published[:4]) if published[:4].isdigit() else None, "publication_type": "preprint",
                "arxiv_id": identifier, "url": entry.findtext("atom:id", default=None, namespaces=namespace),
                "abstract": entry.findtext("atom:summary", default="", namespaces=namespace).strip(), "venue": "arXiv",
            })
        return records, {"content_type": "application/atom+xml", "xml": content.decode("utf-8", errors="replace")}
    raise ValueError(f"provider {provider!r} requires injected results or manual registration")


def _existing_search_records(root: Path, question_payload: dict[str, Any]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    literature = question_payload.get("literature", {}) if isinstance(question_payload.get("literature"), dict) else {}
    for index, ref in enumerate(literature.get("search_receipts", [])):
        path, issues = _current_hashed_ref(root, ref, f"literature.search_receipts[{index}]")
        if not issues and path is not None:
            records.extend(load_json(path).get("results", []))
    return records


def literature_search(root: Path, problem: str, question: str, config: Path) -> dict[str, Any]:
    question_path, question_payload = _question_manifest(root, problem, question)
    if question_payload.get("schema_version") != 3:
        raise ValueError("literature-search requires question schema v3")
    literature = question_payload.get("literature", {})
    plan_path, plan_issues = _current_hashed_ref(root, literature.get("search_plan"), "literature.search_plan")
    if plan_issues or plan_path is None:
        raise ValueError("literature search plan is missing or stale: " + "; ".join(plan_issues))
    plan = load_yaml(plan_path)
    _validate_literature_payload(root, "literature_search_plan.schema.json", plan)
    if plan.get("source_question_manifest_sha256") != question_interface_sha256(question_payload):
        raise ValueError("literature search plan is stale relative to the question interface")
    override = _literature_config(root, config, "literature-search config")
    provider = str(override.get("provider") or "crossref")
    if provider not in LITERATURE_PROVIDERS:
        raise ValueError(f"unsupported literature provider: {provider}")
    queries = {str(item.get("query_id")): item for item in plan.get("queries", []) if isinstance(item, dict)}
    query_id = str(override.get("query_id") or next(iter(queries), ""))
    if query_id not in queries and not override.get("query_text"):
        raise ValueError(f"query id is absent from the search plan: {query_id}")
    query_text = str(override.get("query_text") or queries[query_id].get("query") or "").strip()
    if len(query_text) < 3:
        raise ValueError("literature search query is too short")
    limit = min(int(override.get("limit", plan.get("screening", {}).get("max_candidates", 10)) or 10), 10)
    started = datetime.now(UTC)
    injected = override.get("results", override.get("records"))
    if injected is None:
        records_raw, provider_raw = _provider_search(provider, query_text, limit)
    else:
        if not isinstance(injected, list):
            raise ValueError("literature-search results must be an array")
        records_raw, provider_raw = injected, {"provider": provider, "query": query_text, "results": injected}
    normalized = [_normalize_search_record(item, index) for index, item in enumerate(records_raw, start=1) if isinstance(item, dict)]
    normalized = [item for item in normalized if _academic_source_record(item)]
    unique, merged = deduplicate_literature_records(normalized)
    existing = _existing_search_records(root, question_payload)
    for item in unique:
        duplicate = next((prior for prior in existing if _records_match(prior, item)), None)
        if duplicate is not None:
            item["selected_for_screening"] = False
            item["exclusion_reason"] = f"duplicate of {duplicate.get('canonical_id')}"
    schema_results = [{key: item.get(key) for key in (
        "canonical_id", "title", "authors", "year", "publication_type", "doi", "arxiv_id", "url",
        "abstract_available", "selected_for_screening", "exclusion_reason",
    )} for item in unique[:limit]]
    stamp = started.strftime("%Y%m%d%H%M%S")
    receipt_id = _slug(override.get("receipt_id"), f"litsearch-{question.lower()}-{_slug(provider, 'source')}-{stamp}")
    if not receipt_id.startswith("litsearch-"):
        receipt_id = "litsearch-" + receipt_id
    raw_path = _literature_cache_root(root, problem, question) / f"{receipt_id}.json"
    raw_path.parent.mkdir(parents=True, exist_ok=True)
    raw_path.write_text(json.dumps(provider_raw, ensure_ascii=False, indent=2, default=_json_default) + "\n", encoding="utf-8")
    finished = datetime.now(UTC)
    payload = {
        "schema_version": 1,
        "receipt_id": receipt_id,
        "problem_id": problem,
        "question_id": question,
        "search_plan": relative_path(root, plan_path),
        "search_plan_sha256": sha256(plan_path),
        "query_id": query_id if query_id.startswith("query-") else "query-follow-up",
        "query_text": query_text,
        "provider": provider,
        "started_at_utc": started.isoformat(),
        "finished_at_utc": finished.isoformat(),
        "duration_seconds": max(0.0, (finished - started).total_seconds()),
        "raw_results": {"path": relative_path(root, raw_path), "sha256": sha256(raw_path), "content_type": "application/json"},
        "results": schema_results,
        "deduplication": {
            "input_count": len(normalized), "unique_count": len(unique),
            "key_order": ["doi", "normalized-title-first-author-year"], "merged_records": merged,
        },
        "academic_sources_only": True,
        "status": "SOURCES_VERIFIED" if provider in {"crossref", "openalex", "arxiv", "semantic-scholar"} else "DISCOVERED",
    }
    _validate_literature_payload(root, "literature_search_receipt.schema.json", payload)
    target = _literature_root(root, problem, question) / "searches" / receipt_id / "search_receipt.json"
    dump_json(target, payload)
    receipt_refs = list(literature.get("search_receipts", []))
    receipt_refs = [item for item in receipt_refs if item.get("path") != relative_path(root, target)]
    receipt_refs.append(_hashed_ref(root, target))
    _update_question_literature(
        root,
        question_path,
        search_receipts=receipt_refs,
        status="SOURCES_VERIFIED" if all(load_json(workspace_path(root, item["path"], "search receipt")).get("status") == "SOURCES_VERIFIED" for item in receipt_refs) else "DISCOVERED",
    )
    return {"schema_version": 1, "status": payload["status"], "receipt": relative_path(root, target), "results": len(schema_results), "deduplicated": len(normalized) - len(unique)}


def literature_register(root: Path, problem: str, question: str, config: Path) -> dict[str, Any]:
    payload = _literature_config(root, config, "literature-register config")
    payload["provider"] = str(payload.get("provider") or "user-supplied")
    if payload["provider"] not in {"user-supplied", "manual-google-scholar", "manual-cnki", "manual-wanfang"}:
        raise ValueError("literature-register is restricted to user-supplied or manual database records")
    temporary = _literature_cache_root(root, problem, question) / f"register-{_slug(config.stem, 'source')}.yaml"
    dump_yaml(temporary, payload)
    return literature_search(root, problem, question, temporary)


def _find_receipt_record(root: Path, refs: list[Any], metadata: dict[str, Any]) -> dict[str, Any] | None:
    target_doi = normalize_doi(metadata.get("doi"))
    target_title = str(metadata.get("title") or "")
    target_authors = [item.get("full_name") for item in metadata.get("authors", []) if isinstance(item, dict)]
    target = {"doi": target_doi, "title": target_title, "authors": target_authors, "year": metadata.get("year")}
    for index, ref in enumerate(refs):
        path, issues = _current_hashed_ref(root, ref, f"source_search_receipts[{index}]")
        if issues or path is None:
            continue
        for record in load_json(path).get("results", []):
            if (target_doi and normalize_doi(record.get("doi")) == target_doi) or (target_title and _records_match(target, record)):
                return record
    return None


def _bibtex_key(metadata: dict[str, Any]) -> str:
    authors = metadata.get("authors", [])
    first = authors[0].get("full_name", "ref") if authors and isinstance(authors[0], dict) else "ref"
    surname = _slug(_first_author_key([first]), "ref").replace("-", "")
    year = str(metadata.get("year") or "nd")
    title_tokens = list(_title_tokens(metadata.get("title")))
    suffix = _slug(title_tokens[0] if title_tokens else "work", "work").replace("-", "")[:20]
    value = re.sub(r"[^A-Za-z0-9_:-]", "", f"{surname}{year}{suffix}")
    return value if value and value[0].isalpha() else "ref" + value


def literature_read(root: Path, problem: str, question: str, config: Path) -> dict[str, Any]:
    question_path, question_payload = _question_manifest(root, problem, question)
    if question_payload.get("schema_version") != 3:
        raise ValueError("literature-read requires question schema v3")
    override = _literature_config(root, config, "literature-read config")
    payload = _deep_merge(_literature_template(root, "literature_reference_card.yaml"), override)
    literature = question_payload.get("literature", {})
    refs = override.get("source_search_receipts") or literature.get("search_receipts", [])
    resolved_refs: list[dict[str, str]] = []
    for index, ref in enumerate(refs):
        path, issues = _current_hashed_ref(root, ref, f"source_search_receipts[{index}]")
        if issues or path is None:
            raise ValueError("reference card search receipt is stale: " + "; ".join(issues))
        resolved_refs.append(_hashed_ref(root, path))
    if not resolved_refs:
        raise ValueError("reference card requires at least one search receipt")
    metadata = payload.get("metadata", {}) if isinstance(payload.get("metadata"), dict) else {}
    matched = _find_receipt_record(root, resolved_refs, metadata)
    if matched is None and isinstance(override.get("metadata"), dict):
        raise ValueError("reference-card metadata conflicts with or is absent from the registered search receipts")
    if matched is None:
        for ref in resolved_refs:
            receipt = load_json(workspace_path(root, ref["path"], "source search receipt"))
            matched = next((item for item in receipt.get("results", []) if item.get("selected_for_screening") is True), None)
            if matched is not None:
                metadata = {}
                payload["metadata"] = metadata
                break
    if matched:
        metadata.setdefault("title", matched.get("title"))
        metadata.setdefault("year", matched.get("year"))
        metadata.setdefault("publication_type", matched.get("publication_type"))
        metadata.setdefault("venue", matched.get("venue") or "Academic source")
        metadata.setdefault("doi", matched.get("doi"))
        metadata.setdefault("arxiv_id", matched.get("arxiv_id"))
        metadata.setdefault("url", matched.get("url"))
        if not metadata.get("authors"):
            metadata["authors"] = [{"full_name": author, "orcid": None} for author in matched.get("authors", [])]
    normalized_authors: list[dict[str, Any]] = []
    for author in metadata.get("authors", []):
        if isinstance(author, dict):
            name = str(author.get("full_name") or author.get("name") or "").strip()
            orcid = author.get("orcid")
        else:
            name = str(author).strip()
            orcid = None
        if name:
            normalized_authors.append({"full_name": name, "orcid": orcid})
    metadata["authors"] = normalized_authors
    if not metadata["authors"]:
        raise ValueError("reference card requires at least one verified author")
    try:
        metadata["year"] = int(metadata.get("year"))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"reference card has an invalid year: {metadata.get('year')}") from exc
    metadata["venue"] = str(metadata.get("venue") or "Academic source").strip()
    configured_doi = metadata.get("doi")
    metadata["doi"] = normalize_doi(configured_doi)
    if configured_doi not in (None, "") and metadata["doi"] is None:
        raise ValueError(f"reference card contains an invalid DOI: {configured_doi}")
    metadata["publication_type"] = _publication_type(metadata.get("publication_type"))
    if not isinstance(override.get("metadata"), dict) or "bibtex_key" not in override["metadata"]:
        metadata["bibtex_key"] = _bibtex_key(metadata)
    else:
        metadata["bibtex_key"] = str(metadata.get("bibtex_key") or _bibtex_key(metadata))
    if not _academic_source_record({**metadata, "authors": [item.get("full_name") for item in metadata.get("authors", []) if isinstance(item, dict)]}):
        raise ValueError("reference card source is not an allowed academic publication or appears to be a contest paper")
    card_id = _slug(override.get("card_id"), f"litcard-{question.lower()}-{_slug(metadata.get('bibtex_key'), 'paper')}")
    if not card_id.startswith("litcard-"):
        card_id = "litcard-" + card_id
    metadata_sources = payload.get("metadata_sources", [])
    if not metadata_sources:
        metadata_sources = [{"provider": "user-supplied", "locator": str(metadata.get("doi") or metadata.get("url") or metadata.get("title")), "checked_at_utc": datetime.now(UTC).isoformat()}]
    provider_aliases = {"manual-google-scholar": "google-scholar", "manual-cnki": "cnki", "manual-wanfang": "wanfang"}
    for source in metadata_sources:
        if isinstance(source, dict):
            source["provider"] = provider_aliases.get(str(source.get("provider")), source.get("provider"))
    snapshot_path = _literature_cache_root(root, problem, question) / f"{card_id}-metadata.json"
    dump_json(snapshot_path, {"metadata": metadata, "metadata_sources": metadata_sources})
    source_document = payload.get("source_document", {}) if isinstance(payload.get("source_document"), dict) else {}
    if source_document.get("available") is True:
        source_value = source_document.get("path")
        if not source_value:
            raise ValueError("available source document requires a project-local path")
        source_path = workspace_path(root, str(source_value), "source document")
        if not source_path.is_file():
            raise FileNotFoundError(f"source document is missing: {source_path}")
        cache_path = _literature_cache_root(root, problem, question) / f"{card_id}{source_path.suffix.lower()}"
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        if cache_path.resolve() != source_path.resolve():
            shutil.copy2(source_path, cache_path)
        source_document["path"] = relative_path(root, cache_path)
        source_document["sha256"] = sha256(cache_path)
    else:
        source_document.update({"available": False, "kind": "none", "path": None, "sha256": None})
    depth = str(payload.get("review_depth") or "METADATA_ONLY")
    if depth not in LITERATURE_REVIEW_DEPTHS:
        raise ValueError(f"invalid literature review depth: {depth}")
    if depth in {"METADATA_ONLY", "ABSTRACT_SCREENED"}:
        payload["precision_evidence"] = []
        payload["substantive_citation_eligible"] = False
    payload.update({
        "schema_version": 1,
        "card_id": card_id,
        "problem_id": problem,
        "question_id": question,
        "source_search_receipts": resolved_refs,
        "metadata": metadata,
        "metadata_sources": metadata_sources,
        "metadata_snapshot": _hashed_ref(root, snapshot_path),
        "source_document": source_document,
        "review_depth": depth,
        "substantive_citation_eligible": depth in {"TARGETED_READ", "DEEP_READ"},
        "status": "CARD_READY",
        "created_at_utc": datetime.now(UTC).isoformat(),
    })
    payload.setdefault("paper_handoff", {})["bibtex_key"] = metadata["bibtex_key"]
    _validate_literature_payload(root, "academic_reference_card.schema.json", payload)
    target = _literature_root(root, problem, question) / "cards" / f"{card_id}.yaml"
    dump_yaml(target, payload)
    card_refs = list(literature.get("evidence_cards", []))
    card_refs = [item for item in card_refs if item.get("path") != relative_path(root, target)]
    card_refs.append(_hashed_ref(root, target))
    _update_question_literature(root, question_path, evidence_cards=card_refs, status="CARDS_READY")
    return {"schema_version": 1, "status": "CARD_READY", "card": relative_path(root, target), "review_depth": depth, "bibtex_key": metadata["bibtex_key"]}


def _bibtex_escape(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("&", r"\&")).strip()


def _bibtex_entry(card: dict[str, Any]) -> str:
    metadata = card["metadata"]
    publication_type = metadata["publication_type"]
    entry_type = {"journal-article": "article", "conference-paper": "inproceedings", "preprint": "misc", "thesis": "phdthesis"}[publication_type]
    fields: list[tuple[str, str]] = [
        ("author", " and ".join(_bibtex_escape(item["full_name"]) for item in metadata["authors"])),
        ("title", "{" + _bibtex_escape(metadata["title"]) + "}"),
        ("year", str(metadata["year"])),
    ]
    venue = _bibtex_escape(metadata.get("venue"))
    if entry_type == "article":
        fields.append(("journal", venue))
    elif entry_type == "inproceedings":
        fields.append(("booktitle", venue))
    elif entry_type == "phdthesis":
        fields.append(("school", venue))
    else:
        fields.append(("howpublished", venue or "Preprint"))
    if metadata.get("doi"):
        fields.append(("doi", str(metadata["doi"])))
    if metadata.get("url"):
        fields.append(("url", str(metadata["url"])))
    body = ",\n".join(f"  {key:<10}= {{{value}}}" for key, value in fields)
    return f"@{entry_type}{{{metadata['bibtex_key']},\n{body}\n}}"


def _write_managed_bibliography(root: Path, cards: list[dict[str, Any]], keys: list[str]) -> Path:
    path = root / "paper" / "references.bib"
    path.parent.mkdir(parents=True, exist_ok=True)
    existing = path.read_text(encoding="utf-8") if path.is_file() else ""
    pattern = re.compile(re.escape(LITERATURE_MANAGED_BIB_BEGIN) + r".*?" + re.escape(LITERATURE_MANAGED_BIB_END), re.S)
    all_cards = {card.get("metadata", {}).get("bibtex_key"): card for card in cards}
    all_keys = set(keys)
    for question_path in root.glob("problems/*/questions/Q*/question.yaml"):
        question_payload = load_yaml(question_path)
        if question_payload.get("schema_version") != 3:
            continue
        literature = question_payload.get("literature", {}) if isinstance(question_payload.get("literature"), dict) else {}
        all_keys.update(str(item) for item in literature.get("bib_keys", []) if item)
        for ref in literature.get("evidence_cards", []):
            card_path, issues = _current_hashed_ref(root, ref, "bibliography evidence card")
            if card_path is None or issues:
                continue
            card = load_yaml(card_path)
            key = card.get("metadata", {}).get("bibtex_key")
            if key:
                all_cards[str(key)] = card
    entries = [_bibtex_entry(all_cards[key]) for key in sorted(all_keys) if key in all_cards]
    block = LITERATURE_MANAGED_BIB_BEGIN + "\n" + "\n\n".join(entries) + "\n" + LITERATURE_MANAGED_BIB_END
    updated = pattern.sub(block, existing).strip() if pattern.search(existing) else (existing.strip() + "\n\n" + block).strip()
    path.write_text(updated + "\n", encoding="utf-8")
    return path


def literature_synthesize(root: Path, problem: str, question: str, config: Path) -> dict[str, Any]:
    question_path, question_payload = _question_manifest(root, problem, question)
    if question_payload.get("schema_version") != 3:
        raise ValueError("literature-synthesize requires question schema v3")
    override = _literature_config(root, config, "literature-synthesize config")
    payload = _deep_merge(_literature_template(root, "literature_model_evidence_brief.yaml"), override)
    literature = question_payload.get("literature", {})
    cards: list[dict[str, Any]] = []
    card_refs: list[dict[str, Any]] = []
    requested = override.get("source_cards") or literature.get("evidence_cards", [])
    for index, item in enumerate(requested):
        ref = item if isinstance(item, dict) and set(item) == {"path", "sha256"} else next(
            (candidate for candidate in literature.get("evidence_cards", []) if Path(str(candidate.get("path", ""))).stem == str(item) or str(candidate.get("path", "")).endswith(f"/{item}.yaml")),
            None,
        )
        path, issues = _current_hashed_ref(root, ref, f"source_cards[{index}]")
        if issues or path is None:
            raise ValueError("model evidence source card is stale: " + "; ".join(issues))
        card = load_yaml(path)
        _validate_literature_payload(root, "academic_reference_card.schema.json", card)
        cards.append(card)
        card_refs.append({
            "card_id": card["card_id"], "path": relative_path(root, path), "sha256": sha256(path),
            "review_depth": card["review_depth"], "bibliographic_status": card["bibliographic_status"],
            "bibtex_key": card["metadata"]["bibtex_key"],
        })
    if not cards:
        raise ValueError("model evidence synthesis requires at least one reference card")
    candidates = override.get("candidates")
    if not candidates:
        candidates = []
        for index, card in enumerate(cards, start=1):
            model_name = str(card.get("model", {}).get("name") or f"candidate {index}")
            candidates.append({
                "candidate_id": f"model-{index}", "model_name": model_name,
                "evidence_roles": {
                    "model_family": [card["card_id"]], "mechanism": [], "parameter": [], "validation": [card["card_id"]],
                },
                "fit_to_current_problem": {"matched_conditions": [card["problem_context"]["transferability_note"]], "mismatched_conditions": list(card.get("non_transferable_elements", []))},
                "implementation": {"expected_runtime": "to be measured in Scratch", "dependencies": [card.get("algorithm", {}).get("solver", "unspecified")]},
                "baseline": str(question_payload.get("model_selection", {}).get("baseline") or question_payload.get("method", {}).get("baseline", {}).get("name") or "same-output baseline"),
                "risk_probes": [item["objective"] for item in card.get("suggested_experiments", [])],
                "rejected_when": list(card.get("failure_conditions", [])) or ["reject when core constraints fail"],
                "decision": "retain" if index == 1 else "fallback",
            })
    payload["candidates"] = candidates
    recommendation = payload.get("recommendation", {})
    candidate_ids = {item["candidate_id"]: item for item in candidates}
    if recommendation.get("primary_candidate") not in candidate_ids:
        recommendation["primary_candidate"] = candidates[0]["candidate_id"]
    primary = candidate_ids[recommendation["primary_candidate"]]
    recommendation["baseline"] = str(recommendation.get("baseline") or primary["baseline"])
    payload["recommendation"] = recommendation
    eligible_cards = [card for card in cards if card.get("bibliographic_status") == "VERIFIED" and card.get("substantive_citation_eligible") is True]
    keys = list(dict.fromkeys(card["metadata"]["bibtex_key"] for card in eligible_cards))
    handoff = payload.get("citation_handoff", {})
    handoff["bibtex_keys"] = keys
    if not override.get("citation_handoff", {}).get("paper_targets"):
        handoff["paper_targets"] = [
            {
                "bibtex_key": card["metadata"]["bibtex_key"], "section": "model-selection",
                "purpose": card["paper_handoff"]["citation_note"], "minimum_review_depth": "TARGETED_READ",
            }
            for card in eligible_cards
        ]
    payload["citation_handoff"] = handoff
    high_conflict = any(item.get("severity") == "high" and item.get("resolution_status") == "unresolved" for item in payload.get("literature_conflicts", []))
    if any(card.get("bibliographic_status") == "CONFLICT" for card in cards) or high_conflict:
        payload["model_review_signal"] = "MODEL_REVIEW_SUGGESTED"
    payload.update({
        "schema_version": 1,
        "brief_id": _slug(override.get("brief_id"), f"litevidence-{question.lower()}"),
        "problem_id": problem,
        "question_id": question,
        "source_question_manifest": relative_path(root, question_path),
        "source_question_manifest_sha256": question_interface_sha256(question_payload),
        "source_cards": card_refs,
        "status": "SYNTHESIS_READY",
        "created_at_utc": datetime.now(UTC).isoformat(),
    })
    if not payload["brief_id"].startswith("litevidence-"):
        payload["brief_id"] = "litevidence-" + payload["brief_id"]
    _validate_literature_payload(root, "model_evidence_brief.schema.json", payload)
    target = _literature_root(root, problem, question) / "model_evidence_brief.yaml"
    dump_yaml(target, payload)
    bib_path = _write_managed_bibliography(root, cards, keys)
    _update_question_literature(
        root, question_path, model_evidence_brief=_hashed_ref(root, target), bib_keys=keys, status="SYNTHESIS_READY",
    )
    return {
        "schema_version": 1, "status": "SYNTHESIS_READY", "brief": relative_path(root, target),
        "bibtex": relative_path(root, bib_path), "bib_keys": keys, "model_review_signal": payload["model_review_signal"],
    }


def _tex_citation_keys(root: Path) -> set[str]:
    keys: set[str] = set()
    pattern = re.compile(r"\\(?:cite[a-zA-Z*]*|UpCite)\s*\{([^{}]+)\}")
    for path in sorted((root / "paper").rglob("*.tex")) if (root / "paper").is_dir() else []:
        for group in pattern.findall(path.read_text(encoding="utf-8", errors="replace")):
            keys.update(item.strip() for item in group.split(",") if item.strip())
    return keys


def _bibtex_keys(path: Path) -> set[str]:
    if not path.is_file():
        return set()
    return set(re.findall(r"@\w+\s*\{\s*([^,\s]+)", path.read_text(encoding="utf-8", errors="replace")))


def _literature_claim_source_issues(root: Path, problem: str, question: str | None = None) -> list[str]:
    issues: list[str] = []
    claims_path = root / "results" / problem / "claims.json"
    if claims_path.is_file():
        for claim in load_json(claims_path).get("claims", []):
            if question and claim.get("question_id") != question:
                continue
            try:
                path = locator_path(root, str(claim.get("locator", "")))
            except ValueError:
                continue
            if _is_literature_evidence_path(root, path):
                issues.append(f"claim {claim.get('id')} reads literature/cache evidence")
    contracts_path = root / "paper" / "figure_contracts.yaml"
    if contracts_path.is_file():
        for contract in load_yaml(contracts_path).get("figures", []):
            if question and contract.get("question_id") != question:
                continue
            values = list(contract.get("source_data", []))
            values.extend(item.get("locator") for item in contract.get("evidence_chain", []) if isinstance(item, dict))
            for value in values:
                try:
                    path = locator_path(root, str(value))
                except ValueError:
                    continue
                if _is_literature_evidence_path(root, path):
                    issues.append(f"figure {contract.get('id')} reads literature/cache evidence")
    return issues


def _card_integrity_issues(root: Path, path: Path, card: dict[str, Any]) -> list[str]:
    issues: list[str] = []
    try:
        _validate_literature_payload(root, "academic_reference_card.schema.json", card)
    except ValueError as exc:
        return [str(exc)]
    for index, ref in enumerate(card.get("source_search_receipts", [])):
        _, ref_issues = _current_hashed_ref(root, ref, f"card.source_search_receipts[{index}]")
        issues.extend(ref_issues)
    _, snapshot_issues = _current_hashed_ref(root, card.get("metadata_snapshot"), "card.metadata_snapshot")
    issues.extend(snapshot_issues)
    source = card.get("source_document", {})
    if source.get("available") is True:
        try:
            document = workspace_path(root, str(source.get("path")), "card source document")
            if not document.is_file() or sha256(document) != source.get("sha256"):
                issues.append("card source document is missing or its SHA-256 drifted")
        except ValueError as exc:
            issues.append(str(exc))
    if card.get("review_depth") in {"METADATA_ONLY", "ABSTRACT_SCREENED"} and card.get("precision_evidence"):
        issues.append("metadata/abstract-only card cannot support formulas, parameters, performance, or algorithm steps")
    if not _academic_source_record(card.get("metadata", {})):
        issues.append("card is not an allowed academic publication or appears to be a contest paper")
    return issues


def _existing_literature_issues(root: Path, payload: dict[str, Any]) -> list[str]:
    literature = payload.get("literature", {}) if isinstance(payload.get("literature"), dict) else {}
    issues: list[str] = []
    specifications = (
        ("search_plan", "literature_search_plan.schema.json", "yaml"),
        ("model_evidence_brief", "model_evidence_brief.schema.json", "yaml"),
    )
    for field, schema, kind in specifications:
        ref = literature.get(field)
        if ref is None:
            continue
        path, ref_issues = _current_hashed_ref(root, ref, f"literature.{field}")
        issues.extend(ref_issues)
        if path is not None and not ref_issues:
            try:
                _validate_literature_payload(root, schema, load_yaml(path) if kind == "yaml" else load_json(path))
            except ValueError as exc:
                issues.append(str(exc))
    for field, schema, kind in (
        ("search_receipts", "literature_search_receipt.schema.json", "json"),
        ("evidence_cards", "academic_reference_card.schema.json", "yaml"),
    ):
        for index, ref in enumerate(literature.get(field, [])):
            path, ref_issues = _current_hashed_ref(root, ref, f"literature.{field}[{index}]")
            issues.extend(ref_issues)
            if path is not None and not ref_issues:
                try:
                    _validate_literature_payload(root, schema, load_yaml(path) if kind == "yaml" else load_json(path))
                except ValueError as exc:
                    issues.append(str(exc))
    return issues


def literature_audit(root: Path, problem: str, question: str | None = None, strict: bool = True, write: bool = True) -> dict[str, Any]:
    paths = question_paths(root, problem)
    if question:
        paths = [path for path in paths if path.parent.name == question]
    checks: list[dict[str, Any]] = []
    for question_path in paths:
        payload = load_yaml(question_path)
        question_id = str(payload.get("question_id") or question_path.parent.name)
        if payload.get("schema_version") != 3:
            add_check(checks, "literature_v3_compatibility", True, f"{question_id}: legacy question schema; literature handoff not enforced", str(question_path))
            continue
        literature = payload.get("literature", {}) if isinstance(payload.get("literature"), dict) else {}
        plan_path, plan_ref_issues = _current_hashed_ref(root, literature.get("search_plan"), f"{question_id}.search_plan")
        plan_issues = list(plan_ref_issues)
        if plan_path is not None and not plan_ref_issues:
            plan = load_yaml(plan_path)
            try:
                _validate_literature_payload(root, "literature_search_plan.schema.json", plan)
            except ValueError as exc:
                plan_issues.append(str(exc))
            if plan.get("source_question_manifest_sha256") != question_interface_sha256(payload):
                plan_issues.append("search plan question-interface hash drifted")
        add_check(checks, "literature_plan_current", not plan_issues if strict else True, f"{question_id}: " + ("current" if not plan_issues else "; ".join(plan_issues)), str(question_path))

        receipt_paths: list[Path] = []
        receipt_issues: list[str] = []
        for index, ref in enumerate(literature.get("search_receipts", [])):
            receipt_path, issues = _current_hashed_ref(root, ref, f"{question_id}.search_receipts[{index}]")
            receipt_issues.extend(issues)
            if receipt_path is None or issues:
                continue
            receipt_paths.append(receipt_path)
            receipt = load_json(receipt_path)
            try:
                _validate_literature_payload(root, "literature_search_receipt.schema.json", receipt)
            except ValueError as exc:
                receipt_issues.append(str(exc))
            if plan_path is None or receipt.get("search_plan_sha256") != sha256(plan_path):
                receipt_issues.append(f"{receipt_path.name}: search plan hash drifted")
            raw_path, raw_issues = _current_hashed_ref(root, {"path": receipt.get("raw_results", {}).get("path"), "sha256": receipt.get("raw_results", {}).get("sha256")}, "raw_results")
            receipt_issues.extend(raw_issues)
            if raw_path is not None and not _is_literature_evidence_path(root, raw_path):
                receipt_issues.append("raw search results must remain in the ignored literature cache")
        receipt_ok = bool(receipt_paths) and not receipt_issues
        add_check(checks, "literature_sources_verified", receipt_ok if strict else True, f"{question_id}: receipts={len(receipt_paths)}; issues={receipt_issues or 'none'}", str(question_path))

        cards: list[dict[str, Any]] = []
        card_issues: list[str] = []
        for index, ref in enumerate(literature.get("evidence_cards", [])):
            card_path, issues = _current_hashed_ref(root, ref, f"{question_id}.evidence_cards[{index}]")
            card_issues.extend(issues)
            if card_path is None or issues:
                continue
            card = load_yaml(card_path)
            card_issues.extend(_card_integrity_issues(root, card_path, card))
            cards.append(card)
        substantive = [card for card in cards if card.get("bibliographic_status") == "VERIFIED" and card.get("substantive_citation_eligible") is True]
        cards_ok = bool(substantive) and not card_issues
        add_check(checks, "literature_cards_current", cards_ok if strict else True, f"{question_id}: substantive={len(substantive)}; issues={card_issues or 'none'}", str(question_path))

        brief_path, brief_ref_issues = _current_hashed_ref(root, literature.get("model_evidence_brief"), f"{question_id}.model_evidence_brief")
        brief_issues = list(brief_ref_issues)
        brief: dict[str, Any] = {}
        if brief_path is not None and not brief_ref_issues:
            brief = load_yaml(brief_path)
            try:
                _validate_literature_payload(root, "model_evidence_brief.schema.json", brief)
            except ValueError as exc:
                brief_issues.append(str(exc))
            if brief.get("source_question_manifest_sha256") != question_interface_sha256(payload):
                brief_issues.append("model evidence brief question-interface hash drifted")
            known_cards = {card.get("card_id"): card for card in cards}
            for item in brief.get("source_cards", []):
                path, issues = _current_hashed_ref(root, {"path": item.get("path"), "sha256": item.get("sha256")}, "brief.source_card")
                brief_issues.extend(issues)
                if path is not None and not issues and load_yaml(path).get("card_id") != item.get("card_id"):
                    brief_issues.append("brief source-card identity mismatch")
                if item.get("card_id") not in known_cards:
                    brief_issues.append(f"brief references undeclared card {item.get('card_id')}")
            primary_id = brief.get("recommendation", {}).get("primary_candidate")
            primary_name = next((item.get("model_name") for item in brief.get("candidates", []) if item.get("candidate_id") == primary_id), None)
            declared_primary = str(payload.get("model_selection", {}).get("primary") or "").strip()
            if not declared_primary or primary_name != declared_primary:
                brief_issues.append(f"brief primary model {primary_name!r} does not match question model_selection.primary {declared_primary!r}")
            unresolved = [item for item in brief.get("literature_conflicts", []) if item.get("severity") == "high" and item.get("resolution_status") == "unresolved"]
            if unresolved:
                brief_issues.append("high-severity literature conflict remains unresolved")
        add_check(checks, "literature_model_evidence_brief", not brief_issues if strict else True, f"{question_id}: " + ("current" if not brief_issues else "; ".join(brief_issues)), str(question_path))

        requested_keys = set(str(item) for item in literature.get("bib_keys", []) if item)
        brief_keys = set(str(item) for item in brief.get("citation_handoff", {}).get("bibtex_keys", []) if item)
        bib_path = root / "paper" / "references.bib"
        available_keys = _bibtex_keys(bib_path)
        cited_keys = _tex_citation_keys(root)
        citation_issues: list[str] = []
        if requested_keys != brief_keys:
            citation_issues.append(f"question keys {sorted(requested_keys)} differ from brief keys {sorted(brief_keys)}")
        missing_bib = sorted(requested_keys - available_keys)
        missing_citations = sorted(requested_keys - cited_keys)
        if missing_bib:
            citation_issues.append(f"missing BibTeX keys: {missing_bib}")
        if missing_citations:
            citation_issues.append(f"uncited literature keys: {missing_citations}")
        card_depth = {card["metadata"]["bibtex_key"]: card.get("review_depth") for card in cards}
        for target in brief.get("citation_handoff", {}).get("paper_targets", []):
            key = target.get("bibtex_key")
            if card_depth.get(key) not in {"TARGETED_READ", "DEEP_READ"}:
                citation_issues.append(f"{key} lacks TARGETED_READ/DEEP_READ support")
        add_check(checks, "literature_citation_handoff", bool(requested_keys) and not citation_issues if strict else True, f"{question_id}: keys={sorted(requested_keys)}; issues={citation_issues or 'none'}", str(bib_path))

        evidence_issues = _literature_claim_source_issues(root, problem, question_id)
        add_check(checks, "literature_not_project_evidence", not evidence_issues, f"{question_id}: {evidence_issues or 'project claims and figures use project experiments'}", str(question_path))
        question_passed = not any(not item["passed"] and question_id in str(item.get("detail", "")) for item in checks)
        if write:
            _update_question_literature(root, question_path, status="CITATION_READY" if question_passed and strict else ("STALE" if strict else literature.get("status", "NOT_STARTED")))

    passed = bool(checks) and all(item["passed"] for item in checks)
    report = {
        "schema_version": 1, "status": "PASS" if passed else "FAIL", "passed": passed,
        "problem": problem, "question": question, "strict": strict, "checks": checks,
        "generated_at_utc": datetime.now(UTC).isoformat(),
    }
    if write:
        dump_json(root / "output" / "literature_audit.json", report)
    return report


def _visual_schema_path(root: Path, name: str) -> Path:
    local = root / "config" / "schemas" / name
    if local.is_file():
        return local
    return Path(__file__).resolve().parents[2] / "config" / "schemas" / name


def _validate_visual_payload(root: Path, schema_name: str, payload: dict[str, Any]) -> None:
    from jsonschema import Draft202012Validator

    schema_path = _visual_schema_path(root, schema_name)
    if not schema_path.is_file():
        raise FileNotFoundError(f"visualization schema is missing: {schema_path}")
    validator = Draft202012Validator(load_json(schema_path))
    errors = sorted(validator.iter_errors(payload), key=lambda item: tuple(str(part) for part in item.absolute_path))
    if errors:
        details = []
        for error in errors[:8]:
            location = ".".join(str(item) for item in error.absolute_path) or "<root>"
            details.append(f"{location}: {error.message}")
        raise ValueError(f"{schema_name} validation failed: " + "; ".join(details))


def _nested_field_exists(value: Any, selector: str) -> bool:
    current = value
    for token in selector.removeprefix("$").strip(".").split("."):
        if not token:
            continue
        if isinstance(current, dict) and token in current:
            current = current[token]
        elif isinstance(current, list) and token.isdigit() and int(token) < len(current):
            current = current[int(token)]
        else:
            return False
    return True


def _artifact_shape_and_fields(path: Path, fields: list[str]) -> tuple[int | None, int | None, list[str]]:
    missing: list[str] = []
    suffix = path.suffix.lower()
    if suffix == ".json":
        value = load_json(path)
        missing = [field for field in fields if not _nested_field_exists(value, field)]
        if isinstance(value, list):
            columns = len(value[0]) if value and isinstance(value[0], dict) else None
            return len(value), columns, missing
        if isinstance(value, dict):
            return 1, len(value), missing
        return 1, 1, missing
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            header = reader.fieldnames or []
            row_count = sum(1 for _ in reader)
        missing = [field for field in fields if field not in header]
        return row_count, len(header), missing
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return None, None, []
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        header = [str(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) if value is not None]
        missing = [field for field in fields if field not in header]
        return max(sheet.max_row - 1, 0), len(header), missing
    return None, None, []


def _visual_run(root: Path, problem: str, question: str, run_id: str) -> tuple[Path, dict[str, Any], str]:
    manifest_path = _manifest_path_for_run(root, problem, question, run_id)
    manifest = _upgrade_manifest_v2(root, manifest_path, load_json(manifest_path))
    level = str(manifest.get("run_mode") or manifest.get("level") or manifest_path.parent.parent.name)
    if level not in EXPERIMENT_LEVELS:
        raise ValueError(f"unsupported visualization run level: {level}")
    return manifest_path, manifest, level


def _paper_evidence_is_ready(run_root: Path) -> bool:
    evidence_path = run_root / "paper_evidence_manifest.json"
    if not evidence_path.is_file():
        return False
    try:
        evidence = load_json(evidence_path)
        if evidence.get("status") != "READY":
            return False
        project_root = run_root
        while project_root.parent != project_root and project_root.name != "experiments":
            project_root = project_root.parent
        if project_root.name != "experiments":
            return False
        project_root = project_root.parent
        source = _current_visual_source(
            project_root,
            str(evidence.get("source_manifest", "")),
            str(evidence.get("source_manifest_sha256", "")),
            "paper-evidence parent Formal manifest",
        )
        child = _current_visual_source(
            project_root,
            str(evidence.get("child_manifest", "")),
            str(evidence.get("child_manifest_sha256", "")),
            "paper-evidence child manifest",
        )
        source_payload = load_json(source)
        child_payload = load_json(child)
        source_lifecycle = source_payload.get("lifecycle") if isinstance(source_payload.get("lifecycle"), dict) else {}
        if source_payload.get("run_mode") != "formal" or source_lifecycle.get("formal") is not True or source_lifecycle.get("state") != "FORMAL":
            return False
        if child_payload.get("run_mode") != "paper-evidence" or child_payload.get("status") != "PASS":
            return False
        for item in evidence.get("source_hashes", []):
            if not isinstance(item, dict):
                return False
            source_hash = _current_visual_source(
                project_root,
                str(item.get("path", "")),
                str(item.get("sha256", "")),
                "paper-evidence source",
            )
            if not source_hash.is_file():
                return False
        primary = evidence.get("primary_metric") if isinstance(evidence.get("primary_metric"), dict) else {}
        if primary.get("source_locator"):
            current_value = locator_value(project_root, str(primary["source_locator"]))
            if current_value != primary.get("value"):
                return False
    except (OSError, ValueError, KeyError, json.JSONDecodeError):
        return False
    return True


def _visual_evidence_eligible(level: str, manifest: dict[str, Any], run_root: Path) -> bool:
    if level == "formal":
        lifecycle = manifest.get("lifecycle") if isinstance(manifest.get("lifecycle"), dict) else {}
        return lifecycle.get("formal") is True and lifecycle.get("state") == "FORMAL"
    return level == "paper-evidence" and _paper_evidence_is_ready(run_root)


def _raise_visual_issues(message: str, issues: list[str], level: str) -> None:
    if not issues:
        return
    detail = message + ": " + "; ".join(issues)
    if level == "paper-evidence":
        raise ReopenRequiredError(detail)
    raise ValueError(detail)


def _visual_config(root: Path, path: Path, label: str) -> dict[str, Any]:
    resolved = path.resolve()
    if not path_is_within(resolved, root):
        raise ValueError(f"{label} config must remain inside the selected project root: {resolved}")
    if not resolved.is_file():
        raise FileNotFoundError(f"{label} config does not exist: {resolved}")
    return load_yaml(resolved)


def _current_visual_source(root: Path, path_value: str, hash_value: str, label: str) -> Path:
    path = workspace_path(root, path_value, label)
    if not path.is_file() or sha256(path) != hash_value:
        raise ValueError(f"{label} is missing or stale: {path_value}")
    return path


def _figure_data_manifest_issues(
    root: Path,
    data_path: Path,
    expected_manifest_path: Path | None = None,
) -> list[str]:
    issues: list[str] = []
    if not data_path.is_file():
        return ["figure data manifest is missing"]
    try:
        data = load_yaml(data_path)
        _validate_visual_payload(root, "figure_data_manifest.schema.json", data)
    except (OSError, ValueError) as exc:
        return [str(exc)]
    if data.get("status") != "DATA_READY":
        issues.append(f"figure data status is {data.get('status')}")
    try:
        source_manifest = _current_visual_source(
            root,
            str(data.get("source_run_manifest", "")),
            str(data.get("source_run_manifest_sha256", "")),
            "source run manifest",
        )
    except ValueError as exc:
        issues.append(str(exc))
        source_manifest = None
    if expected_manifest_path is not None and source_manifest is not None and source_manifest.resolve() != expected_manifest_path.resolve():
        issues.append("figure data references a different run manifest")
    declared_fields: set[str] = set()
    for item in data.get("source_artifacts", []):
        if not isinstance(item, dict):
            issues.append("figure data contains an invalid source artifact")
            continue
        declared_fields.update(str(field) for field in item.get("fields", []))
        try:
            source = workspace_path(root, str(item.get("path", "")), "figure-data source")
        except ValueError as exc:
            issues.append(str(exc))
            continue
        if not source.is_file() or item.get("sha256") != sha256(source):
            issues.append(f"figure-data source hash mismatch: {item.get('path')}")
            continue
        _, _, missing = _artifact_shape_and_fields(source, [str(field) for field in item.get("fields", [])])
        if missing:
            issues.append(f"figure-data fields are missing from {item.get('path')}: {missing}")
    profile = data.get("data_profile") if isinstance(data.get("data_profile"), dict) else {}
    variable_types = profile.get("variable_types") if isinstance(profile.get("variable_types"), dict) else {}
    units = profile.get("units") if isinstance(profile.get("units"), dict) else {}
    if declared_fields - {str(key) for key in variable_types}:
        issues.append("figure-data variable types do not cover every source field")
    if declared_fields - {str(key) for key in units}:
        issues.append("figure-data units do not cover every source field")
    level = str(data.get("level", ""))
    if level in NONFORMAL_LEVELS and data.get("contest_evidence_eligible") is not False:
        issues.append(f"{level} figure data cannot be contest evidence")
    if level in {"formal", "paper-evidence"} and data.get("contest_evidence_eligible") is not True:
        issues.append(f"{level} figure data is not eligible contest evidence")
    if level == "paper-evidence" and not _paper_evidence_is_ready(data_path.parent):
        issues.append("paper-evidence parent Formal or child run is stale")
    return issues


def _visual_intent_issues(root: Path, intent_path: Path, data_path: Path) -> list[str]:
    if not intent_path.is_file():
        return ["visual intent is missing"]
    try:
        intent = load_yaml(intent_path)
        _validate_visual_payload(root, "visual_intent.schema.json", intent)
        current_data = _current_visual_source(
            root,
            str(intent.get("source_data_manifest", "")),
            str(intent.get("source_data_manifest_sha256", "")),
            "figure data manifest",
        )
    except (OSError, ValueError) as exc:
        return [str(exc)]
    issues: list[str] = []
    if current_data.resolve() != data_path.resolve():
        issues.append("visual intent references a different figure data manifest")
    if intent.get("status") == "STALE":
        issues.append("visual intent is stale")
    return issues


def _question_visual_decision(
    root: Path,
    question: dict[str, Any],
    allowed: set[str],
) -> tuple[bool, str, str]:
    evidence = question.get("evidence") if isinstance(question.get("evidence"), dict) else {}
    detail = "missing current eligible visual intent"
    for run_ref in evidence.get("runs", []):
        try:
            manifest_path = workspace_path(root, str(run_ref), "question evidence run")
        except ValueError as exc:
            detail = str(exc)
            continue
        data_path = manifest_path.parent / "figure_data_manifest.yaml"
        intent_path = manifest_path.parent / "visual_intent.yaml"
        if not data_path.is_file() or not intent_path.is_file():
            continue
        issues = [
            *_figure_data_manifest_issues(root, data_path, manifest_path),
            *_visual_intent_issues(root, intent_path, data_path),
        ]
        data = load_yaml(data_path)
        intent = load_yaml(intent_path)
        decision = str(intent.get("artifact_decision", ""))
        if data.get("contest_evidence_eligible") is not True:
            issues.append("figure data is not eligible contest evidence")
        if intent.get("contest_evidence_eligible") is not True:
            issues.append("visual intent is not eligible contest evidence")
        if intent.get("status") != "READY":
            issues.append(f"visual intent status is {intent.get('status')}")
        if decision not in allowed:
            issues.append(f"decision={decision or 'missing'}")
        if not issues:
            return True, relative_path(root, intent_path), decision
        detail = "; ".join(issues)
    return False, detail, ""


def _figure_brief_integrity_issues(root: Path, brief: dict[str, Any], data_path: Path, intent_path: Path) -> list[str]:
    issues: list[str] = []
    try:
        _validate_visual_payload(root, "figure_brief.schema.json", brief)
        current_data = _current_visual_source(
            root,
            str(brief.get("source_data_manifest", "")),
            str(brief.get("source_data_manifest_sha256", "")),
            "figure data manifest",
        )
        current_intent = _current_visual_source(
            root,
            str(brief.get("visual_intent", "")),
            str(brief.get("visual_intent_sha256", "")),
            "visual intent",
        )
    except (OSError, ValueError) as exc:
        return [str(exc)]
    if current_data.resolve() != data_path.resolve():
        issues.append("figure brief references a different data manifest")
    if current_intent.resolve() != intent_path.resolve():
        issues.append("figure brief references a different visual intent")
    data = load_yaml(data_path)
    declared_source_fields = {
        str(item.get("path")): {str(field) for field in item.get("fields", [])}
        for item in data.get("source_artifacts", [])
        if isinstance(item, dict)
    }
    declared_sources = set(declared_source_fields)
    if not set(str(value) for value in brief.get("source_data", [])) <= declared_sources:
        issues.append("figure brief reads undeclared source data")
    try:
        source_script = workspace_path(root, str(brief.get("source_script", "")), "figure source script")
        if not source_script.is_file():
            issues.append("figure source script is missing")
        elif brief.get("source_script_sha256") != sha256(source_script):
            issues.append("figure source script hash is stale")
    except ValueError as exc:
        issues.append(str(exc))
    for item in brief.get("evidence_chain", []):
        if not isinstance(item, dict):
            issues.append("figure brief contains an invalid evidence record")
            continue
        try:
            source = locator_path(root, str(item.get("locator", "")))
        except ValueError as exc:
            issues.append(str(exc))
            continue
        if not source.is_file() or item.get("sha256") != sha256(source):
            issues.append(f"figure evidence hash mismatch: {item.get('locator')}")
        source_key = relative_path(root, source) if source.is_file() else ""
        if source_key not in declared_source_fields:
            issues.append(f"figure evidence reads an undeclared source: {item.get('locator')}")
        else:
            undeclared_fields = sorted({str(field) for field in item.get("fields", [])} - declared_source_fields[source_key])
            if undeclared_fields:
                issues.append(f"figure evidence uses undeclared fields: {undeclared_fields}")
    integrity = brief.get("data_integrity") if isinstance(brief.get("data_integrity"), dict) else {}
    for item in integrity.get("source_hashes", []):
        if not isinstance(item, dict):
            issues.append("figure data integrity contains an invalid source record")
            continue
        try:
            source = workspace_path(root, str(item.get("path", "")), "figure integrity source")
        except ValueError as exc:
            issues.append(str(exc))
            continue
        if not source.is_file() or item.get("sha256") != sha256(source):
            issues.append(f"figure integrity hash mismatch: {item.get('path')}")
        if str(item.get("path")) not in declared_sources:
            issues.append(f"figure integrity reads an undeclared source: {item.get('path')}")
    if brief.get("status") == "STALE":
        issues.append("figure brief is stale")
    return issues


def _decision_reference_exists(root: Path, problem: str, question: str, decision_id: str) -> bool:
    def contains(value: Any) -> bool:
        if isinstance(value, dict):
            if str(value.get("id") or value.get("decision_id") or "") == decision_id:
                return True
            return any(contains(item) for item in value.values())
        if isinstance(value, list):
            return any(contains(item) for item in value)
        return False

    state_path = root / "state" / "decision_log.json"
    question_path = root / "problems" / problem / "questions" / question / "question.yaml"
    return (state_path.is_file() and contains(load_json(state_path))) or (question_path.is_file() and contains(load_yaml(question_path)))


def figure_data(root: Path, problem: str, question: str, run_id: str, config_path: Path) -> dict[str, Any]:
    manifest_path, run_manifest, level = _visual_run(root, problem, question, run_id)
    if level == "paper-evidence" and not _paper_evidence_is_ready(manifest_path.parent):
        raise ReopenRequiredError("paper-evidence parent Formal or child run is stale; local G3/G4 reopen is required")
    config = _visual_config(root, config_path, "figure-data")
    if str(config.get("problem_id") or problem) != problem or str(config.get("question_id") or question) != question:
        raise ValueError("figure-data config does not match the selected problem/question")
    run_root = manifest_path.parent
    recorded: dict[str, str] = {}
    for item in [*run_manifest.get("inputs", []), *run_manifest.get("artifacts", [])]:
        if not isinstance(item, dict) or not item.get("path") or not item.get("sha256"):
            continue
        try:
            recorded_path = relative_path(root, workspace_path(root, str(item["path"]), "run artifact"))
        except ValueError:
            continue
        recorded[recorded_path] = str(item["sha256"])
    source_artifacts: list[dict[str, Any]] = []
    observed_rows: list[int] = []
    observed_columns: list[int] = []
    for item in config.get("source_artifacts", []):
        if not isinstance(item, dict):
            raise ValueError("figure-data source_artifacts must contain objects")
        raw_path_value = str(item.get("path", ""))
        fields = [str(field) for field in item.get("fields", []) if str(field).strip()]
        source = workspace_path(root, raw_path_value, "figure-data source")
        path_value = relative_path(root, source)
        if not source.is_file() or not fields:
            raise ValueError(f"figure-data source or fields are missing: {path_value}")
        current_hash = sha256(source)
        if _visual_evidence_eligible(level, run_manifest, run_root) and recorded.get(path_value) != current_hash:
            raise ValueError(f"formal figure data is not registered in the run manifest: {path_value}")
        rows, columns, missing = _artifact_shape_and_fields(source, fields)
        if missing:
            raise ValueError(f"figure-data fields do not exist in {path_value}: {', '.join(missing)}")
        if rows is not None:
            observed_rows.append(rows)
        if columns is not None:
            observed_columns.append(columns)
        source_artifacts.append({"path": path_value, "sha256": current_hash, "fields": fields})
    profile = deepcopy(config.get("data_profile", {}))
    if observed_rows and int(profile.get("row_count", -1)) != max(observed_rows):
        raise ValueError(f"data_profile.row_count does not match observed data: {profile.get('row_count')} != {max(observed_rows)}")
    if observed_columns and int(profile.get("column_count", -1)) < max(observed_columns):
        raise ValueError("data_profile.column_count is smaller than an observed source artifact")
    declared_fields = {
        field for item in source_artifacts for field in item.get("fields", []) if isinstance(field, str)
    }
    variable_types = profile.get("variable_types") if isinstance(profile.get("variable_types"), dict) else {}
    units = profile.get("units") if isinstance(profile.get("units"), dict) else {}
    missing_types = sorted(declared_fields - {str(key) for key in variable_types})
    missing_units = sorted(declared_fields - {str(key) for key in units})
    if missing_types or missing_units:
        raise ValueError(
            "data_profile must define every source field; "
            f"missing variable types={missing_types or 'none'}, units={missing_units or 'none'}"
        )
    eligible = _visual_evidence_eligible(level, run_manifest, run_root)
    payload = {
        "schema_version": 1,
        "manifest_id": str(config.get("manifest_id") or f"figdata-{run_id}"),
        "problem_id": problem,
        "question_id": question,
        "run_id": run_id,
        "level": level,
        "status": "DATA_READY",
        "source_run_manifest": relative_path(root, manifest_path),
        "source_run_manifest_sha256": sha256(manifest_path),
        "source_artifacts": source_artifacts,
        "data_profile": profile,
        "comparators": deepcopy(config.get("comparators", {})),
        "uncertainty": deepcopy(config.get("uncertainty", {})),
        "read_only_transformations": list(config.get("read_only_transformations", [])),
        "reader_question": str(config.get("reader_question", "")),
        "claim_candidates": list(config.get("claim_candidates", [])),
        "paper_targets": list(config.get("paper_targets", [])),
        "contest_evidence_eligible": eligible,
        "created_at_utc": datetime.now(UTC).isoformat(),
    }
    _validate_visual_payload(root, "figure_data_manifest.schema.json", payload)
    target = run_root / "figure_data_manifest.yaml"
    dump_yaml(target, payload)
    return {"schema_version": 1, "status": "DATA_READY", "manifest": relative_path(root, target), "contest_evidence_eligible": eligible}


def figure_intent(root: Path, problem: str, question: str, run_id: str, config_path: Path) -> dict[str, Any]:
    manifest_path, run_manifest, level = _visual_run(root, problem, question, run_id)
    config = _visual_config(root, config_path, "figure-intent")
    data_path = manifest_path.parent / "figure_data_manifest.yaml"
    if not data_path.is_file():
        raise FileNotFoundError("figure-data must run before figure-intent")
    data = load_yaml(data_path)
    data_issues = _figure_data_manifest_issues(root, data_path, manifest_path)
    _raise_visual_issues("figure data is stale or invalid", data_issues, level)
    eligible = bool(data.get("contest_evidence_eligible")) and _visual_evidence_eligible(level, run_manifest, manifest_path.parent)
    payload = {
        "schema_version": 1,
        "intent_id": str(config.get("intent_id") or f"intent-{run_id}"),
        "question_id": question,
        "run_id": run_id,
        "source_data_manifest": relative_path(root, data_path),
        "source_data_manifest_sha256": sha256(data_path),
        "reader_question": str(config.get("reader_question") or data.get("reader_question") or ""),
        "evidence_role": str(config.get("evidence_role") or "exploratory"),
        "artifact_decision": str(config.get("artifact_decision") or "figure"),
        "candidate_archetypes": deepcopy(config.get("candidate_archetypes", [])),
        "required_encodings": deepcopy(config.get("required_encodings", {})),
        "comparison": str(config.get("comparison") or data.get("comparators", {}).get("baseline") or ""),
        "risks": list(config.get("risks", [])),
        "paper_slot": str(config.get("paper_slot") or next(iter(data.get("paper_targets", [])), "")),
        "status": str(config.get("status") or "READY"),
        "contest_evidence_eligible": eligible,
        "created_at_utc": datetime.now(UTC).isoformat(),
    }
    if level == "scratch" and payload["contest_evidence_eligible"]:
        raise ValueError("Scratch visual intent cannot be contest evidence")
    _validate_visual_payload(root, "visual_intent.schema.json", payload)
    target = manifest_path.parent / "visual_intent.yaml"
    dump_yaml(target, payload)
    return {"schema_version": 1, "status": "INTENT_READY", "intent": relative_path(root, target), "artifact_decision": payload["artifact_decision"]}


def figure_brief(root: Path, problem: str, question: str, run_id: str, intent_path: Path, config_path: Path) -> dict[str, Any]:
    manifest_path, run_manifest, level = _visual_run(root, problem, question, run_id)
    if level == "scratch":
        raise ValueError("Scratch visualization stops at visual-intent; promote the run before creating a figure brief")
    config = _visual_config(root, config_path, "figure-brief")
    intent_resolved = intent_path.resolve()
    if not path_is_within(intent_resolved, root) or not intent_resolved.is_file():
        raise ValueError("figure-brief intent must be a project-local file")
    intent = load_yaml(intent_resolved)
    _validate_visual_payload(root, "visual_intent.schema.json", intent)
    if intent.get("question_id") != question or intent.get("run_id") != run_id or intent.get("artifact_decision") != "figure":
        raise ValueError("figure-brief requires a matching figure visual intent")
    data_path = _current_visual_source(root, str(intent["source_data_manifest"]), str(intent["source_data_manifest_sha256"]), "figure data manifest")
    data = load_yaml(data_path)
    data_issues = _figure_data_manifest_issues(root, data_path, manifest_path)
    _raise_visual_issues("figure data is stale or invalid", data_issues, level)
    eligible = bool(data.get("contest_evidence_eligible")) and _visual_evidence_eligible(level, run_manifest, manifest_path.parent)
    status_value = str(config.get("status") or "REVIEWED")
    if level in NONFORMAL_LEVELS and status_value not in {"DRAFT", "REVIEWED"}:
        raise ValueError("Scratch/Candidate figure briefs cannot be approved")
    approval = deepcopy(config.get("approval")) if isinstance(config.get("approval"), dict) else None
    if status_value in {"APPROVED", "RENDERED", "QA_PASSED", "CONTRACT_READY"}:
        if not eligible:
            raise ValueError("only Formal or READY Paper Evidence can approve a figure brief")
        decision_id = str((approval or {}).get("decision_id", ""))
        if not decision_id or not _decision_reference_exists(root, problem, question, decision_id):
            raise ValueError("approved figure brief must reference an existing root decision")
    payload = deepcopy(config)
    payload.update({
        "schema_version": 1,
        "brief_id": str(config.get("brief_id", "")),
        "question_id": question,
        "run_id": run_id,
        "source_data_manifest": relative_path(root, data_path),
        "source_data_manifest_sha256": sha256(data_path),
        "visual_intent": relative_path(root, intent_resolved),
        "visual_intent_sha256": sha256(intent_resolved),
        "status": status_value,
        "contest_evidence_eligible": eligible,
        "created_at_utc": datetime.now(UTC).isoformat(),
    })
    if approval is None:
        payload.pop("approval", None)
    normalized_source_data: list[str] = []
    for value in payload.get("source_data", []):
        source = workspace_path(root, str(value), "figure brief source data")
        if not source.is_file():
            raise FileNotFoundError(f"figure brief source data does not exist: {value}")
        normalized_source_data.append(relative_path(root, source))
    declared_sources = {str(item.get("path")) for item in data.get("source_artifacts", []) if isinstance(item, dict)}
    undeclared_sources = sorted(set(normalized_source_data) - declared_sources)
    if undeclared_sources:
        raise ValueError(f"figure brief reads data not declared by the manifest: {undeclared_sources}")
    payload["source_data"] = normalized_source_data
    source_script = workspace_path(root, str(payload.get("source_script", "")), "figure source script")
    if not source_script.is_file():
        raise FileNotFoundError(f"figure source script does not exist: {payload.get('source_script')}")
    payload["source_script"] = relative_path(root, source_script)
    payload["source_script_sha256"] = sha256(source_script)
    _validate_visual_payload(root, "figure_brief.schema.json", payload)
    for item in payload.get("evidence_chain", []):
        evidence = locator_path(root, str(item.get("locator", "")))
        if not evidence.is_file() or sha256(evidence) != item.get("sha256"):
            raise ValueError(f"figure brief evidence is missing or stale: {item.get('locator')}")
    for item in payload.get("data_integrity", {}).get("source_hashes", []):
        source = workspace_path(root, str(item.get("path", "")), "figure integrity source")
        if not source.is_file() or sha256(source) != item.get("sha256"):
            raise ValueError(f"figure brief integrity source is missing or stale: {item.get('path')}")
    target = manifest_path.parent / "figure_briefs" / f"{payload['brief_id']}.yaml"
    dump_yaml(target, payload)
    phase = "DESIGN_APPROVED" if status_value == "APPROVED" else "BRIEF_READY"
    return {"schema_version": 1, "status": phase, "brief": relative_path(root, target), "brief_status": status_value}


def _render_command_tokens(command: list[Any], root: Path, brief_path: Path, output_dir: Path, data_path: Path) -> list[str]:
    replacements = {
        "{python}": sys.executable,
        "{root}": str(root),
        "{brief}": str(brief_path),
        "{output_dir}": str(output_dir),
        "{data_manifest}": str(data_path),
    }
    return [replacements.get(str(item), str(item)) for item in command]


def _render_command_uses_source_script(root: Path, command: list[str], source_script: Path) -> bool:
    for token in command[1:]:
        if token.startswith("-"):
            continue
        candidate = Path(token)
        resolved = candidate.resolve() if candidate.is_absolute() else (root / candidate).resolve()
        if resolved == source_script.resolve():
            return True
    return False


def _render_protected_snapshot(root: Path, staging_root: Path) -> dict[str, str]:
    snapshot: dict[str, str] = {}
    for path in root.rglob("*"):
        if not path.is_file() or path.is_symlink():
            continue
        relative = path.relative_to(root)
        if relative.parts and relative.parts[0] == ".git":
            continue
        if path_is_within(path.resolve(), staging_root.resolve()):
            continue
        snapshot[relative.as_posix()] = sha256(path)
    return snapshot


def figure_render(root: Path, problem: str, question: str, run_id: str, brief_path: Path) -> dict[str, Any]:
    manifest_path, _, level = _visual_run(root, problem, question, run_id)
    brief_resolved = brief_path.resolve()
    if not path_is_within(brief_resolved, manifest_path.parent) or not brief_resolved.is_file():
        raise ValueError("figure-render brief must belong to the selected run")
    brief = load_yaml(brief_resolved)
    _validate_visual_payload(root, "figure_brief.schema.json", brief)
    allowed = {"REVIEWED"} if level in NONFORMAL_LEVELS else {"APPROVED", "RENDERED"}
    if brief.get("status") not in allowed:
        raise ValueError(f"figure brief status is not renderable: {brief.get('status')}")
    data_path = _current_visual_source(root, str(brief["source_data_manifest"]), str(brief["source_data_manifest_sha256"]), "figure data manifest")
    intent_path = _current_visual_source(root, str(brief["visual_intent"]), str(brief["visual_intent_sha256"]), "visual intent")
    data_issues = _figure_data_manifest_issues(root, data_path, manifest_path)
    brief_issues = _figure_brief_integrity_issues(root, brief, data_path, intent_path)
    _raise_visual_issues("figure design is stale or invalid", [*data_issues, *brief_issues], level)
    stage_root = manifest_path.parent / "figure-staging" / str(brief["brief_id"])
    output_dir = stage_root / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_paths: dict[str, Path] = {}
    for value in brief.get("outputs", {}).values():
        if isinstance(value, str):
            output = workspace_path(root, value, "figure output")
            if not path_is_within(output, output_dir):
                raise ValueError("figure-render outputs must remain in the run-local figure-staging directory")
            output_paths[output.suffix.lower().lstrip(".")] = output
            if output.exists():
                if not output.is_file():
                    raise ValueError(f"figure-render output path is not a file: {relative_path(root, output)}")
                output.unlink()
    command = brief.get("render_command")
    if not isinstance(command, list) or not command:
        raise ValueError("figure brief must declare a non-shell render_command")
    resolved_command = _render_command_tokens(command, root, brief_resolved, output_dir, data_path)
    source_script = workspace_path(root, str(brief["source_script"]), "figure source script")
    shell_names = {"cmd", "cmd.exe", "powershell", "powershell.exe", "pwsh", "pwsh.exe", "bash", "bash.exe", "sh", "sh.exe"}
    if Path(resolved_command[0]).name.lower() in shell_names:
        raise ValueError("figure render_command may not invoke a general-purpose shell")
    if not _render_command_uses_source_script(root, resolved_command, source_script):
        raise ValueError("figure render_command must execute the declared source_script")
    environment = os.environ.copy()
    environment.update({
        "MATHMODEL_PROJECT_ROOT": str(root),
        "MATHMODEL_FIGURE_BRIEF": str(brief_resolved),
        "MATHMODEL_FIGURE_DATA_MANIFEST": str(data_path),
        "MATHMODEL_FIGURE_OUTPUT_DIR": str(output_dir),
    })
    protected_before = _render_protected_snapshot(root, stage_root)
    completed = subprocess.run(resolved_command, cwd=root, env=environment, capture_output=True, text=True, check=False)
    protected_after = _render_protected_snapshot(root, stage_root)
    if protected_after != protected_before:
        changed = sorted(
            key for key in set(protected_before) | set(protected_after)
            if protected_before.get(key) != protected_after.get(key)
        )
        raise RuntimeError("figure renderer modified files outside its staging directory: " + ", ".join(changed[:20]))
    if completed.returncode != 0:
        raise RuntimeError(f"figure render failed ({completed.returncode}): {completed.stderr.strip() or completed.stdout.strip()}")
    post_data_issues = _figure_data_manifest_issues(root, data_path, manifest_path)
    post_brief_issues = _figure_brief_integrity_issues(root, brief, data_path, intent_path)
    if post_data_issues or post_brief_issues:
        raise RuntimeError("figure renderer modified declared evidence or plotting code: " + "; ".join([*post_data_issues, *post_brief_issues]))
    required_formats = ("png",) if level in NONFORMAL_LEVELS else ("pdf", "svg", "png")
    missing = [key for key in required_formats if not output_paths.get(key, Path("__missing__")).is_file()]
    if missing:
        raise FileNotFoundError("figure render did not create: " + ", ".join(missing))
    rendered_status = "PREVIEW_RENDERED" if level in NONFORMAL_LEVELS else "RENDERED"
    if level not in NONFORMAL_LEVELS:
        brief["status"] = "RENDERED"
        dump_yaml(brief_resolved, brief)
    rendered_outputs = {
        key: str(value)
        for key, value in brief["outputs"].items()
        if key == "png_dpi" or (isinstance(value, str) and workspace_path(root, value, "figure output").is_file())
    }
    receipt = {
        "schema_version": 1,
        "status": rendered_status,
        "brief": relative_path(root, brief_resolved),
        "brief_sha256": sha256(brief_resolved),
        "command": resolved_command,
        "stdout": completed.stdout[-4000:],
        "outputs": rendered_outputs,
        "created_at_utc": datetime.now(UTC).isoformat(),
    }
    receipt_path = stage_root / "render_receipt.json"
    dump_json(receipt_path, receipt)
    return {"schema_version": 1, "status": rendered_status, "receipt": relative_path(root, receipt_path)}


def _png_dpi(path: Path) -> float | None:
    try:
        from PIL import Image
    except ImportError:
        return None
    with Image.open(path) as image:
        dpi = image.info.get("dpi")
    if isinstance(dpi, tuple) and dpi:
        return float(dpi[0])
    return float(dpi) if isinstance(dpi, (int, float)) else None


def _png_dimensions(path: Path) -> tuple[int, int] | None:
    try:
        from PIL import Image
    except ImportError:
        return None
    try:
        with Image.open(path) as image:
            image.verify()
        with Image.open(path) as image:
            return int(image.width), int(image.height)
    except OSError:
        return None


def figure_qa(root: Path, problem: str, question: str, run_id: str, brief_path: Path, outputs_dir: Path | None = None) -> dict[str, Any]:
    manifest_path, _, level = _visual_run(root, problem, question, run_id)
    brief_resolved = brief_path.resolve()
    if not path_is_within(brief_resolved, manifest_path.parent) or not brief_resolved.is_file():
        raise ValueError("figure-qa brief must belong to the selected run")
    brief = load_yaml(brief_resolved)
    _validate_visual_payload(root, "figure_brief.schema.json", brief)
    if brief.get("status") != "RENDERED":
        raise ValueError("figure-qa requires a rendered brief")
    data_path = _current_visual_source(root, str(brief["source_data_manifest"]), str(brief["source_data_manifest_sha256"]), "figure data manifest")
    intent_path = _current_visual_source(root, str(brief["visual_intent"]), str(brief["visual_intent_sha256"]), "visual intent")
    data_issues = _figure_data_manifest_issues(root, data_path, manifest_path)
    brief_issues = _figure_brief_integrity_issues(root, brief, data_path, intent_path)
    _raise_visual_issues("figure design is stale or invalid", [*data_issues, *brief_issues], level)
    stage_root = manifest_path.parent / "figure-staging" / str(brief["brief_id"])
    expected_outputs = stage_root / "outputs"
    if outputs_dir is not None and outputs_dir.resolve() != expected_outputs.resolve():
        raise ValueError("figure-qa outputs directory does not match the brief staging directory")
    checks: list[dict[str, Any]] = []
    output_records: list[dict[str, Any]] = []
    for key, suffix in (("pdf", ".pdf"), ("svg", ".svg"), ("png", ".png")):
        path = workspace_path(root, str(brief["outputs"][key]), "figure output")
        passed = path.is_file() and path.suffix.lower() == suffix and path.stat().st_size > 0
        add_check(checks, f"figure_{key}_exists", passed, relative_path(root, path) if path.exists() else str(path))
        if passed:
            output_records.append({"format": key, "path": relative_path(root, path), "sha256": sha256(path), "bytes": path.stat().st_size})
    svg_path = workspace_path(root, str(brief["outputs"]["svg"]), "figure svg")
    svg_text = svg_path.read_text(encoding="utf-8", errors="ignore").lower() if svg_path.is_file() else ""
    editable_svg = bool(svg_text) and "<text" in svg_text and "<image" not in svg_text
    add_check(checks, "svg_editable_text", editable_svg, relative_path(root, svg_path) if svg_path.exists() else str(svg_path))
    allowed_colors = set()
    style_path = root / "config" / "figure_style.yaml"
    if not style_path.is_file():
        style_path = Path(__file__).resolve().parents[2] / "config" / "figure_style.yaml"
    if style_path.is_file():
        style = load_yaml(style_path)
        allowed_colors.update(str(value).lower() for value in style.get("categorical_order", []))
        allowed_colors.update(str(value).lower() for value in style.get("colors", {}).values())
        allowed_colors.update(str(value).lower() for value in style.get("rules", {}).get("allowed_derived_colors", []))
    svg_colors = {value.lower() for value in re.findall(r"#[0-9a-fA-F]{6}", svg_text)}
    add_check(checks, "registered_figure_colors", not (svg_colors - allowed_colors), f"unregistered={sorted(svg_colors - allowed_colors)}")
    png_path = workspace_path(root, str(brief["outputs"]["png"]), "figure png")
    dpi = _png_dpi(png_path) if png_path.is_file() else None
    dpi_ok = brief.get("outputs", {}).get("png_dpi") == 400 and dpi is not None and 395 <= dpi <= 405
    add_check(checks, "png_400_dpi", dpi_ok, f"declared=400, observed={dpi if dpi is not None else 'unavailable'}")
    dimensions = _png_dimensions(png_path) if png_path.is_file() else None
    expected_width_px = float(brief.get("final_width_mm", 0)) / 25.4 * 400
    dimension_ok = dimensions is not None and expected_width_px > 0 and abs(dimensions[0] - expected_width_px) / expected_width_px <= 0.08
    add_check(checks, "png_final_width", dimension_ok, f"observed={dimensions}, expected_width_px={expected_width_px:.1f}")
    pdf_path = workspace_path(root, str(brief["outputs"]["pdf"]), "figure pdf")
    pdf_valid = pdf_path.is_file() and pdf_path.read_bytes()[:5] == b"%PDF-"
    add_check(checks, "pdf_signature", pdf_valid, relative_path(root, pdf_path) if pdf_path.exists() else str(pdf_path))
    collision_checked = isinstance(brief.get("label_strategy"), dict) and brief["label_strategy"].get("collision_checked") is True
    add_check(checks, "label_collision_review", collision_checked, "brief.label_strategy.collision_checked must be true after rendered review")
    eligible = bool(brief.get("contest_evidence_eligible"))
    add_check(checks, "formal_figure_evidence", eligible, f"level={level}")
    passed = all(item["passed"] for item in checks)
    if passed:
        brief["status"] = "QA_PASSED"
        dump_yaml(brief_resolved, brief)
    payload = {
        "schema_version": 1,
        "figure_id": brief.get("brief_id"),
        "problem_id": problem,
        "question_id": question,
        "run_id": run_id,
        "status": "QA_PASSED" if passed else "FAIL",
        "passed": passed,
        "brief": relative_path(root, brief_resolved),
        "brief_sha256": sha256(brief_resolved),
        "data_manifest": relative_path(root, data_path),
        "data_manifest_sha256": sha256(data_path),
        "visual_intent": relative_path(root, intent_path),
        "visual_intent_sha256": sha256(intent_path),
        "checks": checks,
        "outputs": output_records,
        "created_at_utc": datetime.now(UTC).isoformat(),
    }
    target = stage_root / "figure_qa.json"
    dump_json(target, payload)
    payload["qa"] = relative_path(root, target)
    return payload


def figure_promote(root: Path, problem: str, question: str, figure_id: str, brief_path: Path, qa_path: Path) -> dict[str, Any]:
    figure_id = safe_token(figure_id, "figure-id")
    brief_resolved = brief_path.resolve()
    qa_resolved = qa_path.resolve()
    if not path_is_within(brief_resolved, root) or not path_is_within(qa_resolved, root):
        raise ValueError("figure promotion inputs must remain inside the selected project")
    brief = load_yaml(brief_resolved)
    qa = load_json(qa_resolved)
    _validate_visual_payload(root, "figure_brief.schema.json", brief)
    if figure_id != brief.get("brief_id") or brief.get("question_id") != question or brief.get("status") != "QA_PASSED":
        raise ValueError("figure brief is not QA-passed for the selected figure/question")
    if qa.get("passed") is not True or qa.get("status") != "QA_PASSED" or qa.get("brief_sha256") != sha256(brief_resolved):
        raise ValueError("figure QA is missing, failed, or stale")
    if not brief.get("contest_evidence_eligible") or not isinstance(brief.get("approval"), dict):
        raise ValueError("only an approved Formal or READY Paper Evidence brief can be promoted")
    decision_id = str(brief["approval"].get("decision_id", ""))
    if not _decision_reference_exists(root, problem, question, decision_id):
        raise ValueError("figure promotion approval decision is missing")
    data_path = _current_visual_source(root, str(brief["source_data_manifest"]), str(brief["source_data_manifest_sha256"]), "figure data manifest")
    intent_path = _current_visual_source(root, str(brief["visual_intent"]), str(brief["visual_intent_sha256"]), "visual intent")
    manifest_path, _, _ = _visual_run(root, problem, question, str(brief.get("run_id", "")))
    data_issues = _figure_data_manifest_issues(root, data_path, manifest_path)
    brief_issues = _figure_brief_integrity_issues(root, brief, data_path, intent_path)
    if data_issues or brief_issues:
        raise ValueError("figure design is stale or invalid: " + "; ".join([*data_issues, *brief_issues]))
    if qa.get("figure_id") != figure_id or qa.get("question_id") != question or qa.get("run_id") != brief.get("run_id"):
        raise ValueError("figure QA does not belong to the selected figure/question/run")
    if qa.get("data_manifest_sha256") != sha256(data_path) or qa.get("visual_intent_sha256") != sha256(intent_path):
        raise ValueError("figure QA references stale design inputs")
    qa_output_hashes = {
        str(item.get("path")): str(item.get("sha256"))
        for item in qa.get("outputs", [])
        if isinstance(item, dict) and item.get("path") and item.get("sha256")
    }
    for key in ("pdf", "svg", "png"):
        source = workspace_path(root, str(brief["outputs"][key]), "figure output")
        if not source.is_file() or qa_output_hashes.get(relative_path(root, source)) != sha256(source):
            raise ValueError(f"figure QA output is missing or stale: {key}")
    claims_path, claims = load_claims(root, problem)
    frozen_ids = {str(item.get("id")) for item in claims.get("claims", []) if item.get("question_id") == question and item.get("status") == "frozen"}
    if not brief.get("claim_id") or str(brief.get("claim_id")) not in frozen_ids:
        raise ValueError("figure promotion requires a frozen claim from the same question")
    paper_figures = root / "paper" / "figures"
    paper_figures.mkdir(parents=True, exist_ok=True)
    final_outputs: dict[str, Any] = {"png_dpi": 400}
    for key in ("pdf", "svg", "png"):
        source = workspace_path(root, str(brief["outputs"][key]), "figure output")
        destination = paper_figures / f"{figure_id}.{key}"
        shutil.copy2(source, destination)
        final_outputs[key] = relative_path(root, destination)
    contract = {field: deepcopy(brief.get(field)) for field in FIGURE_BRIEF_CONTRACT_FIELDS}
    contract.update({
        "contract_version": "2.0",
        "id": figure_id,
        "question_id": question,
        "kind": "data",
        "archetype": str(brief.get("decision", {}).get("archetype", "")),
        "outputs": final_outputs,
        "design_handoff": {
            "data_manifest": relative_path(root, data_path),
            "data_manifest_sha256": sha256(data_path),
            "visual_intent": relative_path(root, intent_path),
            "visual_intent_sha256": sha256(intent_path),
            "figure_brief": relative_path(root, brief_resolved),
            "figure_brief_sha256": sha256(brief_resolved),
            "design_status": "APPROVED",
            "render_qa": relative_path(root, qa_resolved),
            "render_qa_sha256": sha256(qa_resolved),
        },
    })
    if brief.get("multipanel_justification"):
        contract["multipanel_justification"] = brief["multipanel_justification"]
    contracts_path = root / "paper" / "figure_contracts.yaml"
    manifest = load_yaml(contracts_path) if contracts_path.is_file() else {"schema_version": "2.0", "figures": []}
    figures = manifest.setdefault("figures", [])
    if not isinstance(figures, list):
        raise ValueError("paper/figure_contracts.yaml figures must be a list")
    collisions = [item for item in figures if isinstance(item, dict) and item.get("id") == figure_id and item.get("question_id") != question]
    if collisions:
        raise ValueError("figure id is already owned by another question")
    figures[:] = [item for item in figures if not (isinstance(item, dict) and item.get("id") == figure_id)]
    figures.append(contract)
    dump_yaml(contracts_path, manifest)
    question_path = root / "problems" / problem / "questions" / question / "question.yaml"
    question_data = load_yaml(question_path)
    evidence_figures = question_data.setdefault("evidence", {}).setdefault("figures", [])
    paper_figures_ids = question_data.setdefault("paper", {}).setdefault("figure_ids", [])
    for collection in (evidence_figures, paper_figures_ids):
        if figure_id not in collection:
            collection.append(figure_id)
    dump_yaml(question_path, question_data)
    return {"schema_version": 1, "status": "CONTRACT_READY", "figure_id": figure_id, "contract": relative_path(root, contracts_path), "outputs": final_outputs}


def archive_work(root: Path, problem: str, question: str | None = None) -> dict[str, Any]:
    manifests = lifecycle_manifests(root, problem, question, levels={"scratch"})
    archive_root = root / "output" / "_archive" / "experiments" / problem
    archived: list[str] = []
    for manifest_path in manifests:
        manifest = _upgrade_manifest_v2(root, manifest_path, load_json(manifest_path))
        if manifest.get("lifecycle", {}).get("formal") is True:
            continue
        timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
        destination = archive_root / str(manifest["question_id"]) / f"{manifest['run_id']}-{timestamp}"
        destination.parent.mkdir(parents=True, exist_ok=True)
        if destination.exists():
            raise FileExistsError(f"archive destination already exists: {destination}")
        original_prefix = relative_path(root, manifest_path.parent)
        shutil.move(str(manifest_path.parent), str(destination))
        archived_manifest = destination / "run_manifest.json"
        destination_prefix = relative_path(root, destination)
        for artifact in manifest.get("artifacts", []):
            if isinstance(artifact, dict) and str(artifact.get("path", "")).startswith(original_prefix + "/"):
                artifact["path"] = destination_prefix + str(artifact["path"])[len(original_prefix):]
        for collection in (manifest.get("metrics", []), manifest.get("metric_snapshot", [])):
            for metric in collection:
                if not isinstance(metric, dict):
                    continue
                locator = str(metric.get("locator", ""))
                if locator.startswith(original_prefix + "/"):
                    metric["locator"] = destination_prefix + locator[len(original_prefix):]
        archived_at = datetime.now(UTC).isoformat()
        receipt_path = destination / LIFECYCLE_RECEIPT_NAMES["archive-work"]
        manifest["run_mode"] = "scratch"
        manifest["level"] = "scratch"
        manifest["mode"] = "probe"
        manifest["lifecycle"] = {
            "state": "ARCHIVED",
            "formal": False,
            "updated_at_utc": archived_at,
            "receipt": relative_path(root, receipt_path),
            "promoted_at_utc": None,
            "archived_at_utc": archived_at,
        }
        dump_json(archived_manifest, manifest)
        receipt = _probe_receipt(root, archived_manifest, manifest, "archive-work", "ARCHIVED", "scratch work archived outside formal evidence")
        dump_json(receipt_path, receipt)
        archived.append(relative_path(root, destination))
    return {"schema_version": 1, "action": "archive-work", "problem": problem, "question": question, "status": "ARCHIVED", "archived": archived, "count": len(archived)}


def sprint_path(root: Path, sprint_id: str) -> Path:
    return root / "sprints" / safe_token(sprint_id, "sprint-id")


def resolve_fingerprint_path(
    root: Path, workspace_root: Path | None, value: str, label: str = "input"
) -> Path:
    """Resolve project paths and explicit shared-workbench URI paths.

    Fingerprint manifests use ``workspace://`` to distinguish read-only
    workbench assets from project-local paths.  The URI must be resolved
    against the workbench root, while the original URI is retained in the
    manifest for provenance.
    """

    if value.startswith("workspace://"):
        if workspace_root is None:
            raise ValueError(f"{label} requires workspace-root for URI: {value}")
        relative = value[len("workspace://") :].lstrip("/")
        if not relative or relative.startswith("../") or relative == "..":
            raise ValueError(f"invalid workspace URI: {value}")
        return workspace_path(workspace_root.resolve(), relative, label)
    return workspace_path(root, value, label)


def fingerprint_files(
    root: Path, paths: list[str], workspace_root: Path | None = None
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for value in paths:
        path = resolve_fingerprint_path(root, workspace_root, value, "input")
        if path.is_file():
            kind = "file"
            digest = sha256(path)
        elif path.is_dir():
            kind = "directory"
            digest_source = "\n".join(
                f"{relative_path(path, item)}:{sha256(item)}"
                for item in sorted(path.rglob("*"))
                if item.is_file()
            )
            digest = hashlib.sha256(digest_source.encode("utf-8")).hexdigest()
        else:
            kind = "missing"
            digest = None
        records.append({
            "path": value if value.startswith("workspace://") else relative_path(root, path),
            "kind": kind,
            "exists": kind != "missing",
            "sha256": digest,
        })
    return records


def protected_sprint_snapshot(root: Path, problem: str) -> list[dict[str, Any]]:
    paths = [item.format(problem=problem).rstrip("/") for item in SPRINT_ROOT_ONLY_PATHS]
    return fingerprint_files(root, paths)


def sprint_expected_outputs(write_directory: str, task_id: str) -> list[str]:
    names = Q1_SPRINT_OUTPUTS.get(task_id, ("handoff.json",))
    return [f"{write_directory}/{name}" for name in names]


def prepare_sprint(
    root: Path,
    agent_mode: str,
    max_agents: int,
    problem: str | None = None,
    workspace_root: Path | None = None,
    question: str | None = None,
    sprint_profile: str | None = None,
    input_sprint_id: str | None = None,
) -> dict[str, Any]:
    if agent_mode != "parallel":
        raise ValueError("multi-agent sprint requires explicit agent-mode parallel opt-in")
    if max_agents < 1 or max_agents > 3:
        raise ValueError("max-agents must be between 1 and 3")

    contest = contest_config(root)
    configured_problem = str(problem or contest.get("problem") or "TBD")
    if configured_problem.upper() == "TBD":
        raise ValueError("formal sprint preparation is forbidden while the problem is TBD")
    configured_problem = safe_token(configured_problem, "problem")
    state_path = root / "state" / "decision_log.json"
    if not state_path.is_file():
        raise ValueError("formal sprint preparation requires initialized mathmodel-skill state")
    state = load_json(state_path)
    if str(state.get("problem") or "") != configured_problem:
        raise ValueError("competition state and sprint problem do not match")

    profile = str(sprint_profile or "default").lower()
    if profile not in {"default", "q1-solve", "q1-compose"}:
        raise ValueError("unsupported sprint profile; use default, q1-solve, or q1-compose")
    requested_question = str(question or "").upper()
    if profile in {"q1-solve", "q1-compose"} and requested_question != "Q1":
        raise ValueError(f"{profile} requires -Question Q1")
    if profile == "q1-solve" and max_agents < 2:
        raise ValueError("q1-solve requires at least two worker agents")
    if profile == "q1-compose" and max_agents < 2:
        raise ValueError("q1-compose requires at least two worker agents")
    if profile == "q1-compose" and not input_sprint_id:
        raise ValueError("q1-compose requires -InputSprintId from q1-solve")

    questions = question_paths(root, configured_problem)
    if not questions:
        raise ValueError("formal sprint preparation requires real question manifests")
    question_lookup = {str(load_yaml(path).get("question_id") or path.parent.name).upper(): path for path in questions}
    if requested_question and requested_question not in question_lookup:
        raise ValueError(f"question manifest does not exist: {requested_question}")
    selected = [question_lookup[requested_question]] if requested_question else questions[:max_agents]
    created_at = datetime.now(UTC)
    sprint_id = created_at.strftime("sprint-%Y%m%dT%H%M%S%fZ")
    folder = sprint_path(root, sprint_id)
    if folder.exists():
        raise FileExistsError(f"sprint already exists: {sprint_id}")

    deadline = str(contest.get("deadline") or "")
    task_files: list[str] = []
    write_directories: list[str] = []
    task_template_path = shared_asset(root, workspace_root, "templates/workflow/sprint_task.json")
    task_template = load_json(task_template_path) if task_template_path.is_file() else {}
    if profile == "q1-compose":
        input_folder, input_manifest, input_tasks = load_sprint(root, str(input_sprint_id))
        if input_manifest.get("profile") != "q1-solve":
            raise ValueError("q1-compose input sprint must be a q1-solve sprint")
        if input_manifest.get("status") != "MERGED":
            raise ValueError("q1-compose requires an input sprint with status MERGED")
        if str(input_manifest.get("problem")) != configured_problem:
            raise ValueError("input sprint problem does not match current project")
        input_merge = input_folder / "merge_report.json"
        if not input_merge.is_file():
            raise ValueError("q1-compose input sprint is missing merge_report.json")
        compose_source_paths = [relative_path(root, input_folder / "merged"), relative_path(root, input_merge)]
    else:
        compose_source_paths = []

    task_specs: list[tuple[str, str]]
    if profile == "q1-solve":
        task_specs = [("forecast-q1", "solver"), ("scheduling-q1", "solver")]
    elif profile == "q1-compose":
        task_specs = [("writer-q1", "writer"), ("reviewer-q1", "reviewer")]
    else:
        task_specs = [(f"solver-{str(load_yaml(path).get('question_id') or path.parent.name).lower()}", "solver") for path in selected]

    for index, (task_id, role) in enumerate(task_specs):
        question_path = selected[0] if profile in {"q1-solve", "q1-compose"} else selected[index]
        question = load_yaml(question_path)
        question_id = safe_token(str(question.get("question_id") or question_path.parent.name), "question")
        source_problem = str(question.get("source_problem") or "")
        input_paths = [relative_path(root, question_path), "contest.yaml", "config/workflow.yaml"]
        if source_problem:
            source = workspace_path(root, source_problem, "source-problem")
            if source.is_file():
                input_paths.append(relative_path(root, source))
        if profile == "q1-solve":
            input_paths.extend([
                "problems/C/data/workload_trace.xlsx",
                "problems/C/data/GPU_information.xlsx",
                "problems/C/data/network_latency.xlsx",
                "problems/C/data/region_time_data.xlsx",
                "problems/C/data/power_mapping.xlsx",
            ])
        if profile == "q1-compose":
            input_paths.extend(compose_source_paths)
        input_paths = list(dict.fromkeys(input_paths))
        write_directory = f"sprints/{sprint_id}/staging/{task_id}"
        task = deepcopy(task_template)
        task.update({
            "schema_version": 1,
            "sprint_id": sprint_id,
            "task_id": task_id,
            "role": role,
            "problem": configured_problem,
            "question": question_id,
            "status": "PENDING",
            "attempt": 1,
            "max_attempts": 2,
            "dependencies": [str(input_sprint_id)] if profile == "q1-compose" else [],
            "input_hashes": fingerprint_files(root, input_paths),
            "allowed_read_paths": input_paths + ["src", "experiments", "corpus", "workspace://src", "workspace://corpus"],
            "write_directory": write_directory,
            "expected_outputs": sprint_expected_outputs(write_directory, task_id),
            "target_gate": "G3" if profile == "q1-solve" else "G5",
            "deadline_utc": deadline,
            "root_only_paths": [item.format(problem=configured_problem) for item in SPRINT_ROOT_ONLY_PATHS],
            "sprint_profile": profile,
            "input_sprint_id": input_sprint_id,
        })
        task_path = folder / "tasks" / f"{task_id}.json"
        dump_json(task_path, task)
        workspace_path(root, write_directory, "write-directory").mkdir(parents=True, exist_ok=True)
        task_files.append(relative_path(root, task_path))
        write_directories.append(write_directory)

    if len(write_directories) != len(set(write_directories)):
        raise ValueError("sprint tasks must have unique write directories")
    sprint_template_path = shared_asset(root, workspace_root, "templates/workflow/sprint.json")
    manifest = load_json(sprint_template_path) if sprint_template_path.is_file() else {}
    manifest.update({
        "schema_version": 1,
        "sprint_id": sprint_id,
        "problem": configured_problem,
        "agent_mode": "parallel",
        "max_agents": max_agents,
        "status": "PREPARED",
        "created_at_utc": created_at.isoformat(),
        "deadline_utc": deadline,
        "state_input": {"path": relative_path(root, state_path), "sha256": sha256(state_path)},
        "protected_snapshot": protected_sprint_snapshot(root, configured_problem),
        "task_files": task_files,
        "root_only_paths": [item.format(problem=configured_problem) for item in SPRINT_ROOT_ONLY_PATHS],
        "profile": profile,
        "question": requested_question or None,
        "input_sprint_id": input_sprint_id,
    })
    dump_json(folder / "sprint.json", manifest)
    return {
        "schema_version": 1,
        "status": "PREPARED",
        "passed": True,
        "sprint_id": sprint_id,
        "problem": configured_problem,
        "task_count": len(task_files),
        "task_files": task_files,
    }


def load_sprint(root: Path, sprint_id: str) -> tuple[Path, dict[str, Any], list[tuple[Path, dict[str, Any]]]]:
    folder = sprint_path(root, sprint_id)
    manifest_path = folder / "sprint.json"
    if not manifest_path.is_file():
        raise FileNotFoundError(f"sprint manifest does not exist: {manifest_path}")
    manifest = load_json(manifest_path)
    if manifest.get("sprint_id") != sprint_id:
        raise ValueError("sprint id does not match its manifest")
    tasks: list[tuple[Path, dict[str, Any]]] = []
    for value in manifest.get("task_files", []):
        path = workspace_path(root, str(value), "task-package")
        valid_parent = path_is_within(path, folder / "tasks") or path_is_within(path, folder / "retry")
        if not path.is_file() or not valid_parent:
            raise ValueError(f"invalid sprint task package: {value}")
        tasks.append((path, load_json(path)))
    if not tasks:
        raise ValueError("sprint has no task packages")
    write_dirs = [str(task.get("write_directory") or "") for _, task in tasks]
    if len(write_dirs) != len(set(write_dirs)):
        raise ValueError("sprint task write directories are not unique")
    return folder, manifest, tasks


def changed_fingerprints(
    root: Path, expected: list[dict[str, Any]], workspace_root: Path | None = None
) -> list[str]:
    changed: list[str] = []
    for item in expected:
        value = str(item.get("path") or "")
        current = fingerprint_files(root, [value], workspace_root)[0]
        if current.get("kind") != item.get("kind") or current.get("sha256") != item.get("sha256"):
            changed.append(str(item.get("path")))
    return changed


def inspect_sprint_task(
    root: Path,
    sprint_folder: Path,
    task: dict[str, Any],
    workspace_root: Path | None = None,
) -> dict[str, Any]:
    task_id = str(task.get("task_id") or "")
    issues: list[dict[str, str]] = []
    stale_inputs = changed_fingerprints(root, list(task.get("input_hashes") or []), workspace_root)
    for path in stale_inputs:
        issues.append({"kind": "stale_input", "path": path})

    write_value = str(task.get("write_directory") or "")
    try:
        write_directory = workspace_path(root, write_value, "write-directory")
        expected_parent = sprint_folder
        relative_parts = write_directory.relative_to(expected_parent).parts if path_is_within(write_directory, expected_parent) else ()
        valid_bucket = bool(relative_parts) and relative_parts[0] in {"staging", "retry-staging"}
        if not valid_bucket or protected_sprint_target(root, str(task.get("problem")), write_value):
            issues.append({"kind": "invalid_write_scope", "path": write_value})
    except ValueError:
        write_directory = sprint_folder
        issues.append({"kind": "invalid_write_scope", "path": write_value})

    handoff_path = write_directory / "handoff.json"
    handoff: dict[str, Any] | None = None
    if not handoff_path.is_file():
        issues.append({"kind": "missing_handoff", "path": relative_path(root, handoff_path)})
    else:
        try:
            handoff = load_json(handoff_path)
        except Exception as exc:
            issues.append({"kind": "invalid_handoff", "path": str(exc)})

    expected_paths: list[Path] = []
    for value in task.get("expected_outputs", []):
        try:
            target = workspace_path(root, str(value), "expected-output")
            if not path_is_within(target, write_directory):
                issues.append({"kind": "expected_output_scope", "path": str(value)})
            else:
                expected_paths.append(target)
                if not target.is_file():
                    issues.append({"kind": "missing_expected_output", "path": str(value)})
        except ValueError:
            issues.append({"kind": "expected_output_scope", "path": str(value)})

    declared_status = "MISSING"
    accepted_artifacts: list[dict[str, Any]] = []
    if handoff is not None:
        declared_status = str(handoff.get("status") or "").upper()
        if handoff.get("sprint_id") != task.get("sprint_id"):
            issues.append({"kind": "sprint_id_mismatch", "path": str(handoff.get("sprint_id"))})
        if handoff.get("task_id") != task_id:
            issues.append({"kind": "task_id_mismatch", "path": str(handoff.get("task_id"))})
        if int(handoff.get("attempt") or 0) != int(task.get("attempt") or 0):
            issues.append({"kind": "attempt_mismatch", "path": str(handoff.get("attempt"))})
        if declared_status not in SPRINT_STATUS_VALUES:
            issues.append({"kind": "invalid_status", "path": declared_status})
        if handoff.get("input_hashes") != task.get("input_hashes"):
            issues.append({"kind": "input_snapshot_mismatch", "path": task_id})
        written_paths = [str(item) for item in handoff.get("written_paths", [])]
        for value in written_paths:
            try:
                target = workspace_path(root, value, "written-path")
                if not path_is_within(target, write_directory) or protected_sprint_target(root, str(task.get("problem")), value):
                    issues.append({"kind": "scope_violation", "path": value})
            except ValueError:
                issues.append({"kind": "scope_violation", "path": value})
        declared_artifacts = handoff.get("artifacts", [])
        if declared_status == "SUCCESS" and not declared_artifacts:
            issues.append({"kind": "missing_artifacts", "path": task_id})
        for artifact in declared_artifacts:
            if not isinstance(artifact, dict) or not artifact.get("path"):
                issues.append({"kind": "invalid_artifact", "path": task_id})
                continue
            value = str(artifact["path"])
            try:
                target = workspace_path(root, value, "artifact")
                in_scope = path_is_within(target, write_directory)
                valid_hash = target.is_file() and artifact.get("sha256") == sha256(target)
                if not in_scope:
                    issues.append({"kind": "scope_violation", "path": value})
                elif not valid_hash:
                    issues.append({"kind": "artifact_hash_mismatch", "path": value})
                else:
                    accepted_artifacts.append({"path": value, "sha256": artifact["sha256"]})
            except ValueError:
                issues.append({"kind": "scope_violation", "path": value})
        gate_result = handoff.get("gate_result")
        if declared_status == "SUCCESS":
            valid_gate = isinstance(gate_result, dict) and gate_result.get("gate") == task.get("target_gate") and gate_result.get("passed") is True
            if not valid_gate:
                issues.append({"kind": "gate_not_passed", "path": str(task.get("target_gate"))})
        artifact_paths = {str(item.get("path")) for item in declared_artifacts if isinstance(item, dict)}
        if not artifact_paths.issubset(set(written_paths)):
            issues.append({"kind": "artifact_not_declared_written", "path": task_id})
        for target in expected_paths:
            value = relative_path(root, target)
            if target != handoff_path and (value not in artifact_paths or value not in written_paths):
                issues.append({"kind": "expected_output_not_declared", "path": value})

    blocking = any(item["kind"] in {"stale_input", "invalid_write_scope", "scope_violation", "input_snapshot_mismatch"} for item in issues)
    if not issues and declared_status == "SUCCESS":
        disposition = "ACCEPTED"
    elif blocking:
        disposition = "BLOCKED"
    else:
        disposition = "RETRY_REQUIRED"
    return {
        "task_id": task_id,
        "declared_status": declared_status,
        "disposition": disposition,
        "attempt": int(task.get("attempt") or 0),
        "issues": issues,
        "artifacts": accepted_artifacts,
        "handoff": relative_path(root, handoff_path),
    }


def check_sprint(
    root: Path,
    sprint_id: str,
    write: bool = True,
    workspace_root: Path | None = None,
) -> dict[str, Any]:
    folder, manifest, tasks = load_sprint(root, sprint_id)
    protected_changes = changed_fingerprints(root, list(manifest.get("protected_snapshot") or []), workspace_root)
    reports = [inspect_sprint_task(root, folder, task, workspace_root) for _, task in tasks]
    if protected_changes:
        for report in reports:
            report["issues"].append({"kind": "root_owned_path_changed", "path": ", ".join(protected_changes)})
            if report["disposition"] == "ACCEPTED":
                report["disposition"] = "BLOCKED"
    accepted = [item["task_id"] for item in reports if item["disposition"] == "ACCEPTED"]
    retry = [item["task_id"] for item in reports if item["disposition"] == "RETRY_REQUIRED"]
    blocked = [item["task_id"] for item in reports if item["disposition"] == "BLOCKED"]
    if len(accepted) == len(reports):
        status_value = "READY_TO_MERGE"
    elif accepted:
        status_value = "PARTIAL"
    elif blocked:
        status_value = "BLOCKED"
    else:
        status_value = "RETRY_REQUIRED"
    result = {
        "schema_version": 1,
        "sprint_id": sprint_id,
        "status": status_value,
        "passed": status_value == "READY_TO_MERGE",
        "accepted_tasks": accepted,
        "retry_tasks": retry,
        "blocked_tasks": blocked,
        "protected_changes": protected_changes,
        "tasks": reports,
        "generated_at_utc": datetime.now(UTC).isoformat(),
    }
    if write:
        dump_json(folder / "check_report.json", result)
    return result


def create_retry_package(
    root: Path,
    folder: Path,
    task: dict[str, Any],
    workspace_root: Path | None = None,
) -> str | None:
    attempt = int(task.get("attempt") or 0) + 1
    if attempt > int(task.get("max_attempts") or 1):
        return None
    task_id = str(task["task_id"])
    retry = deepcopy(task)
    retry["attempt"] = attempt
    retry["status"] = "PENDING"
    retry["dependencies"] = list(dict.fromkeys(list(retry.get("dependencies") or []) + [task_id]))
    old_write_directory = workspace_path(root, str(task["write_directory"]), "write-directory")
    retry["write_directory"] = f"sprints/{folder.name}/retry-staging/{task_id}-attempt-{attempt}"
    new_write_directory = workspace_path(root, retry["write_directory"], "retry-write-directory")
    rebased_outputs: list[str] = []
    for value in task.get("expected_outputs", []):
        target = workspace_path(root, str(value), "expected-output")
        relative = target.relative_to(old_write_directory)
        rebased_outputs.append(relative_path(root, new_write_directory / relative))
    retry["expected_outputs"] = rebased_outputs
    retry["input_hashes"] = fingerprint_files(
        root,
        [str(item["path"]) for item in retry.get("input_hashes", [])],
        workspace_root,
    )
    path = folder / "retry" / f"{task_id}-attempt-{attempt}.json"
    workspace_path(root, retry["write_directory"], "retry-write-directory").mkdir(parents=True, exist_ok=True)
    dump_json(path, retry)
    return relative_path(root, path)


def merge_sprint(root: Path, sprint_id: str, workspace_root: Path | None = None) -> dict[str, Any]:
    folder, manifest, tasks = load_sprint(root, sprint_id)
    check = check_sprint(root, sprint_id, write=True, workspace_root=workspace_root)
    reports = {item["task_id"]: item for item in check["tasks"]}
    archived: list[dict[str, Any]] = []
    retry_packages: list[str] = []
    active_task_files: list[str] = []
    for task_path, task in tasks:
        report = reports[str(task["task_id"])]
        if report["disposition"] == "ACCEPTED":
            write_directory = workspace_path(root, str(task["write_directory"]), "write-directory")
            for artifact in report["artifacts"]:
                source = workspace_path(root, str(artifact["path"]), "artifact")
                relative = source.relative_to(write_directory)
                destination = folder / "merged" / str(task["task_id"]) / relative
                destination.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(source, destination)
                if sha256(destination) != artifact["sha256"]:
                    raise RuntimeError(f"merged artifact hash changed: {artifact['path']}")
                archived.append({
                    "task_id": task["task_id"],
                    "source": artifact["path"],
                    "archive": relative_path(root, destination),
                    "sha256": artifact["sha256"],
                })
            active_task_files.append(relative_path(root, task_path))
        elif report["disposition"] == "RETRY_REQUIRED":
            retry_path = create_retry_package(root, folder, task, workspace_root)
            if retry_path:
                retry_packages.append(retry_path)
                active_task_files.append(retry_path)
            else:
                active_task_files.append(relative_path(root, task_path))
        else:
            active_task_files.append(relative_path(root, task_path))

    protected_changes = changed_fingerprints(root, list(manifest.get("protected_snapshot") or []), workspace_root)
    passed = len(check["accepted_tasks"]) == len(tasks) and not protected_changes
    result = {
        "schema_version": 1,
        "sprint_id": sprint_id,
        "status": "MERGED" if passed else ("PARTIAL" if archived else "BLOCKED"),
        "passed": passed,
        "accepted_tasks": check["accepted_tasks"],
        "archived_artifacts": archived,
        "retry_packages": retry_packages,
        "blocked_tasks": check["blocked_tasks"],
        "protected_changes": protected_changes,
        "formal_state_modified": False,
        "generated_at_utc": datetime.now(UTC).isoformat(),
    }
    dump_json(folder / "merge_report.json", result)
    manifest["status"] = result["status"]
    manifest["task_files"] = active_task_files
    manifest["last_merge_report"] = relative_path(root, folder / "merge_report.json")
    dump_json(folder / "sprint.json", manifest)
    return result


def preflight(root: Path, workspace_root: Path | None = None) -> dict[str, Any]:
    required = [
        "contest.yaml",
        "config/workflow.yaml",
        "skills.lock.yaml",
        "templates/figures/figure_contract_v2.schema.json",
        "templates/figures/figure_contract_v2.template.yaml",
    ]
    project = load_yaml(root / "project.yaml") if (root / "project.yaml").is_file() else {}
    if int(project.get("workflow_contract_version", 0) or 0) >= 7:
        required.extend((
            "config/prompt_policy.yaml",
            "config/schemas/prompt_policy.schema.json",
            "config/schemas/prompt_packet.schema.json",
            "config/schemas/prompt_receipt.schema.json",
            "templates/prompts/paper/cumcm-2026.yaml",
        ))
        required.extend(f"templates/prompts/stages/{stage}.yaml" for stage in ("P0", "P1", "P2", "P3a", "P3b", "P4", "P5", "P6"))
        required.extend(f"templates/prompts/roles/{role}.yaml" for role in ("orchestrator", "solver", "literature", "visualization", "paper", "studio_release", "reviewer"))
    if _project_requires_literature_handoff(root):
        required.extend((
            "config/schemas/literature_search_plan.schema.json",
            "config/schemas/literature_search_receipt.schema.json",
            "config/schemas/academic_reference_card.schema.json",
            "config/schemas/model_evidence_brief.schema.json",
            "skill_staging/literature-guided-modeling/SKILL.md",
        ))
    checks = [{"name": item, "passed": shared_asset(root, workspace_root, item).is_file()} for item in required]
    contest = contest_config(root)
    state_exists = (root / "state" / "decision_log.json").exists()
    if str(contest.get("problem", "TBD")).upper() == "TBD":
        checks.append({"name": "precontest_state_absent", "passed": not state_exists})
        checks.append({"name": "precontest_formal_figure_contract_absent", "passed": not (root / "paper" / "figure_contracts.yaml").exists()})
    else:
        checks.append({"name": "formal_figure_contract_present", "passed": (root / "paper" / "figure_contracts.yaml").is_file()})
    report = {"schema_version": 1, "passed": all(item["passed"] for item in checks), "checks": checks, "state_exists": state_exists}
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Local bridge for the mathematics-modeling competition workflow.")
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parents[2])
    parser.add_argument("--workspace-root", type=Path)
    sub = parser.add_subparsers(dest="action", required=True)
    init = sub.add_parser("initialize")
    init.add_argument("--problem", required=True)
    init.add_argument("--problem-file", required=True, type=Path)
    check = sub.add_parser("validate")
    check.add_argument("--problem", required=True)
    check.add_argument("--gate", required=True)
    check.add_argument("--question")
    check.add_argument("--strict", action="store_true", help="require strict question schema v2/v3 and project handoffs")
    lock = sub.add_parser("freeze")
    lock.add_argument("--problem", required=True)
    lock.add_argument("--question", required=True)
    lock.add_argument("--decision-id", required=True)
    prepare = sub.add_parser("prepare-sprint")
    prepare.add_argument("--agent-mode", required=True)
    prepare.add_argument("--max-agents", required=True, type=int)
    prepare.add_argument("--problem")
    prepare.add_argument("--question")
    prepare.add_argument("--sprint-profile", choices=["default", "q1-solve", "q1-compose"], default="default")
    prepare.add_argument("--input-sprint-id")
    sprint_check = sub.add_parser("check-sprint")
    sprint_check.add_argument("--sprint-id", required=True)
    sprint_merge = sub.add_parser("merge-sprint")
    sprint_merge.add_argument("--sprint-id", required=True)
    sub.add_parser("status")
    sub.add_parser("preflight")
    resolve = sub.add_parser("resolve-run-config")
    resolve.add_argument("--config", required=True, type=Path)
    record = sub.add_parser("record-run")
    record.add_argument("--config", required=True, type=Path)
    record.add_argument("--command-json", required=True)
    record.add_argument("--environment-json", required=True)
    record.add_argument("--started-at", required=True)
    record.add_argument("--duration", required=True, type=float)
    record.add_argument("--success", action="store_true")
    quick = sub.add_parser("quickcheck")
    quick.add_argument("--problem", required=True)
    quick.add_argument("--question")
    quick.add_argument("--strict", action="store_true")
    checkpoint_parser = sub.add_parser("checkpoint")
    checkpoint_parser.add_argument("--problem", required=True)
    checkpoint_parser.add_argument("--question")
    checkpoint_parser.add_argument("--strict", action="store_true")
    promote_parser = sub.add_parser("promote")
    promote_parser.add_argument("--problem", required=True)
    promote_parser.add_argument("--question", required=True)
    promote_parser.add_argument("--run-id", required=True)
    paper_evidence_parser = sub.add_parser("paper-evidence")
    paper_evidence_parser.add_argument("--problem", required=True)
    paper_evidence_parser.add_argument("--question", required=True)
    paper_evidence_parser.add_argument("--config", required=True, type=Path)
    paper_evidence_parser.add_argument("--strict", action="store_true")
    literature_plan_parser = sub.add_parser("literature-plan")
    literature_plan_parser.add_argument("--problem", required=True)
    literature_plan_parser.add_argument("--question", required=True)
    literature_plan_parser.add_argument("--config", type=Path)
    literature_search_parser = sub.add_parser("literature-search")
    literature_search_parser.add_argument("--problem", required=True)
    literature_search_parser.add_argument("--question", required=True)
    literature_search_parser.add_argument("--config", required=True, type=Path)
    literature_register_parser = sub.add_parser("literature-register")
    literature_register_parser.add_argument("--problem", required=True)
    literature_register_parser.add_argument("--question", required=True)
    literature_register_parser.add_argument("--config", required=True, type=Path)
    literature_read_parser = sub.add_parser("literature-read")
    literature_read_parser.add_argument("--problem", required=True)
    literature_read_parser.add_argument("--question", required=True)
    literature_read_parser.add_argument("--config", required=True, type=Path)
    literature_synthesize_parser = sub.add_parser("literature-synthesize")
    literature_synthesize_parser.add_argument("--problem", required=True)
    literature_synthesize_parser.add_argument("--question", required=True)
    literature_synthesize_parser.add_argument("--config", required=True, type=Path)
    literature_audit_parser = sub.add_parser("literature-audit")
    literature_audit_parser.add_argument("--problem", required=True)
    literature_audit_parser.add_argument("--question")
    literature_audit_parser.add_argument("--strict", action="store_true")
    figure_data_parser = sub.add_parser("figure-data")
    figure_data_parser.add_argument("--problem", required=True)
    figure_data_parser.add_argument("--question", required=True)
    figure_data_parser.add_argument("--run-id", required=True)
    figure_data_parser.add_argument("--config", required=True, type=Path)
    figure_intent_parser = sub.add_parser("figure-intent")
    figure_intent_parser.add_argument("--problem", required=True)
    figure_intent_parser.add_argument("--question", required=True)
    figure_intent_parser.add_argument("--run-id", required=True)
    figure_intent_parser.add_argument("--config", required=True, type=Path)
    figure_brief_parser = sub.add_parser("figure-brief")
    figure_brief_parser.add_argument("--problem", required=True)
    figure_brief_parser.add_argument("--question", required=True)
    figure_brief_parser.add_argument("--run-id", required=True)
    figure_brief_parser.add_argument("--intent", required=True, type=Path)
    figure_brief_parser.add_argument("--config", required=True, type=Path)
    figure_render_parser = sub.add_parser("figure-render")
    figure_render_parser.add_argument("--problem", required=True)
    figure_render_parser.add_argument("--question", required=True)
    figure_render_parser.add_argument("--run-id", required=True)
    figure_render_parser.add_argument("--brief", required=True, type=Path)
    figure_qa_parser = sub.add_parser("figure-qa")
    figure_qa_parser.add_argument("--problem", required=True)
    figure_qa_parser.add_argument("--question", required=True)
    figure_qa_parser.add_argument("--run-id", required=True)
    figure_qa_parser.add_argument("--brief", required=True, type=Path)
    figure_qa_parser.add_argument("--outputs", type=Path)
    figure_promote_parser = sub.add_parser("figure-promote")
    figure_promote_parser.add_argument("--problem", required=True)
    figure_promote_parser.add_argument("--question", required=True)
    figure_promote_parser.add_argument("--figure-id", required=True)
    figure_promote_parser.add_argument("--brief", required=True, type=Path)
    figure_promote_parser.add_argument("--qa", required=True, type=Path)
    figure_promote_parser.add_argument("--root-authorized", action="store_true", help=argparse.SUPPRESS)
    prompt_parser = sub.add_parser("prompt")
    prompt_parser.add_argument("--project-id", required=True)
    prompt_parser.add_argument("--stage", required=True)
    prompt_parser.add_argument("--role", required=True)
    prompt_parser.add_argument("--question")
    archive_parser = sub.add_parser("archive-work")
    archive_parser.add_argument("--problem", required=True)
    archive_parser.add_argument("--question")
    args = parser.parse_args()
    root = args.root.resolve()
    workspace_root = args.workspace_root.resolve() if args.workspace_root else root
    try:
        if args.action == "initialize":
            result = initialize(root, args.problem, args.problem_file, workspace_root)
        elif args.action == "validate":
            result = validate(root, args.problem, args.gate, args.question, strict=args.strict)
        elif args.action == "freeze":
            result = freeze(root, args.problem, args.question, args.decision_id)
        elif args.action == "prepare-sprint":
            result = prepare_sprint(
                root,
                args.agent_mode,
                args.max_agents,
                args.problem,
                workspace_root,
                args.question,
                args.sprint_profile,
                args.input_sprint_id,
            )
        elif args.action == "check-sprint":
            result = check_sprint(root, args.sprint_id, workspace_root=workspace_root)
        elif args.action == "merge-sprint":
            result = merge_sprint(root, args.sprint_id, workspace_root=workspace_root)
        elif args.action == "status":
            result = status(root)
        elif args.action == "preflight":
            result = preflight(root, workspace_root)
        elif args.action == "resolve-run-config":
            result = resolve_run_config(root, args.config)
        elif args.action == "record-run":
            result = record_run(root, args.config, json.loads(args.command_json), json.loads(args.environment_json), args.started_at, args.duration, args.success)
        elif args.action == "quickcheck":
            result = quickcheck(root, args.problem, args.question, args.strict)
        elif args.action == "checkpoint":
            result = checkpoint(root, args.problem, args.question, args.strict)
        elif args.action == "promote":
            result = promote(root, args.problem, args.question, args.run_id)
        elif args.action == "paper-evidence":
            result = paper_evidence(root, args.problem, args.question, args.config, args.strict)
        elif args.action == "literature-plan":
            result = literature_plan(root, args.problem, args.question, args.config)
        elif args.action == "literature-search":
            result = literature_search(root, args.problem, args.question, args.config)
        elif args.action == "literature-register":
            result = literature_register(root, args.problem, args.question, args.config)
        elif args.action == "literature-read":
            result = literature_read(root, args.problem, args.question, args.config)
        elif args.action == "literature-synthesize":
            result = literature_synthesize(root, args.problem, args.question, args.config)
        elif args.action == "literature-audit":
            result = literature_audit(root, args.problem, args.question, args.strict)
        elif args.action == "figure-data":
            result = figure_data(root, args.problem, args.question, args.run_id, args.config)
        elif args.action == "figure-intent":
            result = figure_intent(root, args.problem, args.question, args.run_id, args.config)
        elif args.action == "figure-brief":
            result = figure_brief(root, args.problem, args.question, args.run_id, args.intent, args.config)
        elif args.action == "figure-render":
            result = figure_render(root, args.problem, args.question, args.run_id, args.brief)
        elif args.action == "figure-qa":
            result = figure_qa(root, args.problem, args.question, args.run_id, args.brief, args.outputs)
        elif args.action == "figure-promote":
            if not args.root_authorized:
                raise ValueError(
                    "figure-promote must be routed through scripts/workflow.ps1; "
                    "the approved decision-log reference is the authoritative root-agent control"
                )
            result = figure_promote(root, args.problem, args.question, args.figure_id, args.brief, args.qa)
        elif args.action == "prompt":
            result = prompt(root, args.project_id, args.stage, args.role, args.question, workspace_root)
        else:
            result = archive_work(root, args.problem, args.question)
    except ReopenRequiredError as exc:
        print(json.dumps({"status": "REOPEN_REQUIRED", "error": str(exc)}, ensure_ascii=False, indent=2))
        return 2
    except Exception as exc:
        print(json.dumps({"status": "ERROR", "error": str(exc)}, ensure_ascii=False, indent=2))
        return 1
    print(json.dumps(result, ensure_ascii=False, indent=2))
    if result.get("status") == "REOPEN_REQUIRED":
        return 2
    return 0 if result.get("passed", True) else 1


if __name__ == "__main__":
    sys.exit(main())
